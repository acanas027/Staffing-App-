"""
Order vs Inventory Matcher — single-file Streamlit app.

Upload a sample-orders workbook and a QPORT inventory report.
The app matches every ordered item to its available quantity and locations,
one row per item, with every stocking location and its quantity listed together.

Output workbook (matches the "updated samples report" layout):
    Regular Orders  — items with a numeric item no
    Samples Orders  — items whose item no starts with a letter (S12150, ...)
    Summary         — headline counts for the day

Run with:
    pip install -r requirements.txt
    streamlit run app.py
"""

import io
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Order vs Inventory Matcher", layout="wide")

# --- QPORT inventory report: 0-based column positions -----------------------
COL_LOC = [0, 1, 2, 3, 4, 5]
COL_SKU_H = 7
COL_SKU_I = 8
COL_QTY = 9
COL_PREV = [10, 11, 12, 13, 14, 15]
COL_DATE2 = 24
COL_TIME2 = 25

# Set to True to keep the quantity number and the unit (CA/EA) in two separate
# columns. False matches the updated samples report, which shows "1 CA" in one
# cell. Either way the CA/EA totals on the Summary sheet are unaffected.
SPLIT_QTY_COLUMNS = False

# Carried through the pipeline but never written to the order sheets.
HIDDEN_COLS = {"_OrderSortKey", "Order Date"}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="2F5496")
LABEL_FILL = PatternFill("solid", start_color="D9E2F3")
ZERO_FILL = PatternFill("solid", start_color="FFF2CC")

COL_WIDTHS = {
    "Order #": 13,
    "Item No": 13,
    "Item Name": 40,
    "Order Qty": 10,
    "U/M": 7,
    "Current Location": 30,
    "Previous Location": 16,
    "Quantity Available": 18,
}


def _clean(v):
    if pd.isna(v):
        return ""
    return str(v).strip()


def normalize_order_no(order_no_numeric, raw_text):
    """
    Display text for Order #.
    Uses the already-parsed numeric value when available (e.g. 3700007249.0 -> "3700007249"),
    otherwise falls back to the original raw text (trimmed).
    """
    if pd.notna(order_no_numeric):
        return str(int(order_no_numeric))
    return _clean(raw_text)


def parse_dep_date(v):
    """
    'Dep dt' arrives as a packed MMDDYY integer (71626 -> 2026-07-16).
    Returns a datetime, or None if it can't be read.
    """
    if pd.isna(v):
        return None

    try:
        s = str(int(float(v))).zfill(6)
        return datetime.strptime(s, "%m%d%y")
    except (ValueError, TypeError):
        return None


def force_order_sort(df: pd.DataFrame) -> pd.DataFrame:
    """
    Final forced sort.

    Sorts by the numeric "_OrderSortKey" only (computed once, at load time).
    Does NOT sort by Item No.
    Keeps the original row order for ties / rows within the same order.
    Rows with no parseable Order # sort to the end.
    """
    df = df.copy()
    df["_line_order"] = range(len(df))

    df = df.sort_values(
        by=["_OrderSortKey", "_line_order"],
        ascending=[True, True],
        na_position="last",
        kind="mergesort",  # stable sort so ties keep their original relative order
    )

    df = df.drop(columns=["_line_order"]).reset_index(drop=True)
    return df


def build_sku(h, i):
    if pd.isna(h) or pd.isna(i):
        return None

    try:
        h = float(h)
        i = float(i)
    except (ValueError, TypeError):
        return None

    int_part = int(h)
    frac = int(round((h - int_part) * 100))
    suffix = int(round(i))

    return f"{int_part}.{frac:02d}{suffix:03d}"


def load_tkreserve(file) -> pd.DataFrame:
    raw = pd.read_excel(file, sheet_name=0, header=0)

    loc = raw.iloc[:, COL_LOC].apply(
        lambda row: "".join([_clean(v) for v in row if _clean(v) != ""]),
        axis=1
    )

    prev_loc = raw.iloc[:, COL_PREV].apply(
        lambda row: "".join([_clean(v) for v in row if _clean(v) != ""]),
        axis=1
    )

    h = pd.to_numeric(raw.iloc[:, COL_SKU_H], errors="coerce")
    i = pd.to_numeric(raw.iloc[:, COL_SKU_I], errors="coerce")

    sku = [build_sku(a, b) for a, b in zip(h, i)]

    qty = pd.to_numeric(raw.iloc[:, COL_QTY], errors="coerce").fillna(0)

    date2 = pd.to_numeric(raw.iloc[:, COL_DATE2], errors="coerce").fillna(0)
    time2 = pd.to_numeric(raw.iloc[:, COL_TIME2], errors="coerce").fillna(0)

    tx_time = date2 * 1_000_000 + time2

    df = pd.DataFrame({
        "SKU": sku,
        "Location": loc,
        "Quantity": qty,
        "Previous Location": prev_loc,
        "TxTime": tx_time,
    })

    df = df.dropna(subset=["SKU"])

    return df[df["SKU"] != ""]


def is_active_location(loc) -> bool:
    """Active pick locations end in a letter; reserve locations end in a digit."""
    loc = str(loc).strip()
    return bool(loc) and loc[-1].isalpha()


def aggregate_inventory(inv: pd.DataFrame) -> pd.DataFrame:
    def summarize(g):
        loc_qty = g.groupby("Location")["Quantity"].sum()

        # Active locations (ending in a letter) list first.
        # Within each group, and when there are no active locations,
        # the order is unchanged: highest quantity first.
        ordered = sorted(
            ((loc, q) for loc, q in loc_qty.items() if loc),
            key=lambda kv: (0 if is_active_location(kv[0]) else 1, -kv[1]),
        )

        loc_str = "; ".join(f"{loc} ({int(q)})" for loc, q in ordered)

        latest = g.loc[g["TxTime"].idxmax(), "Previous Location"]

        return pd.Series({
            "Quantity Available": g["Quantity"].sum(),
            "Current Location": loc_str,
            "Previous Location": latest,
        })

    return inv.groupby("SKU").apply(summarize, include_groups=False).reset_index()


def clean_qty(qty):
    """Numeric order quantity, as an int when it has no decimal part."""
    if pd.isna(qty):
        return 0

    try:
        qty = float(qty)
    except (ValueError, TypeError):
        return 0

    return int(qty) if qty.is_integer() else qty


def clean_um(um):
    """Unit of measure, upper-cased and trimmed (CA, EA, ...)."""
    return _clean(um).upper()


def load_sample_orders(file) -> pd.DataFrame:
    xl = pd.ExcelFile(file)

    frames = []

    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet, header=0)

        if df.empty or "Item no" not in df.columns:
            continue

        df = df.dropna(subset=["Item no"]).copy()

        if df.empty:
            continue

        if "CO no" in df.columns:
            order_no_raw = df["CO no"]
        else:
            order_no_raw = pd.Series([""] * len(df), index=df.index)

        # Parse the numeric sort key ONCE here. Everything downstream
        # (display text and sort order) is derived from this same value,
        # so they can never disagree.
        order_no_numeric = pd.to_numeric(order_no_raw, errors="coerce")

        # Real order date comes from 'Dep dt'. The sheet name is only a
        # fallback for workbooks that don't carry the column.
        if "Dep dt" in df.columns:
            order_date = [parse_dep_date(v) for v in df["Dep dt"]]
        else:
            order_date = [sheet] * len(df)

        if "Order qty" in df.columns:
            qty_num = pd.to_numeric(df["Order qty"], errors="coerce").fillna(0)
        else:
            qty_num = pd.Series([0] * len(df), index=df.index)

        if "U/M" in df.columns:
            um = df["U/M"]
        else:
            um = pd.Series([""] * len(df), index=df.index)

        if "Name" in df.columns:
            item_name = df["Name"]
        else:
            item_name = pd.Series([""] * len(df), index=df.index)

        frames.append(pd.DataFrame({
            "Order Date": order_date,
            "Order #": [normalize_order_no(n, r) for n, r in zip(order_no_numeric, order_no_raw)],
            "_OrderSortKey": order_no_numeric,
            "Item No": df["Item no"].astype(str).str.strip(),
            "Item Name": item_name,
            "Order Qty": [clean_qty(q) for q in qty_num],
            "U/M": [clean_um(u) for u in um],
        }))

    if not frames:
        return pd.DataFrame(columns=[
            "Order Date",
            "Order #",
            "_OrderSortKey",
            "Item No",
            "Item Name",
            "Order Qty",
            "U/M",
        ])

    orders = pd.concat(frames, ignore_index=True)

    return orders


def is_sample_item(item_no) -> bool:
    """Sample items don't start with a digit (S12150, S56032, ...)."""
    item_no = str(item_no).strip()
    return not item_no[:1].isdigit()


def match_orders_to_inventory(orders: pd.DataFrame, tk_file) -> pd.DataFrame:
    inv_raw = load_tkreserve(tk_file)
    inv = aggregate_inventory(inv_raw)

    result = orders.merge(
        inv,
        how="left",
        left_on="Item No",
        right_on="SKU"
    ).drop(columns=["SKU"])

    result["Quantity Available"] = result["Quantity Available"].fillna(0)
    result["Current Location"] = result["Current Location"].fillna("Not found in inventory")
    result["Previous Location"] = result["Previous Location"].fillna("")

    cols = [
        "Order Date",
        "Order #",
        "_OrderSortKey",
        "Item No",
        "Item Name",
        "Order Qty",
        "U/M",
        "Current Location",
        "Previous Location",
        "Quantity Available",
    ]

    result = result[cols]

    # IMPORTANT:
    # This is the only sort.
    # It groups everything by Order #.
    # It does not sort by Item No.
    result = force_order_sort(result)

    return result


def split_regular_samples(result: pd.DataFrame):
    """Regular orders vs samples orders, each already in Order # order."""
    mask = result["Item No"].apply(is_sample_item)

    regular = result[~mask].reset_index(drop=True)
    samples = result[mask].reset_index(drop=True)

    return regular, samples


def to_display(df: pd.DataFrame) -> pd.DataFrame:
    """Drop the internal columns and shape the quantity for output."""
    out = df.drop(columns=[c for c in HIDDEN_COLS if c in df.columns]).copy()

    out["Quantity Available"] = (
        pd.to_numeric(out["Quantity Available"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    if not SPLIT_QTY_COLUMNS:
        out["Order Qty"] = [
            f"{q} {u}".strip() for q, u in zip(out["Order Qty"], out["U/M"])
        ]
        out = out.drop(columns=["U/M"])

    return out


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def summary_date(result: pd.DataFrame):
    """The single order date, or a joined list if the file spans several."""
    dates = [d for d in result["Order Date"].dropna().unique()]

    if not dates:
        return ""

    if len(dates) == 1:
        d = dates[0]
        return d if isinstance(d, str) else pd.Timestamp(d).to_pydatetime()

    return ", ".join(
        d if isinstance(d, str) else pd.Timestamp(d).strftime("%m/%d/%y")
        for d in sorted(dates, key=str)
    )


def day_volume_flag(order_count: int) -> str:
    """Light / Average / Heavy day flag based on number of orders."""
    if order_count < 10:
        return "Light"
    if order_count <= 20:
        return "Average"
    return "Heavy"


def expected_completion_time(order_count: int) -> str:
    """Expected completion time at 40 minutes per order, shown in hours."""
    hours = (order_count * 40) / 60
    return f"{hours:.1f} hours"


def build_summary(result: pd.DataFrame) -> dict:
    """Headline numbers for the Summary sheet, across regular + samples."""
    df = result.copy()
    df["U/M"] = df["U/M"].astype(str).str.strip().str.upper()
    df["Order Qty"] = pd.to_numeric(df["Order Qty"], errors="coerce").fillna(0)

    um_totals = df.groupby("U/M")["Order Qty"].sum()
    orders_count = int(df["Order #"].nunique())

    return {
        "Date": summary_date(df),
        "Orders": orders_count,
        "Day Flag": day_volume_flag(orders_count),
        "Expected Completion Time": expected_completion_time(orders_count),
        "Distinct Items": int(df["Item No"].nunique()),
        "Total Cases (CA)": int(um_totals.get("CA", 0)),
        "Total Each (EA)": int(um_totals.get("EA", 0)),
        "Short on Stock": int((
            (df["Quantity Available"] > 0)
            & (df["Quantity Available"] < df["Order Qty"])
        ).sum()),
    }


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def write_order_sheet(wb, title: str, df: pd.DataFrame):
    """
    One sheet of order lines, with a blank separator row between orders.
    Lines with nothing on hand are highlighted (samples excluded — they're
    never expected to be in inventory).
    """
    ws = wb.create_sheet(title)

    display = to_display(df)
    headers = list(display.columns)

    ws.append(headers)

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    prev_order = None
    r = 1

    for i, row in display.iterrows():
        order_no = row["Order #"]

        # Blank row whenever a new order starts (never before the first one).
        if prev_order is not None and order_no != prev_order:
            r += 1
            ws.append([])

        r += 1
        ws.append([row[c] for c in headers])

        qty_avail = df.iloc[i]["Quantity Available"]
        item_no = df.iloc[i]["Item No"]

        if qty_avail == 0 and not is_sample_item(item_no):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = ZERO_FILL

        prev_order = order_no

    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(h, 14)

    ws.freeze_panes = "A2"

    return ws


def day_flag_fill(flag: str):
    flag = str(flag).strip().lower()
    if flag == "light":
        return PatternFill("solid", start_color="C6EFCE")
    if flag == "average":
        return PatternFill("solid", start_color="FFF2CC")
    if flag == "heavy":
        return PatternFill("solid", start_color="F4CCCC")
    return None


def build_summary_sheet(wb, summary: dict):
    ws = wb.create_sheet("Summary")

    title = ws.cell(row=1, column=1, value="Samples Orders Summary")
    title.font = Font(bold=True, size=14)

    r = 3
    for label, value in summary.items():
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = LABEL_FILL

        vc = ws.cell(row=r, column=2, value=value)
        vc.font = Font(bold=True, size=12)
        vc.alignment = Alignment(horizontal="center")

        if isinstance(value, datetime):
            vc.number_format = "MM/DD/YYYY"

        if label == "Day Flag":
            fill = day_flag_fill(value)
            if fill is not None:
                vc.fill = fill

        r += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14

    return ws


def build_excel(result: pd.DataFrame) -> io.BytesIO:
    # Force the same order again before exporting.
    result = force_order_sort(result)
    regular, samples = split_regular_samples(result)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we name our own

    write_order_sheet(wb, "Regular Orders", regular)
    write_order_sheet(wb, "Samples Orders", samples)
    build_summary_sheet(wb, build_summary(result))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

def order_count(df: pd.DataFrame) -> int:
    """Distinct orders in a set of lines — not the number of lines."""
    return int(df["Order #"].nunique())


def style_zeros(df: pd.DataFrame):
    """Yellow rows for non-sample items with nothing on hand."""
    display = to_display(df)
    n_cols = len(display.columns)

    styles = []
    for _, row in df.iterrows():
        if row["Quantity Available"] == 0 and not is_sample_item(row["Item No"]):
            styles.append(["background-color: #fff2cc"] * n_cols)
        else:
            styles.append([""] * n_cols)

    return display.style.apply(lambda row: styles[row.name], axis=1)


st.title("Sample Orders to Inventory Matcher")

st.write(
    "Upload the **sample orders** workbook and the **TKRESERVE_QPORT** inventory "
    "report. The app matches every ordered item to its available quantity and "
    "every location it's currently stocked in."
)

col1, col2 = st.columns(2)

with col1:
    orders_file = st.file_uploader(
        "Sample Orders (.xlsx)",
        type=["xlsx"],
        key="orders"
    )

with col2:
    tk_file = st.file_uploader(
        "QPORT Inventory Report (.xlsx)",
        type=["xlsx"],
        key="tk"
    )

if orders_file and tk_file:
    try:
        with st.spinner("Reading and matching..."):
            orders = load_sample_orders(orders_file)

            if orders.empty:
                st.error(
                    "No order rows found. Expected a column named 'Item no' "
                    "in the sample orders sheets."
                )
                st.stop()

            result = match_orders_to_inventory(orders, tk_file)
            regular, samples = split_regular_samples(result)

    except Exception as e:
        st.error(f"Something went wrong while processing the files: {e}")
        st.stop()

    summary = build_summary(result)

    date_val = summary["Date"]
    date_txt = (
        date_val.strftime("%m/%d/%Y")
        if isinstance(date_val, datetime)
        else str(date_val)
    )

    st.subheader("Summary")

    m1, m2, m3 = st.columns(3)
    m1.metric("Date", date_txt)
    m2.metric("Orders", summary["Orders"])
    m3.metric("Distinct items", summary["Distinct Items"])

    m4, m5, m6 = st.columns(3)
    m4.metric("Total Cases (CA)", f"{summary['Total Cases (CA)']:,}")
    m5.metric("Total Each (EA)", f"{summary['Total Each (EA)']:,}")
    m6.metric("Short on stock", summary["Short on Stock"])

    m7, m8 = st.columns(2)
    with m7:
        if summary["Day Flag"] == "Light":
            st.success(f"Day Flag: {summary['Day Flag']}")
        elif summary["Day Flag"] == "Average":
            st.warning(f"Day Flag: {summary['Day Flag']}")
        else:
            st.error(f"Day Flag: {summary['Day Flag']}")
    m8.metric("Expected Completion Time", summary["Expected Completion Time"])

    n_reg = order_count(regular)
    n_smp = order_count(samples)

    st.caption(f"{n_reg} regular orders · {n_smp} samples orders")

    st.divider()

    tab1, tab2 = st.tabs(
        [f"Regular Orders ({n_reg})", f"Samples Orders ({n_smp})"]
    )

    with tab1:
        st.dataframe(style_zeros(regular), use_container_width=True, height=420)

    with tab2:
        st.dataframe(style_zeros(samples), use_container_width=True, height=420)

    excel_buf = build_excel(result)

    st.download_button(
        "Download results as Excel",
        data=excel_buf,
        file_name="updated_samples_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("Upload both files to run the match.")
