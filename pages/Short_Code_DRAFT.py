# -*- coding: utf-8 -*-
"""
SKU Data + qPORT/WMS Matcher

Upload the same two Excel files:
1) Data / orders file
2) Big qPORT / WMS file

The app returns one Excel workbook showing only SKUs from the Data file,
matched to every relevant qPORT/WMS row.
"""

import re
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# Basic cleaners
# ============================================================

def clean_str(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def clean_display_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def h_to_str(value):
    if pd.isna(value):
        return None
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def i_to_str(value):
    if pd.isna(value):
        return "000"
    if isinstance(value, float) and value == int(value):
        return str(int(value)).zfill(3)
    return str(value).strip().zfill(3)


# ============================================================
# SKU normalization
# ============================================================

def normalize_sku(value):
    """Normalize SKU values so both files can match.

    Examples:
    06795.48940 -> 6795.4894
    6795.48940  -> 6795.4894
    6795.4894   -> 6795.4894
    79341.0     -> 79341
    """
    if pd.isna(value):
        return ""

    text = str(value).strip().replace(",", "")
    if text == "" or text.lower() == "nan":
        return ""

    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]

    if "." in text:
        whole, decimal = text.split(".", 1)
        whole = whole.strip()
        decimal = decimal.strip()

        if whole.isdigit():
            whole = str(int(whole))

        decimal = decimal.rstrip("0")
        if decimal:
            return f"{whole}.{decimal}"
        return whole

    return text


def build_full_wms_sku(h_value, i_value):
    """Build full SKU from qPORT/WMS columns H and I."""
    if h_value is None:
        return None

    if "." in h_value:
        whole, decimal = h_value.split(".", 1)
        whole = whole.zfill(5)
        decimal = decimal.ljust(2, "0")[:2]
        return f"{whole}.{decimal}{i_value}"

    return h_value.zfill(5)


# ============================================================
# Location normalization
# ============================================================

def normalize_location_text(value):
    text = clean_str(value).upper().replace(" ", "")
    if text == "" or text == "NAN":
        return ""

    if text in {"BLACKHOL", "BLACKHO", "BLACKH", "BLACK"}:
        return "BLACKHOLE"

    return text


def build_location_from_parts(raw_df, start_col, end_col):
    location_parts = raw_df.iloc[:, start_col:end_col].copy()
    for col in location_parts.columns:
        location_parts[col] = location_parts[col].map(clean_str)
    joined = location_parts.agg("".join, axis=1)
    return joined.map(normalize_location_text)


def parse_location(location):
    """Break a compact location into readable pieces.

    Example:
    RC2A30X8 -> RC2-A-30-X8
    """
    compact = normalize_location_text(location)

    if compact == "":
        return pd.Series({
            "Location Normalized": "",
            "Location Zone": "",
            "Location Aisle": "",
            "Location Bay": "",
            "Location Position": "",
        })

    if compact == "BLACKHOLE":
        return pd.Series({
            "Location Normalized": "BLACKHOLE",
            "Location Zone": "BLACKHOLE",
            "Location Aisle": "",
            "Location Bay": "",
            "Location Position": "",
        })

    zone = compact[:3] if len(compact) >= 3 else compact
    remainder = compact[3:] if len(compact) > 3 else ""

    aisle = ""
    if remainder and remainder[0].isalpha():
        aisle = remainder[0]
        remainder = remainder[1:]

    bay = ""
    position = ""
    match = re.match(r"^(\d+)(.*)$", remainder)
    if match:
        bay = match.group(1)
        position = match.group(2)
    else:
        position = remainder

    pieces = [part for part in [zone, aisle, bay, position] if part]
    normalized = "-".join(pieces) if pieces else compact

    return pd.Series({
        "Location Normalized": normalized,
        "Location Zone": zone,
        "Location Aisle": aisle,
        "Location Bay": bay,
        "Location Position": position,
    })


def add_location_fields(df, source_column, prefix):
    parsed = df[source_column].apply(parse_location)
    parsed = parsed.rename(columns={
        "Location Normalized": f"{prefix} Location Normalized",
        "Location Zone": f"{prefix} Location Zone",
        "Location Aisle": f"{prefix} Location Aisle",
        "Location Bay": f"{prefix} Location Bay",
        "Location Position": f"{prefix} Location Position",
    })
    return pd.concat([df, parsed], axis=1)


# ============================================================
# Load files
# ============================================================

def load_data_file(file, exclude_s_items=True, decimal_skus_only=True):
    df = pd.read_excel(file, sheet_name=0)
    df.columns = [str(col).strip() for col in df.columns]

    required_columns = ["Item", "Target Date", "Delivery Date", "Quantity Ordered"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise KeyError(
            "Missing expected column(s) in the Data file: " + ", ".join(missing_columns)
        )

    df = df[df["Item"].notna()].copy()
    df["Data Item"] = df["Item"].map(clean_display_value)
    df["Normalized SKU"] = df["Data Item"].map(normalize_sku)

    if exclude_s_items:
        df = df[~df["Data Item"].str.upper().str.startswith("S")].copy()

    if decimal_skus_only:
        df = df[df["Normalized SKU"].str.contains(r"\.", regex=True, na=False)].copy()

    df = df[df["Normalized SKU"] != ""].copy()

    df["Target Date"] = pd.to_datetime(df["Target Date"], errors="coerce")
    df["Delivery Date"] = pd.to_datetime(df["Delivery Date"], errors="coerce")
    df["Quantity Ordered"] = pd.to_numeric(
        df["Quantity Ordered"], errors="coerce"
    ).fillna(0)

    if "Order" not in df.columns:
        df["Order"] = ""
    else:
        df["Order"] = df["Order"].map(clean_display_value)

    if "Order Status" not in df.columns:
        df["Order Status"] = ""
    else:
        df["Order Status"] = df["Order Status"].map(clean_display_value)

    if "Customer" in df.columns:
        df["Customer"] = df["Customer"].map(clean_display_value)
    else:
        df["Customer"] = ""

    df = df.reset_index(drop=True)
    df["Data Row #"] = df.index + 2
    return df


def load_wms_file(file):
    raw = pd.read_excel(file, sheet_name=0, header=None, skiprows=3)

    if raw.shape[1] < 28:
        raise ValueError(
            "The qPORT/WMS file does not have the expected column layout. "
            "Expected at least 28 columns."
        )

    current_location = build_location_from_parts(raw, 0, 6)
    previous_location = build_location_from_parts(raw, 10, 16)

    lpn = raw.iloc[:, 6].map(clean_display_value)
    sku_h = raw.iloc[:, 7].apply(h_to_str)
    sku_i = raw.iloc[:, 8].apply(i_to_str)
    full_sku = [build_full_wms_sku(h, i) for h, i in zip(sku_h, sku_i)]
    quantity = pd.to_numeric(raw.iloc[:, 9], errors="coerce").fillna(0)
    consumer_priority_raw = raw.iloc[:, 27]

    wms = pd.DataFrame({
        "WMS Row #": raw.index + 4,
        "WMS SKU Number": full_sku,
        "Normalized SKU": [normalize_sku(sku) for sku in full_sku],
        "LPN #": lpn,
        "WMS Quantity": quantity,
        "Current Location Compact": current_location,
        "Previous Location Compact": previous_location,
        "Consumer Priority Date": pd.to_datetime(
            consumer_priority_raw.astype(str), format="%Y%m%d", errors="coerce"
        ),
    })

    wms = wms[wms["Normalized SKU"] != ""].copy()
    wms = add_location_fields(wms, "Current Location Compact", "Current")
    wms = add_location_fields(wms, "Previous Location Compact", "Previous")
    wms = wms.reset_index(drop=True)
    return wms


# ============================================================
# Summaries
# ============================================================

def unique_join(values, limit=50):
    cleaned = []
    for value in values:
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text == "" or text.lower() in {"nan", "nat"}:
            continue
        cleaned.append(text)

    unique_values = sorted(set(cleaned))
    if len(unique_values) > limit:
        shown = unique_values[:limit]
        return ", ".join(shown) + f" ... (+{len(unique_values) - limit} more)"
    return ", ".join(unique_values)


def unique_date_join(values, limit=30):
    converted = pd.to_datetime(pd.Series(values), errors="coerce").dropna()
    if converted.empty:
        return ""
    dates = converted.dt.date.astype(str)
    return unique_join(dates, limit=limit)


def build_data_summary(data_df):
    rows = []

    for sku, group in data_df.groupby("Normalized SKU", dropna=False):
        rows.append({
            "Normalized SKU": sku,
            "Data Item Values": unique_join(group["Data Item"]),
            "Data Lines": len(group),
            "Customer Values": unique_join(group["Customer"]),
            "Order Count": group["Order"].replace("", pd.NA).dropna().nunique(),
            "Orders": unique_join(group["Order"]),
            "Order Statuses": unique_join(group["Order Status"]),
            "Total Quantity Ordered": group["Quantity Ordered"].sum(),
            "Earliest Customer Target Date": pd.to_datetime(
                group["Target Date"], errors="coerce"
            ).min(),
            "Latest Customer Target Date": pd.to_datetime(
                group["Target Date"], errors="coerce"
            ).max(),
            "Customer Target Dates": unique_date_join(group["Target Date"]),
            "Earliest Delivery Date": pd.to_datetime(
                group["Delivery Date"], errors="coerce"
            ).min(),
            "Latest Delivery Date": pd.to_datetime(
                group["Delivery Date"], errors="coerce"
            ).max(),
            "Delivery Dates": unique_date_join(group["Delivery Date"]),
        })

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows).sort_values("Normalized SKU").reset_index(drop=True)


def build_wms_summary(wms_df):
    columns = [
        "Normalized SKU",
        "Matched WMS Rows",
        "Total WMS Quantity",
        "Earliest Consumer Priority Date",
        "Latest Consumer Priority Date",
        "Consumer Priority Dates",
        "Current Location Count",
        "Current Locations",
        "Current Zones",
        "Previous Locations",
        "LPN Count",
    ]

    if wms_df.empty:
        return pd.DataFrame(columns=columns)

    rows = []
    for sku, group in wms_df.groupby("Normalized SKU", dropna=False):
        rows.append({
            "Normalized SKU": sku,
            "Matched WMS Rows": len(group),
            "Total WMS Quantity": group["WMS Quantity"].sum(),
            "Earliest Consumer Priority Date": pd.to_datetime(
                group["Consumer Priority Date"], errors="coerce"
            ).min(),
            "Latest Consumer Priority Date": pd.to_datetime(
                group["Consumer Priority Date"], errors="coerce"
            ).max(),
            "Consumer Priority Dates": unique_date_join(group["Consumer Priority Date"]),
            "Current Location Count": group["Current Location Normalized"].replace(
                "", pd.NA
            ).dropna().nunique(),
            "Current Locations": unique_join(group["Current Location Normalized"]),
            "Current Zones": unique_join(group["Current Location Zone"]),
            "Previous Locations": unique_join(group["Previous Location Normalized"]),
            "LPN Count": group["LPN #"].replace("", pd.NA).dropna().nunique(),
        })

    return pd.DataFrame(rows, columns=columns).sort_values("Normalized SKU").reset_index(drop=True)


def build_outputs(data_df, wms_df):
    data_summary = build_data_summary(data_df)

    if data_summary.empty:
        empty = pd.DataFrame()
        return empty, empty, empty, empty

    data_skus = set(data_summary["Normalized SKU"])
    wms_filtered = wms_df[wms_df["Normalized SKU"].isin(data_skus)].copy()
    wms_summary = build_wms_summary(wms_filtered)

    summary = data_summary.merge(wms_summary, on="Normalized SKU", how="left")

    numeric_fill_cols = [
        "Matched WMS Rows",
        "Total WMS Quantity",
        "Current Location Count",
        "LPN Count",
    ]
    for col in numeric_fill_cols:
        if col in summary.columns:
            summary[col] = summary[col].fillna(0)

    summary["Match Status"] = summary["Matched WMS Rows"].apply(
        lambda value: "MATCHED" if float(value) > 0 else "NO WMS MATCH"
    )
    summary["WMS Qty Minus Ordered Qty"] = (
        summary["Total WMS Quantity"].fillna(0)
        - summary["Total Quantity Ordered"].fillna(0)
    )
    summary["Days: Latest CPD vs Earliest Target"] = (
        pd.to_datetime(summary["Latest Consumer Priority Date"], errors="coerce")
        - pd.to_datetime(summary["Earliest Customer Target Date"], errors="coerce")
    ).dt.days

    detail = wms_filtered.merge(data_summary, on="Normalized SKU", how="left")
    if not detail.empty:
        detail["Days: CPD vs Earliest Target"] = (
            pd.to_datetime(detail["Consumer Priority Date"], errors="coerce")
            - pd.to_datetime(detail["Earliest Customer Target Date"], errors="coerce")
        ).dt.days
        detail["Date Match Check"] = detail["Days: CPD vs Earliest Target"].apply(
            lambda value: "NO DATE"
            if pd.isna(value)
            else ("CPD ON/AFTER TARGET" if value >= 0 else "CPD BEFORE TARGET")
        )

    detail_columns = [
        "Normalized SKU",
        "Data Item Values",
        "Total Quantity Ordered",
        "Order Count",
        "Orders",
        "Customer Values",
        "Earliest Customer Target Date",
        "Latest Customer Target Date",
        "Customer Target Dates",
        "Earliest Delivery Date",
        "Latest Delivery Date",
        "Delivery Dates",
        "WMS SKU Number",
        "WMS Quantity",
        "Consumer Priority Date",
        "Days: CPD vs Earliest Target",
        "Date Match Check",
        "Current Location Compact",
        "Current Location Normalized",
        "Current Location Zone",
        "Current Location Aisle",
        "Current Location Bay",
        "Current Location Position",
        "Previous Location Compact",
        "Previous Location Normalized",
        "Previous Location Zone",
        "Previous Location Aisle",
        "Previous Location Bay",
        "Previous Location Position",
        "LPN #",
        "WMS Row #",
    ]
    detail = detail[[col for col in detail_columns if col in detail.columns]]
    if not detail.empty:
        detail = detail.sort_values(
            ["Normalized SKU", "Consumer Priority Date", "Current Location Normalized"],
            na_position="last",
        ).reset_index(drop=True)

    line_match_cols = [
        "Normalized SKU",
        "Match Status",
        "Matched WMS Rows",
        "Total WMS Quantity",
        "Earliest Consumer Priority Date",
        "Latest Consumer Priority Date",
        "Current Locations",
    ]
    data_lines = data_df.copy().merge(summary[line_match_cols], on="Normalized SKU", how="left")
    data_lines["Days: Latest CPD vs Target Date"] = (
        pd.to_datetime(data_lines["Latest Consumer Priority Date"], errors="coerce")
        - pd.to_datetime(data_lines["Target Date"], errors="coerce")
    ).dt.days

    data_line_columns = [
        "Data Row #",
        "Data Item",
        "Normalized SKU",
        "Customer",
        "Order Status",
        "Order",
        "Quantity Ordered",
        "Target Date",
        "Delivery Date",
        "Match Status",
        "Matched WMS Rows",
        "Total WMS Quantity",
        "Earliest Consumer Priority Date",
        "Latest Consumer Priority Date",
        "Days: Latest CPD vs Target Date",
        "Current Locations",
    ]
    data_lines = data_lines[[col for col in data_line_columns if col in data_lines.columns]]
    data_lines = data_lines.sort_values(
        ["Normalized SKU", "Target Date", "Delivery Date"], na_position="last"
    ).reset_index(drop=True)

    unmatched = summary[summary["Match Status"] == "NO WMS MATCH"].copy()
    unmatched_columns = [
        "Normalized SKU",
        "Data Item Values",
        "Data Lines",
        "Customer Values",
        "Order Count",
        "Orders",
        "Total Quantity Ordered",
        "Earliest Customer Target Date",
        "Latest Customer Target Date",
        "Customer Target Dates",
        "Earliest Delivery Date",
        "Latest Delivery Date",
        "Delivery Dates",
        "Match Status",
    ]
    unmatched = unmatched[[col for col in unmatched_columns if col in unmatched.columns]]
    unmatched = unmatched.reset_index(drop=True)

    summary_columns = [
        "Normalized SKU",
        "Data Item Values",
        "Match Status",
        "Total Quantity Ordered",
        "Total WMS Quantity",
        "WMS Qty Minus Ordered Qty",
        "Matched WMS Rows",
        "LPN Count",
        "Data Lines",
        "Customer Values",
        "Order Count",
        "Orders",
        "Order Statuses",
        "Earliest Customer Target Date",
        "Latest Customer Target Date",
        "Customer Target Dates",
        "Earliest Delivery Date",
        "Latest Delivery Date",
        "Delivery Dates",
        "Earliest Consumer Priority Date",
        "Latest Consumer Priority Date",
        "Consumer Priority Dates",
        "Days: Latest CPD vs Earliest Target",
        "Current Location Count",
        "Current Zones",
        "Current Locations",
        "Previous Locations",
    ]
    summary = summary[[col for col in summary_columns if col in summary.columns]]
    summary = summary.sort_values(["Match Status", "Normalized SKU"]).reset_index(drop=True)

    return summary, detail, data_lines, unmatched


# ============================================================
# Excel output
# ============================================================

DATE_COLUMNS = {
    "Target Date",
    "Delivery Date",
    "Earliest Customer Target Date",
    "Latest Customer Target Date",
    "Earliest Delivery Date",
    "Latest Delivery Date",
    "Consumer Priority Date",
    "Earliest Consumer Priority Date",
    "Latest Consumer Priority Date",
}


def convert_dates_for_excel(df):
    output = df.copy()
    for col in output.columns:
        if col in DATE_COLUMNS:
            converted = pd.to_datetime(output[col], errors="coerce")
            if converted.notna().any():
                output[col] = converted.dt.date
    return output


def write_sheet(writer, df, sheet_name):
    safe_df = convert_dates_for_excel(df)
    safe_df.to_excel(writer, sheet_name=sheet_name, index=False)


def format_workbook(writer):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical="top")

    for sheet in writer.book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            header = str(column_cells[0].value or "")
            max_length = len(header)

            for cell in column_cells[1:2000]:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                header = str(sheet.cell(row=1, column=cell.column).value or "")
                if "Date" in header or "CPD" in header:
                    cell.number_format = "yyyy-mm-dd"
                elif (
                    "Quantity" in header
                    or "Qty" in header
                    or "Rows" in header
                    or "Count" in header
                    or "Days" in header
                ):
                    cell.number_format = "#,##0.##"


def to_excel_bytes(summary_df, detail_df, data_lines_df, unmatched_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_sheet(writer, summary_df, "SKU SUMMARY")
        write_sheet(writer, detail_df, "MATCHED WMS DETAIL")
        write_sheet(writer, data_lines_df, "DATA LINES")
        write_sheet(writer, unmatched_df, "UNMATCHED DATA SKUS")
        format_workbook(writer)

    output.seek(0)
    return output


def date_only_for_display(df):
    output = df.copy()
    for col in output.columns:
        if col in DATE_COLUMNS:
            converted = pd.to_datetime(output[col], errors="coerce")
            if converted.notna().any():
                output[col] = converted.dt.date
    return output


# ============================================================
# Streamlit app
# ============================================================

def main():
    st.set_page_config(page_title="SKU Data + WMS Matcher", layout="wide")
    st.title("SKU Data + WMS Matcher")
    st.caption(
        "Upload the same two Excel files. The output includes only SKUs from the Data file, "
        "matched to the big qPORT/WMS file."
    )

    with st.expander("What this app does"):
        st.markdown(
            "- Normalizes SKU numbers before matching.\n"
            "- Uses the Data file as the master SKU list.\n"
            "- Pulls only matching qPORT/WMS rows for those SKUs.\n"
            "- Normalizes current and previous locations.\n"
            "- Adds ordered quantities, WMS quantities, locations, LPNs, customer target dates, "
            "delivery dates, and consumer priority dates.\n"
            "- Does not allocate inventory and does not create shortage logic."
        )

    with st.sidebar:
        st.header("Upload Files")
        data_file = st.file_uploader("Data / orders file", type=["xlsx"], key="data_file")
        wms_file = st.file_uploader("Big qPORT / WMS file", type=["xlsx"], key="wms_file")

        st.divider()
        st.subheader("Optional filters")
        exclude_s_items = st.checkbox("Exclude Items starting with S", value=True)
        decimal_skus_only = st.checkbox("Only include decimal SKUs", value=True)

        st.divider()
        run_button = st.button("Build Matched Excel", type="primary", use_container_width=True)

    if not run_button:
        st.info("Upload both Excel files in the sidebar, then click Build Matched Excel.")
        return

    if data_file is None or wms_file is None:
        st.error("Please upload both Excel files before running the match.")
        return

    try:
        with st.spinner("Matching SKUs and building Excel..."):
            data_df = load_data_file(
                data_file,
                exclude_s_items=exclude_s_items,
                decimal_skus_only=decimal_skus_only,
            )
            wms_df = load_wms_file(wms_file)
            summary_df, detail_df, data_lines_df, unmatched_df = build_outputs(data_df, wms_df)
            excel_bytes = to_excel_bytes(summary_df, detail_df, data_lines_df, unmatched_df)

    except KeyError as error:
        st.error(str(error))
        return
    except Exception as error:
        st.error(f"Something went wrong while processing the files: {error}")
        return

    st.success("Done.")

    if summary_df.empty:
        st.warning("No Data SKUs were available after the selected filters.")
        return

    matched_count = int((summary_df["Match Status"] == "MATCHED").sum())
    unmatched_count = int((summary_df["Match Status"] == "NO WMS MATCH").sum())

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Data SKUs", summary_df["Normalized SKU"].nunique())
    col2.metric("Matched SKUs", matched_count)
    col3.metric("Unmatched SKUs", unmatched_count)
    col4.metric("Matched WMS Rows", len(detail_df))

    st.download_button(
        "Download matched SKU Excel",
        data=excel_bytes,
        file_name=f"sku_data_wms_match_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    tab1, tab2, tab3, tab4 = st.tabs([
        "SKU SUMMARY",
        "MATCHED WMS DETAIL",
        "DATA LINES",
        "UNMATCHED DATA SKUS",
    ])

    with tab1:
        st.dataframe(date_only_for_display(summary_df), use_container_width=True)

    with tab2:
        st.dataframe(date_only_for_display(detail_df), use_container_width=True)

    with tab3:
        st.dataframe(date_only_for_display(data_lines_df), use_container_width=True)

    with tab4:
        if unmatched_df.empty:
            st.info("Every Data SKU has at least one WMS match.")
        else:
            st.dataframe(date_only_for_display(unmatched_df), use_container_width=True)


if __name__ == "__main__":
    main()
