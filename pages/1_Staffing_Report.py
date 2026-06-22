import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import os
import re
import json
import datetime
from openai import OpenAI

import dc_config
import urllib

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False


st.set_page_config(page_title="Staffing Report Generator", layout="wide")

st.title("Staffing Report Generator")
st.caption("v3.3 — loader hard cap in build_summary")
st.error("⚡ v3.3 LOADED — if you see this, the new file is running")
st.write("Enter daily inputs, select who is present, and generate the staffing report.")

TEMPLATE_FILE = "staffing_template.xlsx"

if not os.path.exists(TEMPLATE_FILE):
    st.error("Template file not found. Put staffing_template.xlsx in the same folder as report.py.")
    st.stop()


# ============================================================
#  OPPORTUNITY CUSTOMER LIST (loaded from Excel)
#  File must be in the same folder as report.py.
#  Sheet: "OC Customer List"
#  Row 6  = headers (skipped by name check)
#  Row 7  = example row — skipped (name contains "market x" / "example")
#  Rows 8+ = real data
#  Columns:
#    A: Resers DC   B: Customer #   C: Customer Name   D: Address
#    E: Profile/Why OC   F: DC Requirements   G: Sign Off (Y/N)
#    H: Pictures (Y/N)   I: Other (Y/N)
# ============================================================
OC_FILE = "OC Cusotmer List.xlsx"
OC_SHEET = "OC Customer List"
OC_DATA_START = 8


@st.cache_data
def load_oc_customer_list():
    """
    Read the OC Excel file and return a (customers_list, error_message) tuple.
    Cached so it only reads once per app session.
    """
    if not os.path.exists(OC_FILE):
        return [], None

    try:
        wb = load_workbook(OC_FILE, data_only=True)
        if OC_SHEET not in wb.sheetnames:
            available = ', '.join(wb.sheetnames)
            return [], f"Sheet '{OC_SHEET}' not found in {OC_FILE}. Available sheets: {available}"


        ws = wb[OC_SHEET]
        customers = []

        for row_idx in range(OC_DATA_START, ws.max_row + 1):
            raw_name = ws.cell(row_idx, 3).value  # col C
            if not raw_name:
                continue

            name_clean = str(raw_name).strip().strip('"').lower()

            if "market x" in name_clean or "example" in name_clean:
                continue

            raw_cust_num = ws.cell(row_idx, 2).value  # col B
            raw_issue    = ws.cell(row_idx, 5).value  # col E
            raw_reqs     = ws.cell(row_idx, 6).value  # col F
            raw_signoff  = ws.cell(row_idx, 7).value  # col G
            raw_pictures = ws.cell(row_idx, 8).value  # col H

            issue = str(raw_issue).strip() if raw_issue else ""
            reqs  = str(raw_reqs).strip()  if raw_reqs  else ""

            sign_off = str(raw_signoff).strip().upper() == "Y" if raw_signoff else False
            pictures = str(raw_pictures).strip().upper() == "Y" if raw_pictures else False

            priority = "HIGH" if (sign_off or pictures) else "HIGH"

            base = name_clean.rstrip(" -").split(" - ")[0].strip()
            aliases = []
            for suffix in [" - all loads", " all loads", " fresh dc", " (olathe)"]:
                if base.endswith(suffix):
                    aliases.append(base.replace(suffix, "").strip())
            if "'" in base:
                aliases.append(base.replace("'", ""))
                aliases.append(base.replace("'s", ""))
            known_aliases = {
                "target rialto":          ["target"],
                "sobey's - all loads":    ["sobeys", "sobey", "sobey's"],
                "sysco kc (olathe)":      ["sysco kc", "sysco kansas city", "sysco olathe", "sysco kc olathe"],
                "pfs virgina":            ["pfs virginia", "pfs va"],
                "metro toronto fresh dc": ["metro toronto", "metro fresh"],
                "jewel's":                ["jewels", "jewel"],
                "awg":                    ["associated wholesale grocers"],
                "whataburguer":           ["whataburger"],
            }
            if name_clean in known_aliases:
                aliases += known_aliases[name_clean]

            aliases = list(dict.fromkeys(
                a for a in aliases if a and a != name_clean
            ))

            customers.append({
                "name": name_clean,
                "aliases": aliases,
                "customer_number": str(raw_cust_num).strip() if raw_cust_num else None,
                "issue": issue,
                "requirements": reqs,
                "sign_off": sign_off,
                "pictures": pictures,
                "priority": priority,
            })

        return customers, None

    except Exception as e:
        return [], f"Error loading OC customer list: {e}"



def find_oc_customers_in_board(board_text):
    oc_list, _ = load_oc_customer_list()
    board_lower = board_text.lower()
    matches = []
    for customer in oc_list:
        search_terms = [customer["name"]] + customer.get("aliases", [])
        found_terms = [term for term in search_terms if term.lower() in board_lower]
        if found_terms:
            matches.append({"customer": customer, "matched_on": found_terms})
    return matches


def build_oc_alert_text(oc_matches):
    if not oc_matches:
        return None
    lines = [
        "=== OPPORTUNITY CUSTOMER (OC) ALERT ===",
        "The following loads belong to customers on the Opportunity Customer List.",
        "These customers have a documented history of complaints and require special handling.",
        "Flag these loads explicitly in your analysis and include action items for each.",
        "",
    ]
    for match in oc_matches:
        c = match["customer"]
        lines.append(f"CUSTOMER: {c['name'].upper()}")
        lines.append(f"Matched on: {', '.join(match['matched_on'])}")
        lines.append(f"Priority: {c['priority']}")
        lines.append(f"DC Requirements: {c['requirements']}")
        if c["sign_off"]:
            lines.append("DC Supervisor Sign-Off REQUIRED before this load ships.")
        if c["pictures"]:
            lines.append("Photos REQUIRED: 3 on dock + 3 during loading (6 total). Email to manager.")
        lines.append("")
    lines += [
        "IMPORTANT: For every OC load identified above:",
        "1. Flag it clearly in your Board Summary section.",
        "2. Add a dedicated OC Action section with specific steps before this load ships.",
        "3. Recommend who should own the sign-off and photo process.",
        "4. Include this as one of the Top 3 Action Items if the load is active or upcoming.",
        "=== END OC ALERT ===",
    ]
    return "\n".join(lines)
# ============================================================
#  TT4 DEVICE LIST (loaded from Excel)
#  File: TT4_CUSTOMERS.xlsx, sheet "TT4 Customers"
#  Col A: TT4 #   Col B: Customer keyword (substring match against board customer)
#  Longest keyword wins so "WALMART CANADA" beats "WALMART".
# ============================================================
TT4_FILE = "TT4_CUSTOMERS.xlsx"
TT4_SHEET = "TT4 Customers"


@st.cache_data
def load_tt4_device_list():
    """Return list of {tt4_number, keyword, caution} from the TT4 device Excel.
    Col A: TT4 #   Col B: Customer keyword   Col D: RELEVANT INFO (caution sentence).
    Rows with a blank customer keyword are skipped (nothing to match on)."""
    if not os.path.exists(TT4_FILE):
        return []
    try:
        wb = load_workbook(TT4_FILE, data_only=True)
        ws = wb[TT4_SHEET] if TT4_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
        devices = []
        for row_idx in range(2, ws.max_row + 1):
            raw_num = ws.cell(row_idx, 1).value      # col A
            raw_kw = ws.cell(row_idx, 2).value       # col B
            raw_caution = ws.cell(row_idx, 4).value  # col D (RELEVANT INFO)
            if raw_num is None or raw_kw is None:
                continue
            keyword = str(raw_kw).strip().lower()
            number = str(raw_num).strip()
            caution = str(raw_caution).strip() if raw_caution is not None else ""
            if not keyword or not number:
                continue
            devices.append({
                "tt4_number": number,
                "keyword": keyword,
                "caution": caution,
            })
        # Longest keyword first so the most specific match wins (WALMART CORNWALL before WALMART).
        devices.sort(key=lambda d: len(d["keyword"]), reverse=True)
        return devices
    except Exception:
        return []


def normalize_tt4_match_text(value):
    """
    Normalize customer text for TT4 matching.
    Makes matching stronger against board names like:
    - SYSCO WEST COAST FLORIDA TK
    - SAFEWAY DENVER DELI TK (009137)
    - JEWEL OSCO DELI NEW - 5433
    """
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")

    # Remove customer/store codes in parentheses.
    text = re.sub(r"\([^)]*\)", " ", text)

    # Convert punctuation to spaces.
    text = re.sub(r"[^a-z0-9]+", " ", text)

    tokens = [t for t in text.split() if t]

    # Words that are common board noise and should not control the match.
    noise_words = {
        "tk", "truck", "load", "loads",
        "new", "fresh", "deli", "dc",
        "inc", "llc", "corp", "corporation",
        "company", "foods", "food",
        "the",
    }

    tokens = [t for t in tokens if t not in noise_words]
    return " ".join(tokens)


def tt4_match_tokens(value):
    return set(normalize_tt4_match_text(value).split())


def tt4_keyword_matches_customer(keyword, customer_name):
    """
    Stronger match:
    1. Normalized substring match.
    2. All keyword tokens appear in customer name.
    3. Common brand match for Sysco/Safeway/Jewel/etc.
    """
    keyword_norm = normalize_tt4_match_text(keyword)
    customer_norm = normalize_tt4_match_text(customer_name)

    if not keyword_norm or not customer_norm:
        return False

    # Example: "safeway denver" in "safeway denver 009137"
    if keyword_norm in customer_norm:
        return True

    keyword_tokens = tt4_match_tokens(keyword)
    customer_tokens = tt4_match_tokens(customer_name)

    if not keyword_tokens or not customer_tokens:
        return False

    # Example: keyword "safeway denver" matches board "SAFEWAY DENVER DELI TK (009137)"
    if keyword_tokens.issubset(customer_tokens):
        return True

    # Brand-family backup match.
    # Example: TT4 keyword "sysco foods" still matches board "SYSCO OKLAHOMA".
    brand_tokens = {
        "sysco",
        "safeway",
        "jewel",
        "jewels",
        "osco",
        "walmart",
        "target",
        "sobeys",
        "sobey",
        "metro",
        "awg",
        "pfs",
        "kroger",
        "costco",
        "heb",
        "aldi",
    }

    if keyword_tokens & customer_tokens & brand_tokens:
        return True

    return False


# Canadian Walmart locations. A board customer containing "WALMART" AND one of
# these words is a Canadian Walmart -> Canada TT4. Any other Walmart -> US TT4.
WALMART_CANADA_KEYWORDS = {"cornwall", "moncton", "brampton", "miss", "mississauga"}
WALMART_CANADA_TT4 = "28077"
WALMART_US_TT4 = "28040/28075"
WALMART_WEBSITE_CAUTION = "Make sure TT4 is logged into their website"


def match_tt4_device_for_customer(customer_name, tt4_devices=None):
    """
    Return (tt4_number, caution) for the keyword that best matches the board customer name,
    or (None, "") if nothing matches. Longest/most-specific keyword wins.

    Walmart is decided explicitly: any Walmart whose name contains a Canadian city keyword
    uses the Canada TT4; every other Walmart uses the US TT4. This prevents the sheet's
    longest-keyword match from sending a US (or unlisted Canadian) Walmart to the wrong device.
    """
    if tt4_devices is None:
        tt4_devices = load_tt4_device_list()

    customer_norm = normalize_tt4_match_text(customer_name)
    if not customer_norm:
        return None, ""

    # Walmart routing decided here, before the generic keyword loop.
    if "walmart" in customer_norm:
        if any(kw in customer_norm for kw in WALMART_CANADA_KEYWORDS):
            return WALMART_CANADA_TT4, WALMART_WEBSITE_CAUTION
        return WALMART_US_TT4, WALMART_WEBSITE_CAUTION

    best_device = None
    best_score = None

    for device in tt4_devices:
        keyword = device.get("keyword", "")

        # Skip the sheet's Walmart rows entirely — Walmart is handled above.
        if "walmart" in keyword:
            continue

        if not tt4_keyword_matches_customer(keyword, customer_name):
            continue

        keyword_norm = normalize_tt4_match_text(keyword)
        keyword_tokens = tt4_match_tokens(keyword)
        score = (len(keyword_tokens), len(keyword_norm))

        if best_score is None or score > best_score:
            best_score = score
            best_device = device

    if best_device:
        return best_device.get("tt4_number"), best_device.get("caution", "")
    return None, ""
    

def get_groq_client():
    if "GROQ_API_KEY" not in st.secrets:
        return None
    return OpenAI(
        api_key=st.secrets["GROQ_API_KEY"],
        base_url="https://api.groq.com/openai/v1",
    )


# ============================================================
#  CROSS DOCK DAILY INPUT ALERTS
#  Daily uploaded Excel. Matches Cross Dock Trip # to board load #.
#  Does NOT go to AI. This is a direct Python alert only.
# ============================================================

def normalize_crossdock_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%m/%d/%Y")
    text = str(value).replace("\n", " ").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_crossdock_load(value):
    """Return a clean 5-9 digit load number from Trip # / load fields."""
    text = normalize_crossdock_text(value)
    if not text:
        return ""
    if re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        return ""
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        return ""
    text = re.sub(r"^LD", "", text, flags=re.IGNORECASE)
    digits = re.sub(r"[^0-9]", "", text)
    return digits if 5 <= len(digits) <= 9 else ""


def parse_crossdock_pallets(value):
    text = normalize_crossdock_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return int(round(float(text)))
    except Exception:
        digits = re.sub(r"[^0-9]", "", text)
        return int(digits) if digits else 0


def normalize_crossdock_header(value):
    return re.sub(r"[^a-z0-9]+", "", normalize_crossdock_text(value).lower())


def detect_crossdock_header_map(values):
    """
    Expected headers in the uploaded Cross Dock file:
    Order# -PO# | Customer | #pallets | Location | Est Dispatch | Trip #
    """
    header_map = {}
    for idx, value in enumerate(values):
        header = normalize_crossdock_header(value)
        if not header:
            continue
        if "order" in header or "po" in header:
            header_map["order_po"] = idx
        elif "customer" in header:
            header_map["customer"] = idx
        elif "pallet" in header:
            header_map["pallets"] = idx
        elif "location" in header:
            header_map["location"] = idx
        elif "dispatch" in header:
            header_map["est_dispatch"] = idx
        elif "trip" in header or "load" in header:
            header_map["trip_load"] = idx
    if "trip_load" in header_map and ("location" in header_map or "pallets" in header_map):
        return header_map
    return None


def get_crossdock_value(values, header_map, key):
    idx = header_map.get(key)
    if idx is None or idx >= len(values):
        return ""
    return values[idx]


def read_crossdock_rows(crossdock_file):
    """Read every routed Cross Dock row from the daily upload."""
    if crossdock_file is None:
        return []

    crossdock_file.seek(0)
    file_name = crossdock_file.name.lower()
    rows = []

    try:
        if file_name.endswith(".xls"):
            sheets = pd.read_excel(crossdock_file, sheet_name=None, header=None, engine="xlrd")
            for sheet_name, df in sheets.items():
                header_map = None
                for idx, row in df.fillna("").iterrows():
                    values = [normalize_crossdock_text(v) for v in row.tolist()]
                    detected = detect_crossdock_header_map(values)
                    if detected:
                        header_map = detected
                        continue
                    if not header_map:
                        continue
                    trip_load = normalize_crossdock_load(get_crossdock_value(values, header_map, "trip_load"))
                    if not trip_load:
                        continue
                    rows.append({
                        "source_sheet": sheet_name,
                        "row_number": int(idx) + 1,
                        "trip_load": trip_load,
                        "order_po": normalize_crossdock_text(get_crossdock_value(values, header_map, "order_po")),
                        "customer": normalize_crossdock_text(get_crossdock_value(values, header_map, "customer")),
                        "pallets": parse_crossdock_pallets(get_crossdock_value(values, header_map, "pallets")),
                        "location": normalize_crossdock_text(get_crossdock_value(values, header_map, "location")),
                        "est_dispatch": normalize_crossdock_text(get_crossdock_value(values, header_map, "est_dispatch")),
                    })
            return rows

        wb = load_workbook(crossdock_file, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_map = None
            for row_idx in range(1, ws.max_row + 1):
                values = [normalize_crossdock_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
                detected = detect_crossdock_header_map(values)
                if detected:
                    header_map = detected
                    continue
                if not header_map:
                    continue
                trip_load = normalize_crossdock_load(get_crossdock_value(values, header_map, "trip_load"))
                if not trip_load:
                    continue
                rows.append({
                    "source_sheet": sheet_name,
                    "row_number": row_idx,
                    "trip_load": trip_load,
                    "order_po": normalize_crossdock_text(get_crossdock_value(values, header_map, "order_po")),
                    "customer": normalize_crossdock_text(get_crossdock_value(values, header_map, "customer")),
                    "pallets": parse_crossdock_pallets(get_crossdock_value(values, header_map, "pallets")),
                    "location": normalize_crossdock_text(get_crossdock_value(values, header_map, "location")),
                    "est_dispatch": normalize_crossdock_text(get_crossdock_value(values, header_map, "est_dispatch")),
                })
        return rows

    finally:
        crossdock_file.seek(0)


def find_crossdock_matches(crossdock_rows, board_rows):
    """Match Cross Dock Trip # to parsed board load number."""
    board_by_load = {}
    for row in board_rows or []:
        load = normalize_crossdock_load(row.get("load") or row.get("load_number"))
        if load and load not in board_by_load:
            board_by_load[load] = row

    matches = []
    for cd in crossdock_rows or []:
        load = normalize_crossdock_load(cd.get("trip_load"))
        if not load or load not in board_by_load:
            continue
        board = board_by_load[load]
        matches.append({
            **cd,
            "load": load,
            "board_customer": board.get("customer", ""),
            "board_time": board.get("time") or board.get("appt_time", ""),
            "board_door": board.get("door", ""),
            "board_status": board.get("status", ""),
            "board_type": board.get("type", ""),
        })
    return matches


# ============================================================
#  TT4 ALERTS
#  Direct Python alert only. Lists today's outbound loads that require TT4.
#  Does NOT go to AI.
# ============================================================

def tt4_value_requires_action(value):
    text = normalize_board_text(value).strip().upper()
    if not text:
        return False
    if text in ["N", "NO", "NONE", "0", "FALSE"]:
        return False
    return True


def find_tt4_required_loads(board_rows, selected_day):
    """
    Find selected-day loads that require TT4.
    A load requires TT4 when:
    - The parsed board flags include TT4-NEEDED from the blue Excel fill, or
    - The TT4 column contains any meaningful value.
    """
    selected_day = str(selected_day or "").strip().lower()
    matches = []
    seen_loads = set()

    for row in board_rows or []:
        row_day = str(row.get("day", "")).strip().lower()

        if selected_day and row_day != selected_day:
            continue

        load = normalize_crossdock_load(row.get("load") or row.get("load_number"))
        if not load or load in seen_loads:
            continue

        flags = row.get("flags", []) or []
        tt4_value = row.get("tt4", "")

        requires_tt4 = "TT4-NEEDED" in flags or tt4_value_requires_action(tt4_value)

        if not requires_tt4:
            continue

        seen_loads.add(load)

        matches.append({
            "load": load,
            "customer": row.get("customer", ""),
            "carrier": row.get("carrier", ""),
            "time": row.get("time") or row.get("appt_time", ""),
            "door": row.get("door", ""),
            "trailer": row.get("trailer", ""),
            "status": row.get("status", ""),
            "type": row.get("type", ""),
            "tt4": tt4_value,
            "flags": flags,
            "comments": row.get("comments", ""),
        })

    return matches


# ============================================================
#  NAME LOADING
#  Reads directly from the staffing sheets (col A), filtered
#  by the selected shift. No caching — avoids stale lists
#  when the user switches between 1st and 2nd shift.
#  1st shift: "Staffing sheet 1ST Shift"  col A rows 2+
#  2nd shift: "Staffing Sheet 2nd Shift"  col A rows 2+
# ============================================================
@st.cache_data
def load_names_for_shift(shift):
    wb = load_workbook(TEMPLATE_FILE, data_only=True)
    if shift == "1st":
        ws = wb["Staffing sheet 1ST Shift"]
    else:
        ws = wb["Staffing Sheet 2nd Shift"]
    names = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val and str(val).strip():
            names.append(str(val).strip())
        elif names:
            break
    return sorted(names)


def whole_workers(value):
    return int(float(value or 0) + 0.7)


def is_present(row):
    return str(row["Present"]).strip().lower() == "x"


def has_skill(row, code):
    """Return True only when the exact skill code is listed in the Skills cell.
    Examples: P/T/L counts as P, T, and L. A random letter inside another word does not count.
    """
    code = str(code or "").strip().upper()
    skills_text = str(row.get("Skills", "") if hasattr(row, "get") else row["Skills"]).upper()
    skills = [s for s in re.split(r"[^A-Z]+", skills_text) if s]
    return code in skills


def best_fit(row, text):
    return text.lower() in str(row["Best Fit"]).lower()


def name_contains(row, text):
    return text.lower() in str(row["Name"]).lower()


# Staffing NEED is sized from TODAY'S TOTAL CASES using fixed day-level splits,
# then multiplied by SHIFT_TARGET (0.52) inside calculate_needed().
#   Picks need  = total_cases * 0.37 * 0.52
#   Pulls need  = total_cases * 0.63 * 0.52  (converted to pallets via CASES_PER_PALLET)
# When actual board picks/pulls (K2 / L2) are available they are used instead
# to allocate against real remaining work — see cases_to_pick_override /
# full_pallets_override in calculate_needed().
PICK_SPLIT = 0.37   # share of total day cases that go through picking
PULL_SPLIT = 0.63   # share of total day cases that go through full-pallet pulls
CASES_PER_PALLET = 75

# Full-shift working hours used ONLY to size the staffing NEED (breaks excluded).
# This is a fixed shift length, not time-remaining: a full-shift quota must be
# divided by full-shift hours, or regenerating the report midday inflates every
# need. hours_remaining is still used everywhere else (goal/handoff math) where
# remaining work is measured against remaining time -- this does NOT touch those.
SHIFT_LABOR_HOURS = 11


def calculate_input_values(day, shift, total_cases):
    # Picks and pulls are sized from the full day's cases using fixed day-level splits.
    # SHIFT_TARGET (0.52) is applied inside calculate_needed(), so the final need is:
    #   Picks need  = total_cases * PICK_SPLIT * SHIFT_TARGET  (= total_cases * 0.37 * 0.52)
    #   Pulls need  = total_cases * PULL_SPLIT * SHIFT_TARGET  (= total_cases * 0.63 * 0.52)
    # `day` and `shift` are kept in the signature for API compatibility.
    cases_to_pick = total_cases * PICK_SPLIT
    full_pallets = (total_cases * PULL_SPLIT) / CASES_PER_PALLET
    return cases_to_pick, full_pallets


def calculate_needed(
    day, shift, total_cases, hours_remaining, total_outbound_loads_actual,
    crossroads_open, deer_creek_open, msb_open,
    cases_to_pick_override=None, full_pallets_override=None,
):
    # NOTE: Picking and Tasking NEED are sized off the FIXED full-shift length
    # (SHIFT_LABOR_HOURS) so regenerating mid-shift doesn't inflate the need.
    # Loading is DIFFERENT: its need is purely time-based (loaders * 1 trailer/hr * hrs_remaining).
    # Using SHIFT_LABOR_HOURS for Loading would undercount loaders needed when hrs_remaining < 11
    # (e.g., at 7 AM with 9.5 hrs left: 45-load target / (1 * 9.5) = 4.7 loaders needed,
    # not 2 as SHIFT_LABOR_HOURS would produce). So Loading need always uses hours_remaining.

    # Old fallback: estimate picks/pulls from total cases.
    # New board flow: use exact board totals from Outbound!L2 and Outbound!K2.
    calculated_cases_to_pick, calculated_full_pallets = calculate_input_values(day, shift, total_cases)
    cases_to_pick = calculated_cases_to_pick if cases_to_pick_override is None else float(cases_to_pick_override or 0)
    full_pallets = calculated_full_pallets if full_pallets_override is None else float(full_pallets_override or 0)
    inbound_pallets = 0
    if crossroads_open == "YES":
        inbound_pallets += 700
    if deer_creek_open == "YES":
        inbound_pallets += 500
    if msb_open == "YES":
        inbound_pallets += 640
    # 52% of the day's volume is the 1st-shift target for picks, pulls, and loading.
    # Inbound (unloading/receiving/putaway) is sized for its full inbound volume — no partial target.
    SHIFT_TARGET = 0.52
    # Safe hours_remaining for the Loading formula. Falls back to SHIFT_LABOR_HOURS only
    # if hours_remaining was not provided (e.g. called without it).
    try:
        _hrs_for_loading = max(0.5, float(hours_remaining or SHIFT_LABOR_HOURS))
    except Exception:
        _hrs_for_loading = SHIFT_LABOR_HOURS
    raw_needed = {
        "Unloading":     (inbound_pallets / 4) / (UNLOAD_RATE * SHIFT_LABOR_HOURS),
        "Receiving":     (inbound_pallets / 4) / (UNLOAD_RATE * SHIFT_LABOR_HOURS),
        "Putaway":       (inbound_pallets / 2) / (PULL_RATE * SHIFT_LABOR_HOURS),
        "Picking":       (cases_to_pick * SHIFT_TARGET) / (PICK_RATE * SHIFT_LABOR_HOURS),
        "Replenishment": ((cases_to_pick * SHIFT_TARGET) / CASES_PER_PALLET) / (PULL_RATE * SHIFT_LABOR_HOURS),
        "Full Pallets":  (full_pallets * SHIFT_TARGET) / (PULL_RATE * SHIFT_LABOR_HOURS),
        # Loading: how many loaders are needed to hit the loading goal in the actual hours left.
        # total_outbound_loads_actual already equals day_loads * LOAD_TARGET_SHARE (the 52% goal).
        # Dividing by (LOAD_RATE * hours_remaining) gives the correct real-time headcount.
        "Loading":       total_outbound_loads_actual / (LOAD_RATE * _hrs_for_loading),
    }
    needed = {
        "Unloading": max(MIN_UNLOADERS, whole_workers(raw_needed["Unloading"])),
        "Receiving":  max(MIN_RECEIVERS, whole_workers(raw_needed["Receiving"])),
        "Picking":    whole_workers(raw_needed["Picking"]),
        "Tasking":    whole_workers(
            raw_needed["Putaway"] + raw_needed["Replenishment"] + raw_needed["Full Pallets"]
        ),
        # Loading: use floor (int) not whole_workers (rounds up at +0.7).
        # whole_workers(5.26) = 6 because 5.26+0.7=5.96 -> int=5... wait actually:
        # whole_workers(5.59) = int(5.59+0.7) = int(6.29) = 6 — overshoots.
        # We want the minimum loaders that can cover the goal, not the rounded-up version.
        # The optimizer already accounts for pick throughput and will cap further if needed.
        "Loading":    max(1, int(raw_needed["Loading"])),
    }
    return needed, raw_needed, cases_to_pick, full_pallets, inbound_pallets


def generate_recommendations(staff, needed):
    needed = dict(needed)          # don't mutate the caller's dict
    needed["Receiving"] = 2        # Receiving is a hard cap of 2, every time
    assigned = {task: 0 for task in needed}
    staff["Recommended Task"] = ""
    present_indexes = staff[staff.apply(is_present, axis=1)].index.tolist()
    locked = set()  # indexes that no later pass may move

    def assign_if_needed(task, idx):
        if assigned[task] < needed[task]:
            staff.at[idx, "Recommended Task"] = task
            assigned[task] += 1
            return True
        return False

    def assign_locked(task, idx):
        """Hard pin: assign regardless of need, protect from every later move."""
        staff.at[idx, "Recommended Task"] = task
        assigned[task] += 1
        locked.add(idx)

    DALE_RECEIVERS = {"dale ferguson", "dale hrenchir"}
    for idx in present_indexes:
        row = staff.loc[idx]
        name_key = str(row["Name"]).strip().lower()
        if name_key in DALE_RECEIVERS and has_skill(row, "R"):
            assign_locked("Receiving", idx)

    for idx in present_indexes:
        if idx in locked:
            continue
        row = staff.loc[idx]
        if name_contains(row, "Alex") and has_skill(row, "U"):
            assign_if_needed("Unloading", idx)

    # Fill Loading early with L-skilled workers before L-skilled taskers can be absorbed into Tasking.
    for idx in present_indexes:
        if assigned["Loading"] >= needed["Loading"]:
            break
        if staff.at[idx, "Recommended Task"] != "":
            continue
        row = staff.loc[idx]
        if best_fit(row, "Load") and has_skill(row, "L"):
            assign_if_needed("Loading", idx)

    for idx in present_indexes:
        if assigned["Loading"] >= needed["Loading"]:
            break
        if staff.at[idx, "Recommended Task"] != "":
            continue
        row = staff.loc[idx]
        if has_skill(row, "L"):
            assign_if_needed("Loading", idx)

    best_fit_steps = [
        ("Unloading", "Unload", "U"),
        ("Loading",   "Load",   "L"),
        ("Receiving", "Receiv", "R"),
        ("Picking",   "Pick",   "P"),
        ("Tasking",   "Task",   "T"),
    ]
    for task, fit_text, skill in best_fit_steps:
        for idx in present_indexes:
            if staff.at[idx, "Recommended Task"] != "":
                continue
            row = staff.loc[idx]
            if best_fit(row, fit_text) and has_skill(row, skill):
                assign_if_needed(task, idx)

    skill_map = {"Unloading": "U", "Receiving": "R", "Loading": "L", "Picking": "P", "Tasking": "T"}
    for task, skill in skill_map.items():
        for idx in present_indexes:
            if assigned[task] >= needed[task]:
                break
            if staff.at[idx, "Recommended Task"] != "":
                continue
            row = staff.loc[idx]
            if has_skill(row, skill):
                assign_if_needed(task, idx)

    backup_tasks = ["Unloading", "Receiving", "Loading", "Picking", "Tasking"]
    backup_skill_map = {
        "Unloading": "U",
        "Receiving": "R",
        "Loading": "L",
        "Picking": "P",
        "Tasking": "T",
    }

    for task in backup_tasks:
        required_skill = backup_skill_map[task]

        while assigned[task] < needed[task]:
            found_worker = False

            # First: use best-fit workers who also have the required skill.
            for idx in present_indexes:
                if staff.at[idx, "Recommended Task"] != "":
                    continue
                row = staff.loc[idx]
                if best_fit(row, task[:5]) and has_skill(row, required_skill):
                    assign_if_needed(task, idx)
                    found_worker = True
                    break

            if found_worker:
                continue

            # Second: use anyone with the required skill.
            for idx in present_indexes:
                if staff.at[idx, "Recommended Task"] != "":
                    continue
                row = staff.loc[idx]
                if has_skill(row, required_skill):
                    assign_if_needed(task, idx)
                    found_worker = True
                    break

            if not found_worker:
                break

    # Final assignment pass: do not park anyone in Lead/Extra.
    # Any remaining present worker is assigned to a function they are actually skilled for.
    # First priority is to fill short functions; if all their skilled functions are full,
    # assign them to their best-fit skilled function as extra capacity.
    final_skill_map = {
        "Unloading": "U",
        "Receiving": "R",
        "Picking": "P",
        "Tasking": "T",
        "Loading": "L",
    }
    shortage_priority = ["Picking", "Tasking", "Loading", "Receiving", "Unloading"]
    fallback_priority = ["Tasking", "Picking", "Loading", "Receiving", "Unloading"]

    def choose_task_for_unassigned(row):
        skilled_tasks = [
            task for task, skill in final_skill_map.items()
            if has_skill(row, skill)
            # Never overflow Receiving or Loading beyond need in the final pass.
            # Receiving has a hard cap of 2. Loading has a hard cap of needed[Loading]
            # so that a pure-L worker doesn't inflate loading when Picking is the bottleneck.
            and not (task == "Receiving" and assigned.get("Receiving", 0) >= needed.get("Receiving", 0))
            and not (task == "Loading" and assigned.get("Loading", 0) >= needed.get("Loading", 0))
        ]
        if not skilled_tasks:
            return "Support"

        # Fill shortages first. This fixes cases like a T/R worker whose best fit is
        # Receiving but Receiving is full while Tasking is still short.
        for task in shortage_priority:
            if task in skilled_tasks and assigned.get(task, 0) < int(needed.get(task, 0) or 0):
                return task

        # If no shortage exists for that worker's skill set, honor best fit when possible.
        best_fit_map = {
            "Unloading": "Unload",
            "Receiving": "Receiv",
            "Picking": "Pick",
            "Tasking": "Task",
            "Loading": "Load",
        }
        for task, fit_text in best_fit_map.items():
            if task in skilled_tasks and best_fit(row, fit_text):
                return task

        # Last resort: assign extra capacity to a real skilled function, never Lead/Extra.
        for task in fallback_priority:
            if task in skilled_tasks:
                return task

        return skilled_tasks[0]

    for idx in present_indexes:
        if staff.at[idx, "Recommended Task"] == "":
            row = staff.loc[idx]
            chosen_task = choose_task_for_unassigned(row)
            staff.at[idx, "Recommended Task"] = chosen_task
            if chosen_task in assigned:
                assigned[chosen_task] += 1

    # Final smart balance pass: never leave a position overstaffed while another
    # position is understaffed if the present crew's skills can cover the shortage.
    # This includes two-step swaps. Example: Receiving is +1, Picking is -1,
    # Tasking is balanced. A T/R receiver can move to Tasking while a P/T tasker
    # moves to Picking, so Receiving becomes balanced and Picking is covered.
    rebalance_priority = ["Picking", "Tasking", "Loading", "Receiving", "Unloading"]
    task_to_skill = {
        "Unloading": "U",
        "Receiving": "R",
        "Picking": "P",
        "Tasking": "T",
        "Loading": "L",
    }
    fit_text_map = {
        "Unloading": "Unload",
        "Receiving": "Receiv",
        "Picking": "Pick",
        "Tasking": "Task",
        "Loading": "Load",
    }

    def _gap(task):
        return int(assigned.get(task, 0)) - int(needed.get(task, 0) or 0)

    def _has_task_skill(row, task):
        return has_skill(row, task_to_skill[task])

    def _move_score(idx, from_task, to_task):
        row = staff.loc[idx]
        # Prefer moving people who are not best-fit in the source and are best-fit in the destination.
        source_fit = best_fit(row, fit_text_map.get(from_task, from_task))
        dest_fit = best_fit(row, fit_text_map.get(to_task, to_task))
        # Then prefer multi-skilled workers for flexibility.
        skill_count = sum(1 for skill in task_to_skill.values() if has_skill(row, skill))
        return (1 if source_fit else 0, 0 if dest_fit else 1, -skill_count)

    moved = True
    while moved:
        moved = False
        short_tasks = [t for t in rebalance_priority if _gap(t) < 0]
        over_tasks = [t for t in rebalance_priority if _gap(t) > 0]
        if not short_tasks or not over_tasks:
            break

        # 1) Direct move: overstaffed task -> understaffed task.
        for short_task in short_tasks:
            if moved:
                break
            for over_task in over_tasks:
                if _gap(short_task) >= 0 or _gap(over_task) <= 0:
                    continue

                candidates = []
                for idx in present_indexes:
                    if idx in locked:
                        continue
                    if staff.at[idx, "Recommended Task"] != over_task:
                        continue
                    row = staff.loc[idx]
                    if _has_task_skill(row, short_task):
                        candidates.append(idx)

                if not candidates:
                    continue

                move_idx = sorted(candidates, key=lambda i: _move_score(i, over_task, short_task))[0]
                staff.at[move_idx, "Recommended Task"] = short_task
                assigned[over_task] -= 1
                assigned[short_task] += 1
                moved = True
                break

        if moved:
            continue

        # 2) Two-step chain move:
        # over worker A can cover a middle task, and middle worker B can cover the short task.
        # Move B middle -> short, then A over -> middle. Middle stays balanced.
        short_tasks = [t for t in rebalance_priority if _gap(t) < 0]
        over_tasks = [t for t in rebalance_priority if _gap(t) > 0]
        for short_task in short_tasks:
            if moved:
                break
            for over_task in over_tasks:
                if moved:
                    break
                if _gap(short_task) >= 0 or _gap(over_task) <= 0:
                    continue

                for middle_task in rebalance_priority:
                    if moved:
                        break
                    if middle_task in (short_task, over_task):
                        continue

                    # The middle task must not be overstaffed already. It can be balanced or short;
                    # since A replaces B, the middle count does not change.
                    if _gap(middle_task) > 0:
                        continue

                    over_candidates = []
                    for idx_a in present_indexes:
                        if idx_a in locked:
                            continue
                        if staff.at[idx_a, "Recommended Task"] != over_task:
                            continue
                        row_a = staff.loc[idx_a]
                        if _has_task_skill(row_a, middle_task):
                            over_candidates.append(idx_a)

                    if not over_candidates:
                        continue

                    middle_candidates = []
                    for idx_b in present_indexes:
                        if idx_b in locked:
                            continue
                        if staff.at[idx_b, "Recommended Task"] != middle_task:
                            continue
                        row_b = staff.loc[idx_b]
                        if _has_task_skill(row_b, short_task):
                            middle_candidates.append(idx_b)

                    if not middle_candidates:
                        continue

                    # Pick the least disruptive pair.
                    best_pair = None
                    best_score = None
                    for idx_a in over_candidates:
                        for idx_b in middle_candidates:
                            if idx_a == idx_b:
                                continue
                            score = (
                                _move_score(idx_a, over_task, middle_task),
                                _move_score(idx_b, middle_task, short_task),
                            )
                            if best_score is None or score < best_score:
                                best_score = score
                                best_pair = (idx_a, idx_b)

                    if best_pair is None:
                        continue

                    idx_a, idx_b = best_pair
                    staff.at[idx_b, "Recommended Task"] = short_task
                    staff.at[idx_a, "Recommended Task"] = middle_task
                    assigned[over_task] -= 1
                    assigned[short_task] += 1
                    # middle_task count is unchanged: one leaves and one enters.
                    moved = True
                    break

    return staff


def build_summary(staff, needed):
    present_recommendations = staff[
        staff["Present"].astype(str).str.strip().str.lower().eq("x")
        & staff["Recommended Task"].astype(str).str.strip().ne("")
    ].copy()

    # Hard cap: if more workers ended up assigned to Loading than needed, move the
    # excess to Picking (the typical bottleneck). This corrects any overspill from
    # generate_recommendations regardless of how it occurred.
    loading_needed = int(needed.get("Loading", 0))
    loading_assigned_idx = present_recommendations.index[
        present_recommendations["Recommended Task"] == "Loading"
    ].tolist()
    if len(loading_assigned_idx) > loading_needed:
        excess_indexes = loading_assigned_idx[loading_needed:]
        for idx in excess_indexes:
            present_recommendations.at[idx, "Recommended Task"] = "Picking"
            staff.at[idx, "Recommended Task"] = "Picking"

    needed_list   = pd.Series(needed, name="Needed")
    assigned_list = present_recommendations["Recommended Task"].value_counts().rename("Assigned")
    summary_table = pd.concat([needed_list, assigned_list], axis=1).fillna(0)
    summary_table["Needed"]     = summary_table["Needed"].astype(int)
    summary_table["Assigned"]   = summary_table["Assigned"].astype(int)
    summary_table["Difference"] = summary_table["Assigned"] - summary_table["Needed"]
    summary_table["Status"]     = summary_table["Difference"].apply(
        lambda x: "Good" if x == 0 else ("Overstaffed" if x > 0 else "Understaffed")
    )
    return present_recommendations, summary_table


def count_present_skill_capacity(staff):
    """Return the maximum available workers by task based on who is present and their listed skills."""
    task_skill_map = {
        "Unloading": "U",
        "Receiving": "R",
        "Picking": "P",
        "Tasking": "T",
        "Loading": "L",
    }
    present_rows = staff[staff.apply(is_present, axis=1)]
    return {
        task: int(present_rows.apply(lambda row: has_skill(row, skill), axis=1).sum())
        for task, skill in task_skill_map.items()
    }


def cap_allocation_to_available_skills(allocation, staff):
    """Never recommend more workers in a task than the present crew can legally staff by skill."""
    caps = count_present_skill_capacity(staff)
    return {
        task: min(int(allocation.get(task, 0) or 0), int(caps.get(task, 0) or 0))
        for task in ["Unloading", "Receiving", "Picking", "Tasking", "Loading"]
    }


def assigned_counts_from_summary(summary_table):
    """Return assigned counts from the staffing summary so high-level allocation matches named workers."""
    task_order = ["Picking", "Tasking", "Loading", "Unloading", "Receiving"]
    return {
        task: int(summary_table.loc[task, "Assigned"]) if task in summary_table.index else 0
        for task in task_order
    }

def compute_labor_availability(summary_table, present_recommendations, lead_extra_count=None):
    """
    Decide whether real surplus labor exists to reallocate.
    Surplus = a task with positive Difference (overstaffed) OR workers parked
    in Lead/Extra. If every task is at or below need and nobody is Lead/Extra,
    there is NO safe move — pulling labor into one function only reopens a gap
    in another. This is the source of truth for both the recommendation engine
    and the AI summary, so neither suggests a phantom move.
    """
    if lead_extra_count is None:
        lead_extra_count = int(
            (present_recommendations["Recommended Task"].astype(str).str.strip() == "Lead/Extra").sum()
        )
    else:
        lead_extra_count = int(lead_extra_count)

    surplus_tasks = {}
    short_tasks = {}
    for task, row in summary_table.iterrows():
        if task == "Lead/Extra":
            continue  # counted via lead_extra_count
        diff = int(row["Difference"])
        if diff > 0:
            surplus_tasks[task] = diff
        elif diff < 0:
            short_tasks[task] = diff

    total_surplus = sum(surplus_tasks.values()) + lead_extra_count
    total_short = sum(abs(v) for v in short_tasks.values())
    net_gap = int(summary_table["Difference"].sum())

    return {
        "lead_extra_count": lead_extra_count,
        "surplus_tasks": surplus_tasks,          # {task: +n}
        "short_tasks": short_tasks,              # {task: -n}
        "total_surplus_workers": total_surplus,
        "total_short_workers": total_short,
        "net_gap": net_gap,
        "has_available_labor": total_surplus > 0,
    }


def build_recommendations(summary_table, present_recommendations, raw_needed, hours_remaining, notes, availability=None):
    recommendations = []
    if availability is None:
        availability = compute_labor_availability(summary_table, present_recommendations)
    surplus_tasks    = availability["surplus_tasks"]
    lead_extra_count = availability["lead_extra_count"]
    has_labor        = availability["has_available_labor"]

    total_labor_gap = int(summary_table["Difference"].sum())
    labor_hours_gap = total_labor_gap * hours_remaining
    recommendations.append(
        f"Current labor balance estimate: {labor_hours_gap:+.1f} labor-hours. "
        f"Positive means extra capacity; negative means short capacity."
    )

    for task, row in summary_table.iterrows():
        diff = int(row["Difference"])
        if diff < 0:
            recommendations.append(
                f"{task}: approximately {abs(diff * hours_remaining):.1f} labor-hours behind based on current staffing vs need."
            )
        elif diff > 0:
            recommendations.append(f"{task}: approximately {diff * hours_remaining:.1f} labor-hours ahead / available capacity.")
        else:
            recommendations.append(f"{task}: Staffing is balanced.")

    # --- No-bench headline -------------------------------------------------
    if not has_labor and total_labor_gap < 0:
        recommendations.append(
            "No surplus labor available to reallocate. Every present worker is assigned and no area "
            "is overstaffed, so moving labor into one function reopens a gap in another. Closing the "
            "remaining gaps requires overtime, an early 2nd-shift start, or borrowing labor."
        )

    picking_gap   = int(summary_table.loc["Picking",   "Difference"]) if "Picking"   in summary_table.index else 0
    tasking_gap   = int(summary_table.loc["Tasking",   "Difference"]) if "Tasking"   in summary_table.index else 0
    receiving_gap = int(summary_table.loc["Receiving", "Difference"]) if "Receiving" in summary_table.index else 0
    unloading_gap = int(summary_table.loc["Unloading", "Difference"]) if "Unloading" in summary_table.index else 0
    loading_gap   = int(summary_table.loc["Loading",   "Difference"]) if "Loading"   in summary_table.index else 0

    if picking_gap < 0:
        recommendations.append("High picking short risk detected.")
        recommendations.append("Avoid pulling pickers into unloading or loading unless outbound service is critical.")
        if tasking_gap > 0:
            recommendations.append(f"Tasking has {tasking_gap} extra worker(s) — move into replenishment to protect pickers.")
        elif lead_extra_count > 0:
            recommendations.append(f"{lead_extra_count} Lead/Extra worker(s) available — flex into replenishment or picking support.")
        else:
            recommendations.append("No safe internal move to protect pickers — close the gap with overtime or an early 2nd-shift start.")

    if unloading_gap < 0 or receiving_gap < 0:
        recommendations.append("Inbound flow risk detected. Falling behind may create dock congestion and delayed putaway.")
        if tasking_gap > 0:
            recommendations.append("Tasking has surplus labor that can temporarily support unloading or receiving.")
        elif lead_extra_count > 0:
            recommendations.append("Lead/Extra labor can temporarily support inbound.")
        else:
            recommendations.append("No surplus labor to support inbound without opening an outbound gap.")

    if loading_gap < 0:
        recommendations.append("Outbound loading risk detected. Late departures and service failures may increase.")
        recommendations.append("Protect loading labor before reallocating to non-critical work.")
        if lead_extra_count > 0:
            recommendations.append("Use Lead/Extra labor to support outbound staging or trailer cleanup.")
        elif not has_labor:
            recommendations.append("No surplus labor to add to loading — prioritize existing loaders on the earliest departures.")

    if total_labor_gap > 1:
        recommendations.append("Operation currently has excess labor capacity.")
        recommendations.append("Consider deep cleaning, trailer audits, replenishment cleanup, or cross-training.")
        recommendations.append("Extra labor could be used proactively to prevent later picking shortages.")

    inbound_pressure  = raw_needed["Unloading"] + raw_needed["Receiving"] + raw_needed["Putaway"]
    outbound_pressure = raw_needed["Picking"] + raw_needed["Loading"]
    if inbound_pressure > outbound_pressure * 1.3:
        recommendations.append("Inbound workload is significantly heavier than outbound.")
        recommendations.append("Focus on unloading, receiving, and putaway to avoid congestion.")
    elif outbound_pressure > inbound_pressure * 1.3:
        recommendations.append("Outbound workload is significantly heavier than inbound.")
        recommendations.append("Prioritize replenishment and picking continuity to avoid shorts.")

    if hours_remaining <= 4:
        recommendations.append("Shift is entering final hours. Prioritize completion work and outbound execution.")
    elif hours_remaining >= 8:
        recommendations.append("Enough shift time remains to strategically rebalance labor before bottlenecks form.")

    lower_notes = notes.lower()
    if "late"  in lower_notes:
        recommendations.append("Manager notes mention late loads. Prioritize outbound execution and trailer readiness.")
    if "short" in lower_notes:
        recommendations.append("Manager notes indicate short risk. Protect replenishment and picking flow.")
    if "live"  in lower_notes:
        recommendations.append("Live loads detected in notes. Prioritize those doors before drop trailers.")
    if "cpu"   in lower_notes:
        recommendations.append("CPU loads referenced. Ensure loading labor is protected.")

    return recommendations


# ============================================================
#  BOARD EXCEL READING
# ============================================================
BOARD_DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]


def normalize_board_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime.datetime):
        return value.strftime("%m/%d/%Y")
    try:
        import math
        if isinstance(value, float) and math.isnan(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\n", " ").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_board_date(value):
    text = normalize_board_text(value)
    if not text:
        return ""
    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"]:
        try:
            return pd.to_datetime(text, format=fmt).strftime("%m/%d/%Y")
        except Exception:
            pass
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%m/%d/%Y")
    except Exception:
        pass
    return text


def normalize_board_time(value):
    text = normalize_board_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        h, m = text.split(":")
        return f"{int(h):02d}:{m}"
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%H:%M")
    except Exception:
        pass
    return text


def looks_like_board_load(value):
    text = normalize_board_text(value)
    if not text:
        return ""
    if text in BOARD_DAY_NAMES:
        return ""
    if re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        return ""
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        return ""
    digits = re.sub(r"[^0-9]", "", text)
    if 5 <= len(digits) <= 9:
        return digits
    return ""


def detect_board_status(value):
    text = normalize_board_text(value).upper()
    if not text:
        return ""
    if "LOADED SHORT" in text:
        return "Loaded Short"
    if "PICKING/SHORT" in text or "PICKING SHORT" in text:
        return "Picking/Short"
    if "READY/SHORT" in text or re.search(r"\bR/S\b", text):
        return "R/S"
    if re.search(r"\bRTL\b", text) or "READY TO LOAD" in text:
        return "RTL"
    if "NO DRIVER" in text:
        return "No Driver"
    if "PICKING" in text:
        return "Picking"
    if "COMPLETED" in text or re.search(r"\bCOMPLETE\b", text):
        return "Completed"
    if re.search(r"\bLATE\b", text):
        return "Late"
    if re.search(r"\bLOADED\b", text):
        return "Loaded"
    return ""


def detect_trailer_field_late(trailer_value):
    text = normalize_board_text(trailer_value).upper()
    if not text:
        return False
    if re.search(r"\bLATE\b", text):
        return True
    if re.match(r"^ETA\b", text):
        return True
    return False


def board_cell_flags(cell):
    flags = []
    fill_color = ""
    font_color = ""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            if fill.fgColor.type == "rgb":
                fill_color = str(fill.fgColor.rgb).upper()
            elif fill.fgColor.type == "indexed":
                fill_color = str(fill.fgColor.indexed).upper()
    except Exception:
        pass
    try:
        font = cell.font
        if font and font.color and font.color.type == "rgb":
            font_color = str(font.color.rgb).upper()
    except Exception:
        pass
    if fill_color in ("FFFFFF00", "00FFFF00", "FFFF00", "0000000D"):
        flags.append("LOAD-CHECK")
    if fill_color in ("FFADD8E6", "FF87CEEB", "FFADD8FF", "FFB0E0E6", "FF00BFFF"):
        flags.append("TT4-NEEDED")
    if font_color in ("FFFF0000", "00FF0000"):
        flags.append("CANADIAN")
    return flags


def parse_number(value):
    """
    Parse board count cells safely.
    Important: Excel formula results can be numeric decimals like 607.0666666667.
    The old parser converted that to text and stripped the decimal point, which became 60706666667.
    For actual numeric Excel values, round to the nearest whole count instead.
    """
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        try:
            return int(round(float(value)))
        except Exception:
            return 0

    text = normalize_board_text(value)
    if not text or text.strip() in ("", " "):
        return 0

    # If text still looks like a decimal number, parse it as a number first.
    cleaned = text.replace(",", "").strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
        try:
            return int(round(float(cleaned)))
        except Exception:
            pass

    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else 0


# ============================================================
#  OUTBOUND BOARD COLUMN MAPPING
#  Supports both board layouts:
#  OLD: A Load | B Customer | C Carrier | D Time | E Door | F Trailer | G Status | H TT4 | I Loader | J Comments | K Pulls | L Picks
#  NEW: A Load | B Customer | C Carrier | D Type | E Time | F Door | G Trailer | H Status | I TT4 | J Loader | K Pulls | L Picks | O Comments
# ============================================================

def clean_header_text(value):
    return re.sub(r"[^a-z0-9#]+", " ", normalize_board_text(value).lower()).strip()


def parse_board_day_name(value):
    text = normalize_board_text(value).strip()
    for day_name in BOARD_DAY_NAMES:
        if text.lower() == day_name.lower():
            return day_name
    return ""


def default_old_col_map():
    return {
        "load_number": 0,
        "customer": 1,
        "carrier": 2,
        "type": None,
        "appt_time": 3,
        "door": 4,
        "trailer": 5,
        "status": 6,
        "tt4": 7,
        "loader": 8,
        "comments": 9,
        "pulls": 10,
        "picks": 11,
    }


def default_new_col_map():
    return {
        "load_number": 0,
        "customer": 1,
        "carrier": 2,
        "type": 3,
        "appt_time": 4,
        "door": 5,
        "trailer": 6,
        "status": 7,
        "tt4": 8,
        "loader": 9,
        "comments": 14,
        "pulls": 10,
        "picks": 11,
    }


def detect_board_layout_from_header(values):
    """Return the correct column map from the header row when possible."""
    headers = [clean_header_text(v) for v in values]
    joined = " | ".join(headers)

    has_board_header = any(h in ("load #", "load", "load number", "destination", "customer", "carrier", "status") for h in headers)
    if not has_board_header:
        return None

    # If Type is in column D, this is the refreshed layout.
    if len(headers) > 3 and headers[3] in ("type", "load type"):
        return default_new_col_map()

    # If Status is in column H, this is the refreshed layout even if Type header is blank.
    if len(headers) > 7 and headers[7] == "status":
        return default_new_col_map()

    # If Status is in column G, this is the older layout.
    if len(headers) > 6 and headers[6] == "status":
        return default_old_col_map()

    # Fallback: use header positions if labels are present.
    col_map = default_new_col_map()
    label_map = {
        "load_number": ["load #", "load", "load number", "ld"],
        "customer": ["customer", "destination", "ship to", "consignee"],
        "carrier": ["carrier"],
        "type": ["type", "load type"],
        "appt_time": ["time", "appt", "appointment", "appointment time"],
        "door": ["door"],
        "trailer": ["trailer", "tr", "trailer #"],
        "status": ["status"],
        "tt4": ["tt4"],
        "loader": ["loader"],
        "comments": ["comments", "comment", "notes", "note"],
        "pulls": ["pulls", "pull"],
        "picks": ["picks", "pick"],
    }
    for key, aliases in label_map.items():
        for idx, header in enumerate(headers):
            if header in aliases:
                col_map[key] = idx
                break
    return col_map


def get_mapped_value(values, col_map, key):
    idx = col_map.get(key)
    if idx is None or idx < 0 or idx >= len(values):
        return ""
    return values[idx]


def derive_board_type(type_value, trailer_value, raw_values):
    explicit = normalize_board_text(type_value)
    if explicit:
        return explicit

    raw_upper = " ".join(normalize_board_text(v) for v in raw_values).upper()
    trailer_upper = normalize_board_text(trailer_value).upper()

    if "CPU" in raw_upper and "LIVE" in raw_upper:
        return "CPU - Live"
    if "CPU" in raw_upper and "DROP" in raw_upper:
        return "CPU - Drop"
    if "LIVE" in trailer_upper or "LIVE" in raw_upper:
        return "Live"
    if "DROP" in trailer_upper or "DROP" in raw_upper:
        return "Drop"
    if "CPU" in raw_upper:
        return "CPU"
    return ""


def build_board_row_from_values(values, col_map, source, current_day, current_date, row_number=None, flags=None):
    load_number = looks_like_board_load(get_mapped_value(values, col_map, "load_number"))
    if not load_number:
        return None

    trailer_value = get_mapped_value(values, col_map, "trailer")
    status_value = get_mapped_value(values, col_map, "status")

    if detect_trailer_field_late(trailer_value):
        status = "Late"
    else:
        status = detect_board_status(status_value)
    if not status:
        status = detect_board_status(" ".join(values))

    row = {
        "source": source,
        "day": current_day,
        "date": current_date,
        "load_number": load_number,
        "customer": get_mapped_value(values, col_map, "customer"),
        "carrier": get_mapped_value(values, col_map, "carrier"),
        "appt_time": normalize_board_time(get_mapped_value(values, col_map, "appt_time")),
        "door": get_mapped_value(values, col_map, "door"),
        "trailer": trailer_value,
        "status": status,
        "type": derive_board_type(get_mapped_value(values, col_map, "type"), trailer_value, values),
        "tt4": get_mapped_value(values, col_map, "tt4"),
        "loader": get_mapped_value(values, col_map, "loader"),
        "comments": get_mapped_value(values, col_map, "comments"),
        "pulls": parse_number(get_mapped_value(values, col_map, "pulls")),
        "picks": parse_number(get_mapped_value(values, col_map, "picks")),
        "flags": sorted(set(flags or [])),
        "raw_row": " | ".join(v for v in values if v),
    }
    if row_number is not None:
        row["row_number"] = row_number
    return row


def read_board_today_totals_from_excel(board_file):
    """
    New board requirement:
    K2 = pulls left for today
    L2 = picks left for today
    """
    board_file.seek(0)
    try:
        wb = load_workbook(board_file, data_only=True)
        sheet_name = "Outbound" if "Outbound" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        return {
            "pulls_left_today": parse_number(ws["K2"].value),
            "picks_left_today": parse_number(ws["L2"].value),
        }
    except Exception:
        return {"pulls_left_today": 0, "picks_left_today": 0}
    finally:
        board_file.seek(0)


def board_records_from_excel(board_file):
    board_file.seek(0)
    file_name = board_file.name.lower()
    all_rows = []

    if file_name.endswith(".xls"):
        sheets = pd.read_excel(board_file, sheet_name=None, header=None, engine="xlrd")
        for sheet_name, df in sheets.items():
            df = df.fillna("")
            current_day = ""
            current_date = ""
            col_map = default_new_col_map()
            for idx, row in df.iterrows():
                values = [normalize_board_text(v) for v in row.tolist()]
                while len(values) < 16:
                    values.append("")

                detected_map = detect_board_layout_from_header(values)
                if detected_map:
                    col_map = detected_map
                    continue

                day_name = parse_board_day_name(values[0])
                if day_name:
                    current_day = day_name
                    current_date = normalize_board_date(values[1])
                    continue

                parsed = build_board_row_from_values(values, col_map, sheet_name, current_day, current_date, flags=[])
                if parsed:
                    all_rows.append(parsed)
        return all_rows

    wb = load_workbook(board_file, data_only=True)
    outbound_sheet = None
    for candidate in ["Outbound", "outbound", "OUTBOUND"]:
        if candidate in wb.sheetnames:
            outbound_sheet = candidate
            break
    sheets_to_read = [outbound_sheet] if outbound_sheet else wb.sheetnames

    for sheet_name in sheets_to_read:
        ws = wb[sheet_name]
        current_day = ""
        current_date = ""
        consecutive_empty = 0
        col_map = default_new_col_map()

        for row_idx in range(1, ws.max_row + 1):
            values = []
            flags = []
            has_content = False
            # Read farther than the old board because Comments moved to column O.
            for col_idx in range(1, 17):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    has_content = True
                values.append(normalize_board_text(cell.value))
                for flag in board_cell_flags(cell):
                    flags.append(flag)

            if not has_content:
                consecutive_empty += 1
                if consecutive_empty >= 15:
                    break
                continue
            consecutive_empty = 0

            detected_map = detect_board_layout_from_header(values)
            if detected_map:
                col_map = detected_map
                continue

            day_name = parse_board_day_name(values[0])
            if day_name:
                current_day = day_name
                current_date = normalize_board_date(values[1])
                continue

            parsed = build_board_row_from_values(
                values, col_map, sheet_name, current_day, current_date,
                row_number=row_idx, flags=flags
            )
            if parsed:
                all_rows.append(parsed)

    return all_rows


def board_records_from_csv(board_file):
    board_file.seek(0)
    df = pd.read_csv(board_file, header=None).fillna("")
    current_day = ""
    current_date = ""
    all_rows = []
    col_map = default_new_col_map()

    for idx, row in df.iterrows():
        values = [normalize_board_text(v) for v in row.tolist()]
        while len(values) < 16:
            values.append("")

        detected_map = detect_board_layout_from_header(values)
        if detected_map:
            col_map = detected_map
            continue

        day_name = parse_board_day_name(values[0])
        if day_name:
            current_day = day_name
            current_date = normalize_board_date(values[1])
            continue

        parsed = build_board_row_from_values(values, col_map, "CSV Board", current_day, current_date, flags=[])
        if parsed:
            all_rows.append(parsed)

    return all_rows

def board_records_from_inbound_sheet(board_file):
    board_file.seek(0)
    try:
        wb = load_workbook(board_file, data_only=True)
    except Exception:
        return []

    inbound_sheet = None
    for candidate in ["Inbound", "inbound", "INBOUND"]:
        if candidate in wb.sheetnames:
            inbound_sheet = candidate
            break
    if not inbound_sheet:
        return []

    ws = wb[inbound_sheet]
    all_rows = []
    current_day = ""
    current_date = ""

    for row_idx in range(1, ws.max_row + 1):
        has_content = any(ws.cell(row_idx, c).value is not None for c in range(1, 12))
        if not has_content:
            continue

        col1 = normalize_board_text(ws.cell(row_idx, 1).value)
        col1_title = col1.strip().title()
        if col1_title in BOARD_DAY_NAMES:
            current_day = col1_title
            current_date = normalize_board_date(ws.cell(row_idx, 2).value)
            continue

        if col1.lower() in ("load number", "load #", "load"):
            continue

        load_number = looks_like_board_load(col1)
        if not load_number:
            continue

        all_rows.append({
            "source":      inbound_sheet,
            "day":         current_day,
            "date":        current_date,
            "load_number": load_number,
            "carrier":     normalize_board_text(ws.cell(row_idx, 2).value),
            "appt_time":   normalize_board_time(ws.cell(row_idx, 3).value),
            "type":        normalize_board_text(ws.cell(row_idx, 4).value),
            "trailer":     normalize_board_text(ws.cell(row_idx, 5).value),
            "status":      normalize_board_text(ws.cell(row_idx, 6).value),
            "receiver":    normalize_board_text(ws.cell(row_idx, 7).value),
            "origin":      normalize_board_text(ws.cell(row_idx, 8).value),
            "or_number":   normalize_board_text(ws.cell(row_idx, 9).value),
            "notes":       normalize_board_text(ws.cell(row_idx, 10).value),
        })

    return all_rows


def build_python_inbound_summary(inbound_rows):
    summary = {
        "loads_read_from_inbound": len(inbound_rows),
        "loads_by_day": {},
        "live_loads": 0,
        "drop_loads": 0,
        "on_lot": 0,
        "at_door": 0,
        "loads_with_receiver": 0,
        "loads_missing_receiver": 0,
    }
    for row in inbound_rows:
        day_key = row.get("day") or "Unknown Day"
        summary["loads_by_day"][day_key] = summary["loads_by_day"].get(day_key, 0) + 1

        type_upper   = row.get("type",   "").upper()
        status_upper = row.get("status", "").upper()

        if "LIVE"   in type_upper:
            summary["live_loads"] += 1
        if "DROP"   in type_upper:
            summary["drop_loads"] += 1
        if "ON LOT" in status_upper:
            summary["on_lot"] += 1
        if "DOOR"   in status_upper:
            summary["at_door"] += 1
        if row.get("receiver"):
            summary["loads_with_receiver"] += 1
        else:
            summary["loads_missing_receiver"] += 1

    return summary


def build_python_board_summary(board_rows):
    summary = {
        "loads_read_from_board":        len(board_rows),
        "loads_by_day":                 {},
        "loads_by_date":                {},
        "status_counts":                {},
        "late_loads":                   0,
        "rtl_loads":                    0,
        "rs_loads":                     0,
        "picking_loads":                0,
        "picking_short_loads":          0,
        "loaded_short_loads":           0,
        "completed_loads":              0,
        "blank_or_not_started_loads":   0,
        "live_loads":                   0,
        "drop_loads":                   0,
        "cpu_loads":                    0,
        "tt4_needed_loads":             0,
        "load_check_loads":             0,
        "canadian_loads":               0,
        "loads_with_loader_assigned":   0,
        "loads_missing_loader":         0,
        "late_load_details":            [],
        "rs_load_details":              [],
        "picking_short_details":        [],
        "loaded_short_details":         [],
        "rtl_details":                  [],
        "blank_or_not_started_details": [],
        "priority_load_details":        [],
    }

    for row in board_rows:
        day_key      = row.get("day")    or "Unknown Day"
        date_key     = row.get("date")   or "Unknown Date"
        status       = row.get("status") or "Blank/Not Started"
        status_upper = status.upper()
        raw_upper    = row.get("raw_row", "").upper()
        flags        = row.get("flags", [])

        summary["loads_by_day"][day_key]   = summary["loads_by_day"].get(day_key, 0)   + 1
        summary["loads_by_date"][date_key] = summary["loads_by_date"].get(date_key, 0) + 1
        summary["status_counts"][status]   = summary["status_counts"].get(status, 0)   + 1

        if "LATE" in status_upper or "LATE " in f" {raw_upper} ":
            summary["late_loads"] += 1
            summary["late_load_details"].append(row)
        if status_upper == "RTL" or "READY TO LOAD" in status_upper:
            summary["rtl_loads"] += 1
            summary["rtl_details"].append(row)
        if status_upper in ["R/S", "READY/SHORT"] or "R/S" in raw_upper:
            summary["rs_loads"] += 1
            summary["rs_load_details"].append(row)
        if status_upper == "PICKING":
            summary["picking_loads"] += 1
        if "PICKING/SHORT" in status_upper or "PICKING SHORT" in status_upper:
            summary["picking_short_loads"] += 1
            summary["picking_short_details"].append(row)
        if "LOADED SHORT" in status_upper:
            summary["loaded_short_loads"] += 1
            summary["loaded_short_details"].append(row)
        if "COMPLETED" in status_upper or status_upper == "COMPLETE":
            summary["completed_loads"] += 1
        if not row.get("status"):
            summary["blank_or_not_started_loads"] += 1
            summary["blank_or_not_started_details"].append(row)
        if "LIVE" in raw_upper:
            summary["live_loads"] += 1
            summary["priority_load_details"].append(row)
        if "DROP" in raw_upper:
            summary["drop_loads"] += 1
        if "CPU" in raw_upper:
            summary["cpu_loads"] += 1
            summary["priority_load_details"].append(row)
        if "TT4-NEEDED" in flags:
            summary["tt4_needed_loads"] += 1
            summary["priority_load_details"].append(row)
        if "LOAD-CHECK" in flags:
            summary["load_check_loads"] += 1
            summary["priority_load_details"].append(row)
        if "CANADIAN" in flags:
            summary["canadian_loads"] += 1
            summary["priority_load_details"].append(row)
        if row.get("loader"):
            summary["loads_with_loader_assigned"] += 1
        else:
            summary["loads_missing_loader"] += 1

    seen = set()
    unique_priority = []
    for item in summary["priority_load_details"]:
        key = (item.get("load_number"), item.get("row_number"), item.get("source"))
        if key not in seen:
            seen.add(key)
            unique_priority.append(item)
    summary["priority_load_details"] = unique_priority

    return summary


def compact_board_rows_for_ai(board_rows):
    """Strip raw_row and row_number — send only what the AI needs."""
    compact_rows = []
    for row in board_rows:
        compact_rows.append({
            "day":      row.get("day", ""),
            "date":     row.get("date", ""),
            "load":     row.get("load_number", ""),
            "customer": row.get("customer", ""),
            "carrier":  row.get("carrier", ""),
            "time":     row.get("appt_time", ""),
            "door":     row.get("door", ""),
            "trailer":  row.get("trailer", ""),
            "status":   row.get("status", ""),
            "type":     row.get("type", ""),
            "tt4":      row.get("tt4", ""),
            "loader":   row.get("loader", ""),
            "picks":    row.get("picks", 0),
            "pulls":    row.get("pulls", 0),
            "flags":    row.get("flags", []),
            "comments": row.get("comments", ""),
        })
    return compact_rows


def slim_summary_for_ai(board_summary):
    """Drop all *_details arrays — they duplicate row data and waste tokens."""
    detail_keys = {k for k in board_summary if k.endswith("_details")}
    return {k: v for k, v in board_summary.items() if k not in detail_keys}

def slim_pacing_for_ai(selected_day_pacing):
    """Drop the row-list arrays from pacing — they duplicate the actionable table.
    Keep all counts and the pacing verdict."""
    drop_keys = {"due_not_RTL_loads_first_10", "future_done_loads_first_10"}
    return {k: v for k, v in (selected_day_pacing or {}).items() if k not in drop_keys}


def actionable_rows_for_ai(board_rows):
    """
    Two buckets sent to the AI:
    1. actionable — loads needing attention (notable status, flagged, or blank).
       Full detail so the AI can recommend specific actions.
    2. completed — slim records (load, customer, appt time, day) so the AI can
       judge pacing (how many done vs remaining, are we ahead or behind schedule).
    Plain "Loaded" rows with no flag are omitted entirely — done and no action needed.
    """
    COMPLETED_STATUSES = {"Completed", "Complete"}
    SKIP_STATUSES = {"Loaded"}
    actionable = []
    completed = []
    for row in board_rows:
        status = (row.get("status") or "").strip()
        flags  = row.get("flags", [])
        is_blank       = not status
        is_completed   = status in COMPLETED_STATUSES
        is_skip        = status in SKIP_STATUSES and not flags
        notable_status = status and not is_completed and not is_skip

        if is_completed:
            completed.append({
                "day":      row.get("day", ""),
                "load":     row.get("load_number", ""),
                "customer": row.get("customer", ""),
                "time":     row.get("appt_time", ""),
                "status":   status,
            })
        elif notable_status or bool(flags) or is_blank:
            actionable.append({
                "day":      row.get("day", ""),
                "load":     row.get("load_number", ""),
                "customer": row.get("customer", ""),
                "time":     row.get("appt_time", ""),
                "door":     row.get("door", ""),
                "trailer":  row.get("trailer", ""),
                "status":   status or "Blank/Not Started",
                "type":     row.get("type", ""),
                "loader":   row.get("loader", ""),
                "pulls":    row.get("pulls", 0),
                "picks":    row.get("picks", 0),
                "flags":    flags,
                "comments": row.get("comments", ""),
            })
    return actionable, completed


def read_board_file_to_text(board_file):
    """
    Main entry point: reads outbound and inbound sheets, builds Python-verified
    summaries, and returns a compact JSON string for the AI prompt.
    Only scalar counts go in the summary. Only actionable rows are sent.
    """
    board_file.seek(0)
    file_name = board_file.name.lower()

    try:
        if file_name.endswith(".csv"):
            board_rows   = board_records_from_csv(board_file)
            inbound_rows = []
            today_totals = {"pulls_left_today": 0, "picks_left_today": 0}
        else:
            today_totals = read_board_today_totals_from_excel(board_file)
            board_rows = board_records_from_excel(board_file)
            board_file.seek(0)
            inbound_rows = board_records_from_inbound_sheet(board_file)

        board_summary   = build_python_board_summary(board_rows)
        inbound_summary = build_python_inbound_summary(inbound_rows)
        actionable_rows, completed_rows = actionable_rows_for_ai(board_rows)
        all_outbound_rows = compact_board_rows_for_ai(board_rows)

        payload = {
            "python_verified_outbound_summary": slim_summary_for_ai(board_summary),
            "python_verified_inbound_summary":  inbound_summary,
            "python_verified_today_totals":     today_totals,
            "actionable_outbound_rows":         actionable_rows,
            "completed_outbound_rows":          completed_rows,
            "all_outbound_rows":                all_outbound_rows,
            "instructions_for_ai": [
                "Use python_verified_outbound_summary for ALL outbound counts — do not recount from rows.",
                "Use python_verified_inbound_summary for ALL inbound counts.",
                "actionable_outbound_rows = loads needing attention (notable status, flags, or blank).",
                "completed_outbound_rows = slim records of finished loads. Use appt times to judge pacing: are completed loads early/on-time/late in the day relative to hours remaining?",
                "Outbound and inbound are separate — never mix their counts.",
                "All times use 24-hour clock.",
                "Blank status means load not yet started.",
                "Flags: LOAD-CHECK=yellow fill, TT4-NEEDED=blue fill, CANADIAN=red font.",
            ],
        }

        return json.dumps(payload, indent=2, ensure_ascii=False)

    except Exception as e:
        error_message = str(e)
        st.error(f"BOARD PARSER ERROR: {error_message}")
        st.exception(e)
        return json.dumps({
            "error": f"Could not read board file: {error_message}",
            "python_verified_outbound_summary": {},
            "python_verified_inbound_summary":  {},
            "actionable_outbound_rows": [],
        }, indent=2, ensure_ascii=False)


# ============================================================
#  SINGLE-CALL GROQ ANALYSIS
#  Python does ALL counting. AI gets only pre-computed summaries
#  + the compact load rows for context.
# ============================================================
def _rows_to_table(rows, columns):
    """Format a list of dicts as a compact pipe-delimited text table."""
    if not rows:
        return "(none)"
    header = " | ".join(columns)
    sep    = "-" * len(header)
    lines  = [header, sep]
    for r in rows:
        lines.append(" | ".join(str(r.get(c, "")).strip() for c in columns))
    return "\n".join(lines)


def rows_for_selected_day(rows, selected_day):
    """Return rows that belong only to the selected app day."""
    selected_day = str(selected_day or "").strip().lower()
    return [
        r for r in rows
        if str(r.get("day", "")).strip().lower() == selected_day
    ]


def rows_not_selected_day(rows, selected_day):
    """Return rows from days other than the selected app day."""
    selected_day = str(selected_day or "").strip().lower()
    return [
        r for r in rows
        if str(r.get("day", "")).strip().lower() != selected_day
    ]


def status_bucket_for_pacing(status):
    status = str(status or "").strip()
    return status if status else "Blank/Not Started"


def is_done_for_pacing(status):
    status_upper = str(status or "").strip().upper()
    return status_upper in {"COMPLETED", "COMPLETE", "LOADED"}


def board_minutes_for_pacing(value):
    text = normalize_board_time(value)
    if not text or not re.fullmatch(r"\d{1,2}:\d{2}", text):
        return None
    h, m = text.split(":")
    return int(h) * 60 + int(m)


def format_minutes_for_pacing(minutes):
    if minutes is None:
        return "unknown"
    minutes = int(minutes) % (24 * 60)
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


# Unpaid break inside a shift. hours_remaining is entered as WORKING hours
# (breaks excluded), but clock time keeps moving during the break. The break is
# modeled as a single block that is FINISHED by SHIFT_BREAK_DONE_BY (clock time):
#   - Before the break, the whole break is still ahead, so clock-time-left =
#     working-left + break. (1st shift: 8.5 working left = 07:00.)
#   - After the break, working time and clock time run 1:1, so zero working left
#     = the exact shift end (16:30).
# This keeps all three points exact -- 9.5 -> 06:00, 8.5 -> 07:00, 0 -> 16:30 --
# with clean whole/half-hour readings in between.
SHIFT_BREAK_HOURS = 1.0


def estimated_current_minutes_from_shift(shift, hours_remaining):
    """Estimate current clock time from WORKING hours remaining (breaks excluded)."""
    try:
        working_remaining = max(0.0, float(hours_remaining or 0))
    except Exception:
        return None

    break_minutes = int(round(SHIFT_BREAK_HOURS * 60))
    remaining_minutes = int(round(working_remaining * 60))

    shift_lower = str(shift or "").lower()
    if "1" in shift_lower:
        end_minutes = 16 * 60 + 30        # 1st shift ends 16:30
        break_done_by = 12 * 60           # breaks finished by 12:00
    else:
        end_minutes = 23 * 60 + 30        # 2nd shift estimate
        break_done_by = 19 * 60           # breaks finished by 19:00

    # Working hours still left at the moment the break is finished.
    working_left_at_break_done = end_minutes - break_done_by

    if remaining_minutes > working_left_at_break_done:
        # Before the break: the full break is still ahead of you.
        return max(0, end_minutes - remaining_minutes - break_minutes)
    # After the break: working time and clock time run 1:1 to the shift end.
    return max(0, end_minutes - remaining_minutes)


def build_selected_day_pacing(all_rows, selected_day, shift, hours_remaining):
    """
    Build a Python-computed selected-day pacing guardrail.
    This prevents the AI from mixing days or inventing pacing math.
    """
    selected_rows = rows_for_selected_day(all_rows, selected_day)
    current_minutes = estimated_current_minutes_from_shift(shift, hours_remaining)

    status_counts = {}
    completed_count = 0
    loaded_count = 0
    done_count = 0
    due_by_now = 0
    due_done = 0
    future_done = 0
    due_not_RTL_rows = []
    future_done_rows = []

    for row in selected_rows:
        status = status_bucket_for_pacing(row.get("status"))
        status_upper = status.upper()
        status_counts[status] = status_counts.get(status, 0) + 1

        if status_upper in {"COMPLETED", "COMPLETE"}:
            completed_count += 1
        if status_upper == "LOADED":
            loaded_count += 1

        done = is_done_for_pacing(status)
        if done:
            done_count += 1

        # Controlled for pacing = picking/pulling is finished.
        # Completed/Loaded are done; RTL is staged on the door (picks/pulls done,
        # only loading remains) -> NOT behind. Still Picking / Picking-Short /
        # Blank with a passed appt time IS behind.
        controlled_for_pacing = done or status_is_rtl(status)

        appt_minutes = board_minutes_for_pacing(row.get("time") or row.get("appt_time"))
        slim_row = {
            "load": row.get("load") or row.get("load_number", ""),
            "customer": row.get("customer", ""),
            "time": row.get("time") or row.get("appt_time", ""),
            "door": row.get("door", ""),
            "status": status,
        }

        if current_minutes is not None and appt_minutes is not None:
            if appt_minutes <= current_minutes:
                due_by_now += 1
                if controlled_for_pacing:
                    due_done += 1
                else:
                    due_not_RTL_rows.append(slim_row)
            elif controlled_for_pacing:
                future_done += 1
                future_done_rows.append(slim_row)

    due_not_RTL = max(0, due_by_now - due_done)
    actionable_or_not_done = max(0, len(selected_rows) - done_count)

    if due_not_RTL > 0:
        pacing = "BEHIND"
        reason = f"{due_not_RTL} selected-day load(s) due by now are still picking/pulling."
    elif future_done > 0:
        pacing = "AHEAD"
        reason = f"{future_done} future selected-day load(s) are already done."
    else:
        pacing = "ON TRACK"
        reason = "No selected-day loads due by the estimated current time are unfinished."

    return {
        "selected_day": selected_day,
        "estimated_current_time": format_minutes_for_pacing(current_minutes),
        "selected_day_total_loads": len(selected_rows),
        "selected_day_status_counts": status_counts,
        "completed_count": completed_count,
        "loaded_count": loaded_count,
        "done_count_completed_plus_loaded": done_count,
        "actionable_or_not_done_count": actionable_or_not_done,
        "due_by_now": due_by_now,
        "due_done": due_done,
        "due_not_RTL": due_not_RTL,
        "future_done": future_done,
        "pacing": pacing,
        "reason": reason,
        "due_not_RTL_loads_first_10": due_not_RTL_rows[:10],
        "future_done_loads_first_10": future_done_rows[:10],
    }


# ============================================================
#  THROUGHPUT-OPTIMAL ALLOCATION  (paste near compute_python_shift_goal_preview)
#  Goal: control (pick + pull + stage) as many of TODAY's loads as possible
#  before 2nd shift takes over at shift end. Picking is usually the wall, but
#  Tasking has a hard floor of 4 (replenishment + putaway) before any full-pallet
#  taskers are added, and Loading must keep up or staged loads pile up.
# ============================================================

# Operating constants now come from dc_config so the staffing report, the shift
# closeout, and the README standard can never drift apart. Aliased to module-level
# names here so every existing reference below keeps working unchanged.

PICK_RATE = dc_config.PICK_RATE              # cases/hr/person
PULL_RATE = dc_config.PULL_RATE              # full pallets/hr/person
LOAD_RATE = dc_config.LOAD_RATE              # trailers/hr/person
UNLOAD_RATE = dc_config.UNLOAD_RATE          # inbound pallets/hr/person
TASK_FLOOR = dc_config.TASK_FLOOR            # always-on replenishment + putaway, before full-pallet pulls
MIN_UNLOADERS = dc_config.MIN_UNLOADERS      # reserved inbound crew before the split
MIN_RECEIVERS = dc_config.MIN_RECEIVERS      # reserved inbound crew before the split
LOAD_TARGET_SHARE = dc_config.LOAD_TARGET_SHARE  # 1st-shift loading target = share of selected-day outbound loads


def status_is_completed_or_loaded(status):
    """Physically completed/loaded now. These count toward the loading goal already banked."""
    status_upper = str(status or "").strip().upper()
    return status_upper in {"COMPLETED", "COMPLETE", "LOADED"}


def status_is_rtl(status):
    """Ready-to-load: controlled for appointment cutoff, but still needs loader work."""
    status_upper = str(status or "").strip().upper()
    return status_upper in {"RTL", "READY TO LOAD"}

def status_is_controlled_appointment(status):
    status_upper = str(status or "").strip().upper()
    return (
        status_is_completed_or_loaded(status)
        or status_is_rtl(status)
        or status_upper in {"R/S", "READY/SHORT", "LOADED SHORT"}
    )


def status_is_excluded_from_new_control(status):
    """
    Rows excluded from new labor-control capacity.
    Loaded Short is a service issue, not a normal capacity wave.
    R/S is controlled for appointment cutoff and therefore handled by status_is_controlled_appointment().
    """
    status_upper = str(status or "").strip().upper()
    return status_upper in {"LOADED SHORT"}


def compute_throughput_optimal_allocation(
    picks_left, pulls_left, total_loads, hours_remaining, present_total,
    min_unload=MIN_UNLOADERS, min_receive=MIN_RECEIVERS, task_floor=TASK_FLOOR,
    completed_or_loaded_now=0,
):
    """
    Distribute present workers to MAXIMIZE what can actually be controlled by shift end,
    without over-assigning loaders ahead of the freight Picking/Tasking can create.

    Logic:
    1. Protect Unloading and Receiving minimums.
    2. Protect Tasking floor first for replenishment + putaway.
    3. Test every possible split of the remaining crew across Picking, extra Tasking
       for full-pallet pulls, and Loading.
    4. Pick the split that controls the most loads by shift end.
    5. If two splits control the same number of loads, prefer the one with less idle
       loading capacity and more Picking/Tasking feed for 2nd shift.

    Returns a dict: Unloading, Receiving, Picking, Tasking, Loading.
    """
    try:
        hrs = max(0.0, float(hours_remaining or 0))
    except Exception:
        hrs = 0.0

    try:
        present_total = int(present_total or 0)
    except Exception:
        present_total = 0

    try:
        total_loads = max(0, int(round(float(total_loads or 0))))
    except Exception:
        total_loads = 0

    # Loading is not optimized against the entire day board.
    # New rule: the loading goal is everything already completed/loaded now
    # PLUS 52% of the selected-day loads. Completed/loaded loads are already banked,
    # so the loader allocation is aimed at the remaining loading goal.
    try:
        completed_or_loaded_now = max(0, int(round(float(completed_or_loaded_now or 0))))
    except Exception:
        completed_or_loaded_now = 0

    base_loading_goal = int(round(total_loads * LOAD_TARGET_SHARE)) if total_loads > 0 else 0
    if total_loads > 0 and base_loading_goal <= 0:
        base_loading_goal = 1

    # Loading goal is a flat 52% of selected-day loads. Completed loads are no longer
    # added to inflate the goal, nor subtracted from it.
    loading_target_loads = min(total_loads, base_loading_goal)
    remaining_loading_goal = loading_target_loads

    picks_left = max(0.0, float(picks_left or 0))
    pulls_left = max(0.0, float(pulls_left or 0))

    # Protect the fixed minimums, but never allocate more people than are present.
    alloc = {"Unloading": 0, "Receiving": 0, "Picking": 0, "Tasking": 0, "Loading": 0}
    remaining = present_total

    alloc["Unloading"] = min(int(min_unload), remaining)
    remaining -= alloc["Unloading"]

    alloc["Receiving"] = min(int(min_receive), remaining)
    remaining -= alloc["Receiving"]

    alloc["Tasking"] = min(int(task_floor), remaining)
    remaining -= alloc["Tasking"]

    if remaining <= 0 or hrs <= 0:
        return alloc

    # If there are outbound loads, keep at least one loader. More loaders must be
    # earned by enough Picking/Tasking feed; otherwise they are idle capacity.
    min_loading = 1 if remaining_loading_goal > 0 else 0
    if min_loading and remaining > 0:
        alloc["Loading"] = 1
        remaining -= 1

    best = None
    best_score = None

    def _loads_feedable_by_pick(pickers):
        if total_loads <= 0:
            return 0.0
        if picks_left <= 0:
            return float(total_loads)
        return min(float(total_loads), (pickers * PICK_RATE * hrs / picks_left) * total_loads)

    def _loads_feedable_by_pull(extra_pull_taskers):
        if total_loads <= 0:
            return 0.0
        if pulls_left <= 0:
            return float(total_loads)
        return min(float(total_loads), (extra_pull_taskers * PULL_RATE * hrs / pulls_left) * total_loads)

    # Brute force all integer allocations of the flexible crew. This is small and
    # reliable, and it prevents the old issue where Loading took too many people
    # just because total day loads were high.
    for add_picking in range(remaining + 1):
        for add_loading in range(remaining - add_picking + 1):
            add_pull_extra = remaining - add_picking - add_loading

            pickers = alloc["Picking"] + add_picking
            loaders = alloc["Loading"] + add_loading
            pull_extra = add_pull_extra
            taskers = alloc["Tasking"] + pull_extra

            pick_feed = _loads_feedable_by_pick(pickers)
            pull_feed = _loads_feedable_by_pull(pull_extra)
            freight_feed = min(pick_feed, pull_feed, float(total_loads))
            # Loading target is completed/loaded now + 52% of selected-day loads.
            # The completed/loaded portion is already banked; the flexible loader
            # allocation only covers the remaining loading goal.
            loading_capacity = min(float(remaining_loading_goal), loaders * LOAD_RATE * hrs)
            controlled = min(freight_feed, loading_capacity)

            # Do not reward loaders that cannot be fed by Picking/Tasking.
            idle_loading_capacity = max(0.0, loading_capacity - freight_feed)
            unmet_feed_capacity = max(0.0, freight_feed - loading_capacity)

            # Objective order:
            # 1) control the most loads by shift end
            # 2) avoid idle loaders that Picking/Tasking cannot feed
            # 3) keep more ready-freight creation capacity for 2nd shift
            # 4) prefer fewer loaders when the same work is controlled
            score = (
                round(controlled, 4),
                -round(idle_loading_capacity, 4),
                round(freight_feed, 4),
                -loaders,
                -abs(pick_feed - pull_feed),
                -round(unmet_feed_capacity, 4),
            )

            candidate = {
                "Unloading": alloc["Unloading"],
                "Receiving": alloc["Receiving"],
                "Picking": pickers,
                "Tasking": taskers,
                "Loading": loaders,
            }

            if best_score is None or score > best_score:
                best_score = score
                best = candidate

    return best or alloc

def appointment_controlled_by_allocation(
    board_text, day, shift, hours_remaining, summary_table_or_counts,
    task_floor=TASK_FLOOR,
):
    """
    Given an allocation, compute how many of TODAY's loads it can control
    (pick + pull + stage) by shift end, and the appointment time of the last
    load in that controlled wave.

    Logic:
    - Completed / Loaded / RTL already count as controlled for appointment cutoff.
    - RTL still needs loader work in the report, but it is already controlled for cutoff.
    - R/S counts as controlled for appointment cutoff, but it does not create new pick/stage capacity demand.
    - Loading goal = Completed/Loaded now + 52% of selected-day loads.
    """
    blank = {
        "loads_controlled": 0,
        "selected_day_loads": 0,
        "already_controlled_loads": 0,
        "completed_or_loaded_now": 0,
        "rtl_controlled_loads": 0,
        "excluded_from_new_control_loads": 0,
        "cutoff": "n/a",
        "binding": "Unknown",
        "pick_frac": 0.0, "pull_frac": 0.0, "load_frac": 0.0,
        "loading_target_loads": 0,
        "base_loading_goal_loads": 0,
        "additional_loads_controlled": 0,
        "note": "No board / allocation data.",
    }
    if not board_text:
        return blank
    try:
        payload = json.loads(board_text or "{}")
    except Exception:
        return blank

    all_rows = payload.get("all_outbound_rows", []) or []
    today_totals = payload.get("python_verified_today_totals", {}) or {}
    selected_rows = rows_for_selected_day(all_rows, day)
    if not selected_rows:
        return blank

    def _count(key):
        try:
            if hasattr(summary_table_or_counts, "loc") and key in summary_table_or_counts.index:
                return int(summary_table_or_counts.loc[key, "Assigned"])
        except Exception:
            pass
        try:
            return int((summary_table_or_counts or {}).get(key, 0))
        except Exception:
            return 0

    pickers = _count("Picking")
    taskers = _count("Tasking")
    loaders = _count("Loading")

    try:
        hrs = max(0.0, float(hours_remaining or 0))
    except Exception:
        hrs = 0.0

    picks_left = pdf_number(today_totals.get("picks_left_today", 0))
    pulls_left = pdf_number(today_totals.get("pulls_left_today", 0))
    total_loads = len(selected_rows)

    completed_or_loaded_now = sum(
        1 for r in selected_rows if status_is_completed_or_loaded(r.get("status"))
    )
    rtl_controlled = sum(
        1 for r in selected_rows if status_is_rtl(r.get("status"))
    )
    # R/S loads: product not available in the DC. They count toward appointment cutoff
    # (the pick wave is done or moot) but they are NOT DC-controllable — no loader or
    # picker action will ship them until product arrives. Surfaced separately so the
    # report never silently inflates the "controlled" number with unshippable loads.
    rs_count = sum(
        1 for r in selected_rows
        if str(r.get("status") or "").strip().upper() in {"R/S", "READY/SHORT"}
    )
    already_controlled = sum(
        1 for r in selected_rows if status_is_controlled_appointment(r.get("status"))
    )
    excluded_from_new_control = sum(
        1 for r in selected_rows
        if (not status_is_controlled_appointment(r.get("status"))
            and status_is_excluded_from_new_control(r.get("status")))
    )

    remaining_candidate_loads = max(
        0,
        total_loads - already_controlled - excluded_from_new_control,
    )

    # Loading goal = Completed/Loaded already banked + 52% of total selected-day loads.
    # RTL is NOT included in completed_or_loaded_now because it still needs loader work.
    base_loading_goal_loads = int(round(total_loads * LOAD_TARGET_SHARE)) if total_loads > 0 else 0
    if total_loads > 0 and base_loading_goal_loads <= 0:
        base_loading_goal_loads = 1
    # Loading goal is a flat 52% of selected-day loads (completed not added or subtracted).
    loading_target_loads = min(total_loads, base_loading_goal_loads)
    remaining_loading_goal = loading_target_loads

    pull_workers = max(0, taskers - task_floor)   # only above-floor taskers do pulls

    pick_frac = min(1.0, (pickers * PICK_RATE * hrs) / picks_left) if picks_left > 0 else 1.0
    pull_frac = min(1.0, (pull_workers * PULL_RATE * hrs) / pulls_left) if pulls_left > 0 else 1.0

    loading_capacity_left = loaders * LOAD_RATE * hrs
    load_frac = (
        min(1.0, loading_capacity_left / loading_target_loads)
        if loading_target_loads > 0 else 1.0
    )

    # Convert each stream into ADDITIONAL load ceilings.
    # Already-controlled loads are banked and never removed by a later bottleneck.
    pick_supported_additional = int(round(pick_frac * remaining_candidate_loads))
    pull_supported_additional = int(round(pull_frac * remaining_candidate_loads))
    load_supported_additional = int(round(min(loading_capacity_left, remaining_loading_goal, remaining_candidate_loads)))

    additional_controlled = max(
        0,
        min(
            remaining_candidate_loads,
            pick_supported_additional,
            pull_supported_additional,
            load_supported_additional,
        )
    )

    loads_controlled = min(total_loads, already_controlled + additional_controlled)

    support_map = {
        "Picking": already_controlled + pick_supported_additional,
        "Tasking/Pulls": already_controlled + pull_supported_additional,
        "Loading": already_controlled + load_supported_additional,
    }
    if remaining_candidate_loads <= 0 or loads_controlled >= total_loads:
        binding_name = "None (controls full day)"
    else:
        binding_name = min(support_map, key=support_map.get)

   # Cutoff = appt time of the last load in the controlled wave (sorted by appt time).
    timed = []
    for r in selected_rows:
        m = board_minutes_for_pacing(r.get("time") or r.get("appt_time"))
        if m is not None:
            timed.append(m)
    timed.sort()

    # When the allocation controls every selected-day load, report the LATEST
    # appointment time of the day and label it as fully controlled.
    all_today_controlled = total_loads > 0 and loads_controlled >= total_loads

    if all_today_controlled:
        if timed:
            cutoff = f"{format_minutes_for_pacing(timed[-1])} (all loads for today controlled)"
        else:
            cutoff = "all loads for today controlled"
    elif not timed:
        cutoff = "no appt times on board"
    elif loads_controlled <= 0:
        cutoff = "none — before first appt"
    elif loads_controlled >= len(timed):
        cutoff = f"{format_minutes_for_pacing(timed[-1])} (all today)"
    else:
        cutoff = format_minutes_for_pacing(timed[loads_controlled - 1])

    return {
        "loads_controlled": loads_controlled,
        "selected_day_loads": total_loads,
        "already_controlled_loads": already_controlled,
        "completed_or_loaded_now": completed_or_loaded_now,
        "rtl_controlled_loads": rtl_controlled,
        "rs_count": rs_count,
        "excluded_from_new_control_loads": excluded_from_new_control,
        "cutoff": cutoff,
        "binding": binding_name,
        "pick_frac": round(pick_frac, 3),
        "pull_frac": round(pull_frac, 3),
        "load_frac": round(load_frac, 3),
        "loading_target_loads": loading_target_loads,
        "base_loading_goal_loads": base_loading_goal_loads,
        "additional_loads_controlled": additional_controlled,
        "note": (
            f"Already controlled: Completed/Loaded/RTL/R/S ({already_controlled}). "
            f"RTL = picking done, staged on door, still needs loader work. "
            + (f"R/S ({rs_count} load(s)): picking is done, load is waiting on product not yet in the DC. "
               f"Counted as controlled since no pick work remains; status expected to resolve before departure. "
               if rs_count > 0 else "")
            + f"Loading goal = 52% of selected-day loads = {loading_target_loads}."
        ),
    }


def render_allocation_controls_preview(label, controlled):
    """Streamlit display of what one allocation controls by shift end."""
    if not controlled:
        return
    st.markdown(f"**{label}**")
    a, b, c = st.columns(3)
    a.metric("Loads controlled by shift end",
             f"{controlled['loads_controlled']} / {controlled['selected_day_loads']}")
    b.metric("Controlled through appt", controlled["cutoff"])
    c.metric("Bottleneck", controlled["binding"])
    rs_count = int(controlled.get("rs_count", 0) or 0)
    rs_note = (
        f" Note: {rs_count} of those are currently R/S (Ready/Short) — "
        f"picking is done but product is not yet available in the DC to complete the load. "
        f"These are counted as controlled because no further pick work is needed, "
        f"but their status should resolve before departure as product arrives."
        if rs_count > 0 else ""
    )
    st.caption(
        f"Coverage — picking {int(controlled['pick_frac']*100)}%, "
        f"pulls {int(controlled['pull_frac']*100)}%, loading {int(controlled['load_frac']*100)}%. "
        f"Already controlled: Completed/Loaded/RTL/R/S ({controlled.get('already_controlled_loads', 0)}). "
        f"Loading goal = 52% of selected-day loads = {controlled.get('loading_target_loads', 0)}."
        + rs_note
    )


# ============================================================
#  PYTHON SHIFT GOAL / APPOINTMENT TARGET PREVIEW
#  Source of truth for pre-report decision, AI goal, and PDF goal.
# ============================================================

def shift_end_label(shift):
    """Return the planned shift-end label used in the Python goal."""
    return "16:30" if "1" in str(shift).lower() else "23:30"


def build_summary_table_from_counts(needed, assigned_counts):
    """Build the same Staffing Summary format from a manual/actual allocation."""
    task_order = ["Unloading", "Receiving", "Picking", "Tasking", "Loading"]
    needed_series = pd.Series({t: int(needed.get(t, 0)) for t in task_order}, name="Needed")
    assigned_series = pd.Series({t: int(assigned_counts.get(t, 0)) for t in task_order}, name="Assigned")
    summary_table = pd.concat([needed_series, assigned_series], axis=1).fillna(0)
    summary_table["Needed"] = summary_table["Needed"].astype(int)
    summary_table["Assigned"] = summary_table["Assigned"].astype(int)
    summary_table["Difference"] = summary_table["Assigned"] - summary_table["Needed"]
    summary_table["Status"] = summary_table["Difference"].apply(
        lambda x: "Good" if x == 0 else ("Overstaffed" if x > 0 else "Understaffed")
    )
    return summary_table


def compute_python_shift_goal_preview(board_text, day, shift, hours_remaining, summary_table):
    """
    Python-only appointment target preview.

    This uses appointment_controlled_by_allocation() as the source of truth for:
    cutoff, bottleneck, loads controlled, Completed/Loaded/RTL already banked,
    and the loading goal.
    """
    empty = {
        "goal": "Upload a board and compute allocation to generate the appointment target.",
        "confidence": "UNKNOWN",
        "main_constraint": "Unknown",
        "target_cutoff": "",
        "target_load_count": 0,
        "reason": "Board data was not available.",
        "suggested_adjustment": "Upload the board, select present workers, and compute allocation.",
        "pick_capacity": 0,
        "pull_capacity": 0,
        "loading_capacity": 0,
        "picks_left": 0,
        "pulls_left": 0,
        "loads_to_stage_for_target": 0,
        "loads_controlled": 0,
        "selected_day_loads": 0,
        "already_controlled_loads": 0,
        "completed_or_loaded_now": 0,
        "rtl_controlled_loads": 0,
        "loading_target_loads": 0,
        "base_loading_goal_loads": 0,
        "controlled_through_appt": "",
        "pick_coverage_pct": 0,
        "pull_coverage_pct": 0,
        "load_coverage_pct": 0,
    }

    if not board_text:
        return empty

    try:
        payload = json.loads(board_text or "{}")
    except Exception:
        return empty

    selected_rows = rows_for_selected_day(payload.get("all_outbound_rows", []) or [], day)
    today_totals = payload.get("python_verified_today_totals", {}) or {}

    if not selected_rows:
        empty["goal"] = f"No selected-day board rows were found for {day}."
        empty["reason"] = "The board parser did not find rows matching the selected day."
        return empty

    controlled = appointment_controlled_by_allocation(
        board_text=board_text,
        day=day,
        shift=shift,
        hours_remaining=hours_remaining,
        summary_table_or_counts=summary_table,
    )

    try:
        hrs = max(0.0, float(hours_remaining or 0))
    except Exception:
        hrs = 0.0

    def _assigned(task):
        try:
            if summary_table is not None and task in summary_table.index:
                return int(summary_table.loc[task, "Assigned"])
        except Exception:
            pass
        return 0

    pickers = _assigned("Picking")
    taskers = _assigned("Tasking")
    loaders = _assigned("Loading")

    picks_left = pdf_number(today_totals.get("picks_left_today", 0))
    pulls_left = pdf_number(today_totals.get("pulls_left_today", 0))
    pull_workers = max(0, taskers - TASK_FLOOR)

    pick_capacity = pickers * PICK_RATE * hrs
    pull_capacity = pull_workers * PULL_RATE * hrs
    loading_capacity = loaders * LOAD_RATE * hrs

    total_loads = int(controlled.get("selected_day_loads", 0) or 0)
    loads_controlled = int(controlled.get("loads_controlled", 0) or 0)
    already_controlled = int(controlled.get("already_controlled_loads", 0) or 0)
    completed_or_loaded_now = int(controlled.get("completed_or_loaded_now", 0) or 0)
    rtl_controlled = int(controlled.get("rtl_controlled_loads", 0) or 0)
    rs_count = int(controlled.get("rs_count", 0) or 0)
    loading_target_loads = int(controlled.get("loading_target_loads", 0) or 0)
    base_loading_goal_loads = int(controlled.get("base_loading_goal_loads", 0) or 0)
    cutoff = str(controlled.get("cutoff", "n/a"))
    bottleneck = str(controlled.get("binding", "Unknown"))

    if total_loads <= 0:
        return empty

    confidence = "YES" if loads_controlled > 0 else "NO"

    goal = (
        f"Pick & stage every load through appointment {cutoff} "
        f"({loads_controlled} of {total_loads} selected-day loads) by shift end {shift_end_label(shift)}."
    )

    rs_note = (
        f" Note: {rs_count} of those are currently R/S — picking is done but product is not yet "
        f"in the DC. Counted as controlled since no pick work remains; expected to resolve before departure."
        if rs_count > 0 else ""
    )

    if loads_controlled >= total_loads:
        reason = (
            f"This allocation can control the full selected-day board by shift end: "
            f"{loads_controlled} of {total_loads} load(s).{rs_note} "
            f"Already controlled now: {already_controlled} "
            f"(Completed/Loaded: {completed_or_loaded_now}, RTL: {rtl_controlled})."
        )
        suggested_adjustment = "Protect current pickers, taskers, and loaders until the full board is controlled."
    elif loads_controlled > 0:
        reason = (
            f"This allocation controls {loads_controlled} of {total_loads} selected-day load(s) by shift end.{rs_note} "
            f"Already controlled now: {already_controlled} "
            f"(Completed/Loaded: {completed_or_loaded_now}, RTL: {rtl_controlled}). "
            f"The limiting stream is {bottleneck}."
        )
        if "Pick" in bottleneck:
            suggested_adjustment = "Add/protect Picking first if you want to push the cutoff later."
        elif "Task" in bottleneck or "Pull" in bottleneck:
            suggested_adjustment = "Add/protect Tasking above the floor if you want to push the cutoff later."
        elif "Load" in bottleneck:
            suggested_adjustment = "Add/protect Loading if you want to push the cutoff later."
        else:
            suggested_adjustment = "Protect current allocation until the controlled wave is complete."
    else:
        reason = f"This allocation does not control the first selected-day appointment wave by shift end. Bottleneck: {bottleneck}.{rs_note}"
        suggested_adjustment = "Add labor to the bottleneck before generating the final report."

    return {
        "goal": goal,
        "confidence": confidence,
        "main_constraint": bottleneck,
        "target_cutoff": cutoff,
        "target_load_count": loads_controlled,
        "reason": reason,
        "suggested_adjustment": suggested_adjustment,
        "pick_capacity": round(pick_capacity),
        "pull_capacity": round(pull_capacity),
        "loading_capacity": round(loading_capacity, 1),
        "picks_left": picks_left,
        "pulls_left": pulls_left,
        "loads_to_stage_for_target": loads_controlled,
        "estimated_picks_for_target": round(picks_left * float(controlled.get("pick_frac", 0) or 0)),
        "estimated_pulls_for_target": round(pulls_left * float(controlled.get("pull_frac", 0) or 0)),
        "loads_controlled": loads_controlled,
        "selected_day_loads": total_loads,
        "already_controlled_loads": already_controlled,
        "completed_or_loaded_now": completed_or_loaded_now,
        "rtl_controlled_loads": rtl_controlled,
        "rs_count": rs_count,
        "loading_target_loads": loading_target_loads,
        "base_loading_goal_loads": base_loading_goal_loads,
        "controlled_through_appt": cutoff,
        "pick_coverage_pct": int(float(controlled.get("pick_frac", 0) or 0) * 100),
        "pull_coverage_pct": int(float(controlled.get("pull_frac", 0) or 0) * 100),
        "load_coverage_pct": int(float(controlled.get("load_frac", 0) or 0) * 100),
    }


def render_python_shift_goal_preview(preview):
    """Streamlit display for the pre-report appointment target."""
    if not preview:
        return

    st.markdown("---")
    st.subheader("Pre-Report Appointment Target Preview")

    confidence = str(preview.get("confidence", "UNKNOWN")).upper()
    if confidence == "YES":
        st.success(preview.get("goal", ""))
    elif confidence == "RISKY":
        st.warning(preview.get("goal", ""))
    elif confidence == "NO":
        st.error(preview.get("goal", ""))
    else:
        st.info(preview.get("goal", ""))

    c1, c2, c3 = st.columns(3)
    c1.metric("Can this allocation hit target?", confidence)
    c2.metric("Main constraint", preview.get("main_constraint", "Unknown"))
    c3.metric("Target cutoff", preview.get("target_cutoff", ""))

    c4, c5, c6 = st.columns(3)
    c4.metric("Picking capacity left", f"{preview.get('pick_capacity', 0):,} cases")
    c5.metric("Tasking / pull capacity", f"{preview.get('pull_capacity', 0):,} pallets")
    c6.metric("Loading capacity left", f"{preview.get('loading_capacity', 0)} loads")

    st.caption(
        f"Picks left: {preview.get('picks_left', 0):,} | "
        f"Pulls left: {preview.get('pulls_left', 0):,} | "
        f"Loads controlled in target wave: {preview.get('loads_to_stage_for_target', 0)} | "
        f"Already controlled now: {preview.get('already_controlled_loads', 0)} | "
        f"RTL controlled/not loaded: {preview.get('rtl_controlled_loads', 0)} | "
        + (f"R/S (picking done, waiting on product): {preview.get('rs_count', 0)} | "
           if preview.get('rs_count', 0) > 0 else "")
        + f"Loading goal: {preview.get('loading_target_loads', 0)} loads "
        f"(52% of day)"
    )
    st.write(f"**Why:** {preview.get('reason', '')}")
    st.write(f"**Suggested decision:** {preview.get('suggested_adjustment', '')}")


def build_second_shift_handoff_forecast(board_text, day, shift, hours_remaining, summary_table_or_counts):
    """Forecast what this allocation leaves for 2nd shift using the same controlled-through logic."""
    blank = {
        "loads_controlled": 0,
        "selected_day_loads": 0,
        "loads_left_for_next_shift": 0,
        "controlled_through_appt": "n/a",
        "bottleneck": "Unknown",
        "picks_left_now": 0,
        "pulls_left_now": 0,
        "estimated_picks_left_at_handoff": 0,
        "estimated_pulls_left_at_handoff": 0,
        "handoff_message": "No board / allocation data available for handoff forecast.",
    }
    if not board_text:
        return blank
    try:
        payload = json.loads(board_text or "{}")
    except Exception:
        return blank

    today_totals = payload.get("python_verified_today_totals", {}) or {}
    controlled = appointment_controlled_by_allocation(
        board_text=board_text,
        day=day,
        shift=shift,
        hours_remaining=hours_remaining,
        summary_table_or_counts=summary_table_or_counts,
    )

    total = int(controlled.get("selected_day_loads", 0) or 0)
    done = int(controlled.get("loads_controlled", 0) or 0)
    left = max(0, total - done)
    picks_left = pdf_number(today_totals.get("picks_left_today", 0))
    pulls_left = pdf_number(today_totals.get("pulls_left_today", 0))

    pick_frac = max(0.0, min(1.0, float(controlled.get("pick_frac", 0) or 0)))
    pull_frac = max(0.0, min(1.0, float(controlled.get("pull_frac", 0) or 0)))

    est_picks_left = int(round(picks_left * (1.0 - pick_frac)))
    est_pulls_left = int(round(pulls_left * (1.0 - pull_frac)))
    cutoff = controlled.get("cutoff", "n/a")
    bottleneck = controlled.get("binding", "Unknown")

    if total <= 0:
        message = "No selected-day outbound loads found for the handoff forecast."
    elif left <= 0:
        message = "This allocation is projected to control the full selected-day board before 2nd shift."
    else:
        message = (
            f"2nd shift should expect about {left} selected-day load(s) left after the controlled wave. "
            f"Main carryover constraint: {bottleneck}."
        )

    return {
        "loads_controlled": done,
        "selected_day_loads": total,
        "loads_left_for_next_shift": left,
        "controlled_through_appt": cutoff,
        "bottleneck": bottleneck,
        "picks_left_now": picks_left,
        "pulls_left_now": pulls_left,
        "estimated_picks_left_at_handoff": est_picks_left,
        "estimated_pulls_left_at_handoff": est_pulls_left,
        "handoff_message": message,
    }

def analyze_board_with_groq(
    board_text, day, shift, total_cases, hours_remaining, total_outbound_loads,
    crossroads_open, deer_creek_open, msb_open, needed, summary_table,
    cases_to_pick, inbound_pallets, notes, oc_alert_text=None,
    recommended_allocation=None, deviation_reason=None, python_shift_goal_preview=None,
):
    client = get_groq_client()
    if client is None:
        return (
            "Board analysis could not be completed because GROQ_API_KEY is missing. "
            "Add GROQ_API_KEY in Streamlit Cloud Secrets."
        )

    try:
        board_payload   = json.loads(board_text)
        py_summary      = board_payload.get("python_verified_outbound_summary", {})
        py_inbound      = board_payload.get("python_verified_inbound_summary", {})
        py_today_totals = board_payload.get("python_verified_today_totals", {})
        actionable_rows = board_payload.get("actionable_outbound_rows", [])
        completed_rows  = board_payload.get("completed_outbound_rows", [])
        all_rows        = board_payload.get("all_outbound_rows", [])
    except Exception:
        py_summary      = {}
        py_inbound      = {}
        py_today_totals = {}
        actionable_rows = []
        completed_rows  = []
        all_rows        = []

    if not all_rows:
        all_rows = actionable_rows + completed_rows

    staffing_lines = []
    for task, row in summary_table.iterrows():
        staffing_lines.append(
            f"  {task}: need {int(row['Needed'])}, have {int(row['Assigned'])}, "
            f"gap {int(row['Difference'])} ({row['Status']})"
        )
    staffing_summary = "\n".join(staffing_lines)

    allocation_block = ""
    if recommended_allocation:
        label_map = {
            "Picking": "Pickers",
            "Tasking": "Taskers",
            "Loading": "Loaders",
            "Unloading": "Unloaders",
            "Receiving": "Receivers",
        }
        rec_line = ", ".join(
            f"{label_map.get(t, t)} {int(n)}" for t, n in recommended_allocation.items()
        )
        allocation_block = (
            "ALLOCATION MODE - SUPERVISOR OVERRIDE:\n"
            "The 'have' numbers in the STAFFING table are the ACTUAL crew the supervisor has on each "
            "function right now - NOT the tool's recommendation. Treat them as fact.\n"
            f"Tool's recommended placement (for comparison only): {rec_line}.\n"
        )
        if deviation_reason and deviation_reason.strip():
            allocation_block += (
                f"Supervisor's stated reason for running it differently: {deviation_reason.strip()}\n"
            )
        allocation_block += (
            "When recommending moves: compare ACTUAL vs NEEDED to find the real gaps, and pull only from "
            "functions that are OVERSTAFFED vs need (positive gap). Respect the supervisor's stated reason - "
            "if a gap is intentionally covered by that reason, acknowledge it instead of flagging it as a problem.\n"
        )

    plants_open = [
        p for p, s in [("Crossroads", crossroads_open), ("Deer Creek", deer_creek_open), ("MSB", msb_open)]
        if s == "YES"
    ]

    oc_section = f"\n{oc_alert_text}\n" if oc_alert_text else ""

    selected_day_pacing = build_selected_day_pacing(all_rows, day, shift, hours_remaining)
    selected_status_text = json.dumps(selected_day_pacing.get("selected_day_status_counts", {}), ensure_ascii=False)

    loads_by_day = py_summary.get("loads_by_day", {})
    day_str = ", ".join(f"{d}:{n}" for d, n in loads_by_day.items())
    verified_counts = (
        f"SELECTED DAY COUNTS ONLY (Python — use this first for today's analysis):\n"
        f"Day:{selected_day_pacing.get('selected_day')}  "
        f"Total:{selected_day_pacing.get('selected_day_total_loads',0)}  "
        f"Completed:{selected_day_pacing.get('completed_count',0)}  "
        f"Loaded:{selected_day_pacing.get('loaded_count',0)}  "
        f"DoneCompletedPlusLoaded:{selected_day_pacing.get('done_count_completed_plus_loaded',0)}  "
        f"ActionableOrNotDone:{selected_day_pacing.get('actionable_or_not_done_count',0)}\n"
        f"DueByNow:{selected_day_pacing.get('due_by_now',0)}  "
        f"DueDone:{selected_day_pacing.get('due_done',0)}  "
        f"DueNotDone:{selected_day_pacing.get('due_not_RTL',0)}  "
        f"FutureDone:{selected_day_pacing.get('future_done',0)}  "
        f"Pacing:{selected_day_pacing.get('pacing','UNKNOWN')}  "
        f"CurrentTimeEstimate:{selected_day_pacing.get('estimated_current_time','unknown')}\n"
        f"Selected day status counts: {selected_status_text}\n"
        f"ALL BOARD COUNTS BY DAY / CONTEXT ONLY (do not use this as today's selected-day total):\n"
        f"Total:{py_summary.get('loads_read_from_board',0)}  "
        f"Late:{py_summary.get('late_loads',0)}  "
        f"RTL:{py_summary.get('rtl_loads',0)}  "
        f"R/S:{py_summary.get('rs_loads',0)}  "
        f"Picking:{py_summary.get('picking_loads',0)}  "
        f"Pick/Short:{py_summary.get('picking_short_loads',0)}  "
        f"LoadedShort:{py_summary.get('loaded_short_loads',0)}  "
        f"Completed:{py_summary.get('completed_loads',0)}  "
        f"Blank:{py_summary.get('blank_or_not_started_loads',0)}\n"
        f"Live:{py_summary.get('live_loads',0)}  "
        f"CPU:{py_summary.get('cpu_loads',0)}  "
        f"Canadian:{py_summary.get('canadian_loads',0)}  "
        f"TT4:{py_summary.get('tt4_needed_loads',0)}  "
        f"LoadCheck:{py_summary.get('load_check_loads',0)}  "
        f"LoaderAssigned:{py_summary.get('loads_with_loader_assigned',0)}  "
        f"MissingLoader:{py_summary.get('loads_missing_loader',0)}\n"
        f"By day: {day_str}"
    )

    ib_day_str = ", ".join(f"{d}:{n}" for d, n in py_inbound.get("loads_by_day", {}).items())
    verified_inbound = (
        f"VERIFIED INBOUND COUNTS (Python — do not recount):\n"
        f"Total:{py_inbound.get('loads_read_from_inbound',0)}  "
        f"Live:{py_inbound.get('live_loads',0)}  "
        f"Drop:{py_inbound.get('drop_loads',0)}  "
        f"OnLot:{py_inbound.get('on_lot',0)}  "
        f"AtDoor:{py_inbound.get('at_door',0)}  "
        f"OnLotOrAtDoor:{py_inbound.get('on_lot',0)+py_inbound.get('at_door',0)}\n"
        f"By day: {ib_day_str}"
    )

    pulls_left_today = py_today_totals.get("pulls_left_today", 0)
    picks_left_today = py_today_totals.get("picks_left_today", 0)

    today_actionable_rows = rows_for_selected_day(actionable_rows, day)
    other_day_actionable_rows = rows_not_selected_day(actionable_rows, day)
    today_done_rows = [
        r for r in rows_for_selected_day(all_rows, day)
        if is_done_for_pacing(r.get("status"))
    ]

    day_pacing_text = json.dumps(slim_pacing_for_ai(selected_day_pacing), ensure_ascii=False)

    python_goal_text = ""
    if python_shift_goal_preview:
        python_goal_text = (
            f"PYTHON-COMPUTED SHIFT GOAL (SOURCE OF TRUTH - use this exact goal):\n"
            f"Goal: {python_shift_goal_preview.get('goal', '')}\n"
            f"Confidence: {python_shift_goal_preview.get('confidence', '')}\n"
            f"Main constraint: {python_shift_goal_preview.get('main_constraint', '')}\n"
            f"Reason: {python_shift_goal_preview.get('reason', '')}\n"
            f"Suggested decision: {python_shift_goal_preview.get('suggested_adjustment', '')}\n"
            "Do not create a different shift goal. Explain this goal and use the same appointment cutoff everywhere.\n"
        )
    actionable_table = _rows_to_table(
        today_actionable_rows,
        ["day","load","customer","time","door","status","type","flags","comments"]
    )

    other_day_actionable_count = len(other_day_actionable_rows)
    other_day_actionable_table = f"{other_day_actionable_count} other-day actionable load(s) on the board (context only — not today's work)."


    completed_table = _rows_to_table(
        today_done_rows,
        ["day","load","customer","time","status"]
    )

    prompt = f"""You are an outbound warehouse shift manager. Data comes from Excel cells — treat it as accurate. Short bullets. No corporate fluff. Be direct, practical, specific. This is for operational execution.

CONTEXT: High-volume grocery DC. 1st shift 06:00-16:30. 24-hour clock. Is 9.5 working hours and 1 hour of breaks that we don't take into consideration for calculations.

OPERATING PRIORITIES: 1) Prevent shorts. 2) Protect departures. 3) Protect picking flow. 4) Protect inbound flow. 5) Use proactive labor early. 6) Set up 2nd shift.

STATUS DEFINITIONS:
- RTL = ALREADY STAGED ON DOOR, ready to load. R/S = ready/short (waiting for product, not actionable).
- Picking = being picked. Picking/Short = inventory shortage, picking.
- Loaded Short = trailer loaded but missing product. 
- Late = missed appointment. Completed = done. Loaded = done for pacing.
- Blank = not started. Live = at dock (higher priority than Drop). Drop = drop trailer.
- CPU = customer pickup (protect timing, they may leave). 
- Only positions are: pickers, taskers, loaders, unloaders, receivers, lead and extra

RATES: Pick 185 cases/hr/person. Load 1 trailer/hr/person. Unload 44 pallets/hr/person. Tasking 25 pallets/hr/person. Tasking protects pickers via replenishment/full pallets/putaways. Mention manufacturing help only if it genuinely avoids shorts or protects outbound today.

LABOR RULES:
- Keep pickers picking. Protect loading labor. Protect Tasking when picks/pulls are high. Use Extra proactively.
- Every move: source area → destination area → reason.
- Only move from Lead/Extra WHEN THERE IS AVAILABLE or from an area with positive surplus shown in the staffing table NOW. Never pull from an understaffed, zero, or negative-gap area. Never say "if surplus exists."
- If no safe move exists, say so clearly.

RATE MATH RULES:
- For every labor move AND short-risk call, internally calculate the payoff using RATES when inputs allow.
- Do not compute an exact completion time unless remaining work AND assigned labor are both clearly given. If fuzzy, state the move/risk with no fake precision. Never invent numbers.
- Print format: action → rate math → result. Example: "Move 2 to Picking → +370 cases/hr → reduces Picking gap from -4 to -2."

SOURCE OF TRUTH:
- Python verified counts are the source of truth. Do NOT recount from rows.
- Use SELECTED DAY COUNTS first; ALL BOARD COUNTS are context only.
- Pulls left = board K2 = Inputs!B6. Picks left = board L2 = Inputs!B5.
- Completed loads are separated from actionable on purpose — don't assume they're missing. Completed rows are for pacing.
- Outbound and inbound counts are separate; never mix them.
- Never invent load numbers, customers, doors, statuses, times, counts, goals, or problems.

SHIFT GOAL: Use the PYTHON-COMPUTED SHIFT GOAL exactly as the source of truth. Do not invent a different appointment cutoff or target. You may explain why it matters, but the goal text and cutoff must match Python.

TODAY SELECTED IN APP:
{day} {shift} shift | Total cases: {total_cases:,} | Pulls left from K2: {pulls_left_today} | Picks left from L2: {picks_left_today} | Hours left: {hours_remaining} | Total outbound loads today: {total_outbound_loads} | Plants open: {", ".join(plants_open) if plants_open else "none"} | Notes: {notes.strip() or "none"}

{python_goal_text}

PYTHON DAY-SPECIFIC PACING GUARDRAIL (use first):
{day_pacing_text}

STAFFING — PYTHON COMPUTED:
{staffing_summary}

{allocation_block}

{verified_counts}

{verified_inbound}

{oc_section}

TODAY ACTIONABLE LOADS — SELECTED DAY {day} (notable, flagged, late, short, blank, not-started only):
{actionable_table}

OTHER-DAY ACTIONABLE LOADS (mention separately only, not for today's pacing):
{other_day_actionable_table}

TODAY COMPLETED LOADS — SELECTED DAY {day} (for pacing):
{completed_table}

===== OUTPUT =====
 - OUTPUT ONLY Bottom Line, Shift Health, Prioritization, and Top Action Items. Do NOT add an OC Actions section, OC table, owner/deadline table, or any markdown table (no pipe `|` tables). You may annotate an OC load inline within Prioritization or Action Items, but never as a separate table or section.

BOTTOM LINE (one sentence, max 30 words, before everything): pace status + the shift appointment-time cutoff ( ALWAYS use actual appointment times from today's board) + the single biggest threat to meeting expectations. No data dump.
Example: "Behind 5 loads: should have 17 done by now, have done 12. Target have RTL all loads to 14:00. Move surplus labor to bottleneck."
- Start: "SHIFT HEALTH: GREEN / YELLOW / RED." + one reason.

3. PRIORITIZATION — group as A) Past due / immediate risk  B) Next 2 hours  C) Later today high risk.
- No other-day loads as today's work. Never prioritize any load with Reser's in the customer column.
- If we are running behind and we have loads that are DROPS (specify that only drops) from a reser's customer (priority 4) State that if have not been started yet and are picking heavy, recommend to not start those and make sure to be on time for every CPU, OC and lives. 

4. TOP ACTION ITEMS — Next 30 minutes: exactly 3. Next 2 hours: exactly 3.
- Give specific actionable actions to keep the operations on time, put us back on track or get us ahead.
- No vague verbs unless tied to a specific deadline and load. Don't repeat a 30-min action in the 2-hr block.
- Don't invent actions and don't tell to move workers from extra to a function if there is no extra (bench).


"""

    try:
        response = client.chat.completions.create(
            model="openai/gpt-oss-120b",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_completion_tokens=2850,
            extra_body={"include_reasoning" : False},
        )
        return response.choices[0].message.content

    except Exception as e:
        return f"Board analysis could not be completed: {str(e)}"


def _build_mailto(to_addr, subject, body):
    """Build a mailto: link that prefills recipient, subject, and body."""
    params = urllib.parse.urlencode(
        {"subject": subject, "body": body},
        quote_via=urllib.parse.quote,
    )
    return f"mailto:{to_addr.strip()}?{params}"

# ============================================================
#  EMAIL DRAFT
#  workbook stays as the attachment. Otherwise fall back to the detailed body.
# ============================================================
def build_email_draft(
    day, shift, total_cases, hours_remaining, total_outbound_loads_day,
    summary_table, present_recommendations, recommendations,
    board_analysis_text=None, oc_matches=None,
    board_text="", shift_goal_preview=None,
):
    subject = f"{day} {shift} Shift – Staffing & Board Summary"

    # ---- Python-only concise recap of the PDF first page ----
    try:
        payload = json.loads(board_text or "{}")
    except Exception:
        payload = {}
    py_out = payload.get("python_verified_outbound_summary", {})
    py_in = payload.get("python_verified_inbound_summary", {})
    py_today = payload.get("python_verified_today_totals", {})
    all_rows = payload.get("all_outbound_rows", [])

    pacing = build_selected_day_pacing(all_rows, day, shift, hours_remaining) if all_rows else {}
    health = derive_shift_health(summary_table, pacing, py_out)

    net_gap = int(summary_table["Difference"].sum()) if summary_table is not None and "Difference" in summary_table else 0
    total_present = len(present_recommendations)
    picks_left = py_today.get("picks_left_today", 0)
    pulls_left = py_today.get("pulls_left_today", 0)
    onlot_atdoor = (py_in.get("on_lot", 0) or 0) + (py_in.get("at_door", 0) or 0)

    goal = shift_goal_preview.get("goal", "") if shift_goal_preview else ""

    # Per-function assigned count + gap, from the summary table.
    def _assigned(task):
        try:
            if summary_table is not None and task in summary_table.index:
                return int(summary_table.loc[task, "Assigned"])
        except Exception:
            pass
        return 0

    def _gap(task):
        try:
            if summary_table is not None and task in summary_table.index:
                return int(summary_table.loc[task, "Difference"])
        except Exception:
            pass
        return 0

    # Where the gap is: list short and over functions explicitly.
    short_bits = [f"{t} {_gap(t)}" for t in ["Picking", "Tasking", "Loading", "Unloading", "Receiving"] if _gap(t) < 0]
    over_bits  = [f"{t} +{_gap(t)}" for t in ["Picking", "Tasking", "Loading", "Unloading", "Receiving"] if _gap(t) > 0]
    if short_bits:
        gap_where = "Short: " + ", ".join(short_bits) + ("; Over: " + ", ".join(over_bits) if over_bits else "")
    elif over_bits:
        gap_where = "Over: " + ", ".join(over_bits)
    else:
        gap_where = "balanced across all functions"

    # Coverage % per stream from the same controlled-through logic the PDF uses.
    controlled = appointment_controlled_by_allocation(
        board_text, day, shift, hours_remaining, summary_table
    ) if board_text else {}
    pick_pct = int(float(controlled.get("pick_frac", 0) or 0) * 100)
    pull_pct = int(float(controlled.get("pull_frac", 0) or 0) * 100)
    load_pct = int(float(controlled.get("load_frac", 0) or 0) * 100)

    lines = [f"Shift health: {health}."]
    if goal:
        lines.append(f"Goal: {goal}")
    lines.append(
        f"Loads: {pacing.get('selected_day_total_loads', 0)} today | "
        f"Completed {pacing.get('completed_count', 0)} | "
        f"Due now {pacing.get('due_by_now', 0)} | "
        f"Due not RTL {pacing.get('due_not_RTL', 0)} | "
        f"Pacing {pacing.get('pacing', 'n/a')}."
    )
    lines.append(f"Picks left {picks_left:,} | Pulls left {pulls_left:,} | Hours left {hours_remaining}.")
    lines.append(f"Staffing: {total_present} present | net gap {net_gap:+d} | {gap_where}.")
    lines.append(
        f"Allocation: Pickers {_assigned('Picking')} (can do {pick_pct}% of loads) | "
        f"Taskers {_assigned('Tasking')} (can do {pull_pct}% of day) | "
        f"Loaders {_assigned('Loading')} (can do {load_pct}% of loads) | "
        f"Unloaders {_assigned('Unloading')} | Receivers {_assigned('Receiving')}."
    )
    lines.append(f"Inbound: {py_in.get('loads_read_from_inbound', 0)} loads | {onlot_atdoor} on lot/at door.")
    if oc_matches:
        oc_names = ", ".join(m["customer"]["name"].upper() for m in oc_matches)
        lines.append(f"OC alert: {oc_names} — see report for handling.")

    recap = "\n".join(f"- {ln}" for ln in lines)

    body = (
        "Good morning,\n\n"
        f"{recap}\n\n"
        "Full prioritization, action items, and load alerts are in the attached report.\n\n"
        "Thanks,"
    )
    return subject, body


# ============================================================
#  PDF REPORT GENERATION
#  Generates the final organized report as PDF instead of Excel.
# ============================================================


def clean_pdf_text(value):
    """Clean AI/Markdown text so the PDF prints plain readable text only."""
    if value is None:
        return ""

    text = str(value)

    replacements = {
        "—": "-", "–": "-", "−": "-", "‐": "-", "‑": "-",
        "→": "->", "≤": "<=", "≥": ">=",
        "•": "-", "▪": "-", "■": "-", "□": "-", "●": "-", "◦": "-", "·": "-",
        "“": '"', "”": '"', "‘": "'", "’": "'",
        "\u00a0": " ", "\u200b": "", "\u200c": "", "\u200d": "", "\ufeff": "",
    }

    for src, dst in replacements.items():
        text = text.replace(src, dst)

    # Remove invisible/control characters that ReportLab can render as black boxes.
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", text)
    text = re.sub(r"[\u2000-\u200f\u2028-\u202f\u205f-\u206f]", " ", text)

    # Remove markdown artifacts that should not print in the PDF.
    text = re.sub(r"^\s*#{1,6}\s*", "", text)
    text = text.replace("**", "").replace("__", "")
    text = text.replace("###", "").replace("####", "")
    text = text.replace("---", "")
    text = re.sub(r"`([^`]*)`", r"\1", text)

    # Normalize bullet starters and extra spaces.
    text = re.sub(r"^\s*[-*]+\s*", "", text).strip()
    text = re.sub(r"\s+", " ", text).strip()

    return text

def pdf_safe(value):
    """Clean text for ReportLab paragraphs."""
    if value is None:
        return ""
    text = clean_pdf_text(value)
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return text

def pdf_number(value, default=0):
    try:
        return int(float(value))
    except Exception:
        return default


def pdf_status_color(health):
    health = str(health or "").upper()
    if health == "GREEN":
        return colors.HexColor("#C6EFCE")
    if health == "RED":
        return colors.HexColor("#FFC7CE")
    return colors.HexColor("#FFEB9C")


def derive_shift_health(summary_table, pacing, py_out):
    """Simple transparent health rule for the PDF headline."""
    net_gap = int(summary_table["Difference"].sum()) if summary_table is not None and "Difference" in summary_table else 0
    due_not_RTL = pdf_number(pacing.get("due_not_RTL", 0)) if pacing else 0
    loaded_short = pdf_number(py_out.get("loaded_short_loads", 0)) if py_out else 0
    picking_short = pdf_number(py_out.get("picking_short_loads", 0)) if py_out else 0

    if due_not_RTL > 3 or net_gap <= -5:
        return "RED"
    if net_gap < 2 or picking_short > 1 or str(pacing.get("pacing", "")).upper() == "BEHIND":
        return "YELLOW"
    return "GREEN"


def appointment_cutoff_from_rows(rows):
    times = []
    for row in rows or []:
        m = board_minutes_for_pacing(row.get("time") or row.get("appt_time"))
        if m is not None:
            times.append(m)
    if not times:
        return "today's appointment wave"
    return format_minutes_for_pacing(max(times))


def find_oc_load_matches(board_rows, selected_day):
    """Load-level OC matches for the PDF report."""
    oc_list, _ = load_oc_customer_list()
    selected_day = str(selected_day or "").strip().lower()
    matches = []
    seen = set()

    for row in board_rows or []:
        if selected_day and str(row.get("day", "")).strip().lower() != selected_day:
            continue
        customer_text = str(row.get("customer", "")).strip().lower()
        raw_text = str(row.get("raw_row", "")).strip().lower()
        search_space = f"{customer_text} {raw_text}"
        load = normalize_crossdock_load(row.get("load") or row.get("load_number"))

        for customer in oc_list:
            terms = [customer.get("name", "")] + customer.get("aliases", [])
            found = [t for t in terms if t and str(t).lower() in search_space]
            if not found:
                continue
            key = (load, customer.get("name"))
            if key in seen:
                continue
            seen.add(key)
            matches.append({
                "load": load,
                "customer_on_board": row.get("customer", ""),
                "time": row.get("time") or row.get("appt_time", ""),
                "door": row.get("door", ""),
                "status": row.get("status", ""),
                "oc_name": customer.get("name", ""),
                "priority": customer.get("priority", ""),
                "requirements": customer.get("requirements", ""),
                "sign_off": customer.get("sign_off", False),
                "pictures": customer.get("pictures", False),
                "matched_on": found,
            })
    return matches


def pdf_add_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawString(0.5 * inch, 0.35 * inch, "Staffing Report Generator")
    canvas.drawRightString(7.8 * inch, 0.35 * inch, f"Page {doc.page}")
    canvas.restoreState()


def pdf_table(data, col_widths=None, header_fill="#0F5B78"):
    """
    Standard PDF table that wraps cell text.
    This prevents long staffing/capacity values from running into neighboring cells.
    """
    header_style = ParagraphStyle(
        "PdfTableHeader",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9.5,
        textColor=colors.white,
        wordWrap="CJK",
    )
    body_style = ParagraphStyle(
        "PdfTableBody",
        fontName="Helvetica",
        fontSize=7.1,
        leading=8.6,
        textColor=colors.black,
        wordWrap="CJK",
    )

    def _wrap_cell(value, style):
        # Leave existing ReportLab flowables alone.
        if hasattr(value, "wrapOn"):
            return value
        return Paragraph(pdf_safe(value), style)

    wrapped_data = []
    for row_idx, row in enumerate(data):
        style_for_row = header_style if row_idx == 0 else body_style
        wrapped_data.append([_wrap_cell(cell, style_for_row) for cell in row])

    table = Table(wrapped_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_fill)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.1),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7B7B7")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    table.setStyle(style)
    return table


def pdf_paragraph_list_large(items, styles):
    """Bigger bullet text for the PDF Prioritization and Top Action Items pages."""
    story = []
    for item in items or []:
        if item is None or str(item).strip() == "":
            continue
        story.append(Paragraph(f"- {pdf_safe(item)}", styles["ActionBody"]))
    return story

def extract_ai_shift_goal(board_analysis_text):
    """Pull the Shift goal line from the AI board analysis for the PDF headline."""
    text = str(board_analysis_text or "")
    if not text.strip():
        return ""

    lines = [line.strip() for line in text.splitlines()]
    for i, line in enumerate(lines):
        clean = re.sub(r"^[#\-*\s]+", "", line).strip()
        if re.search(r"shift\s*goal", clean, flags=re.IGNORECASE):
            # Handles: **Shift goal:** Pick and stage...
            clean = re.sub(r"\*", "", clean).strip()
            parts = re.split(r":", clean, maxsplit=1)
            if len(parts) == 2 and parts[1].strip():
                return parts[1].strip()
            # Handles a title line followed by the actual goal.
            for nxt in lines[i + 1:i + 4]:
                nxt_clean = re.sub(r"^[#\-*\s]+", "", nxt).replace("**", "").strip()
                if nxt_clean and not re.search(r"shift\s*goal", nxt_clean, flags=re.IGNORECASE):
                    return nxt_clean
    return ""
    
def _is_markdown_table_line(text):
    """True for markdown-table rows/separators (e.g. '| Load | Customer |' or '|---|---|')
    so stray table fragments never print as garbled pipe rows in the PDF.
    Action lines use '->' not '|', so pipes are safe to treat as table markers."""
    t = str(text or "").strip()
    if not t:
        return False
    if t.startswith("|") or t.endswith("|"):
        return True
    if re.fullmatch(r"[\s|:\-]+", t):
        return True
    if t.count("|") >= 2:
        return True
    return False

def extract_ai_prioritization_lines(board_analysis_text):
    """Return only the AI prioritization section. No Python-generated priority list."""
    text = str(board_analysis_text or "")
    if not text.strip():
        return []

    lines = [line.rstrip() for line in text.splitlines()]
    output = []
    in_priority = False

    for line in lines:
        raw = line.strip()
        upper = raw.upper()

        if "PRIORITIZATION" in upper:
            in_priority = True

        if in_priority:
            # Stop before Top Action Items or the next major section after prioritization.
            if output and (
                re.match(r"^#{0,6}\s*4\.\s", raw)
                or "TOP ACTION" in upper
                or "PICKING & SHORT" in upper
                or "BOARD SUMMARY" in upper
            ):
                break
            if raw and not _is_markdown_table_line(raw) :
                clean = clean_pdf_text(raw)
                clean = re.sub(r"^\d+\.\s*", "", clean).strip()
                if clean:
                    output.append(clean)

    return output


def extract_ai_top_action_items_lines(board_analysis_text):
    """Return only the AI Top Action Items section for the final PDF page."""
    text = str(board_analysis_text or "")
    if not text.strip():
        return []

    lines = [line.rstrip() for line in text.splitlines()]
    output = []
    in_actions = False

    for line in lines:
        raw = line.strip()
        upper = raw.upper()

        if "TOP ACTION ITEMS" in upper:
            in_actions = True

        if in_actions:
            # Stop at OC Actions, an owner/deadline table, or the next major numbered section.
            if output and (
                re.match(r"^#{0,6}\s*[5-9]\.\s", raw)
                or "OC ACTION" in upper
            ):
                break
            if raw and not _is_markdown_table_line(raw):
                clean = clean_pdf_text(raw)
                clean = re.sub(r"^\d+\.\s*", "", clean).strip()
                if clean:
                    output.append(clean)

    return output


def build_pdf_board_summary_rows(selected_rows):
    """Small first-page board summary table for the selected day only."""
    counts = {
        "Total loads": 0,
        "Completed": 0,
        "RTL": 0,
        "R/S + Loaded Short": 0,
        "Picking/Short": 0,
        "Picking no short": 0,
        "Blank/Not Started": 0,
    }

    for row in selected_rows or []:
        counts["Total loads"] += 1
        status = str(row.get("status", "") or "").strip()
        status_upper = status.upper()

        if status_upper in {"COMPLETED", "COMPLETE"}:
            counts["Completed"] += 1
        elif status_upper == "RTL" or "READY TO LOAD" in status_upper:
            counts["RTL"] += 1
        elif (
            status_upper in {"R/S", "READY/SHORT"}
            or "R/S" in status_upper
            or "LOADED SHORT" in status_upper
        ):
            counts["R/S + Loaded Short"] += 1
        elif "PICKING/SHORT" in status_upper or "PICKING SHORT" in status_upper:
            counts["Picking/Short"] += 1
        elif status_upper == "PICKING":
            counts["Picking no short"] += 1
        elif status_upper == "LOADED":
            pass
        elif not status:
            counts["Blank/Not Started"] += 1
        else:
            # Treat unknown active statuses as not-started/needs review for the compact summary.
            counts["Blank/Not Started"] += 1

    return [
        ["Outbound - selected day", ""],
        ["Total loads", counts["Total loads"]],
        ["Completed", counts["Completed"]],
        ["RTL", counts["RTL"]],
        ["R/S + Loaded Short", counts["R/S + Loaded Short"]],
        ["Picking/Short", counts["Picking/Short"]],
        ["Picking no short", counts["Picking no short"]],
        ["Blank/Not Started", counts["Blank/Not Started"]],
    ]


def derive_service_risk_level(summary_table, pacing, py_out, oc_load_matches, crossdock_matches, tt4_matches):
    """First-page service risk level based on actual allocation gaps and selected-day board risk."""
    net_gap = int(summary_table["Difference"].sum()) if summary_table is not None and "Difference" in summary_table else 0
    due_not_RTL = pdf_number(pacing.get("due_not_RTL", 0)) if pacing else 0
    picking_short = pdf_number(py_out.get("picking_short_loads", 0)) if py_out else 0
    loaded_short = pdf_number(py_out.get("loaded_short_loads", 0)) if py_out else 0
    alert_count = len(oc_load_matches or []) + len(crossdock_matches or []) + len(tt4_matches or [])

    if due_not_RTL > 2 or loaded_short > 0 or net_gap <= -5:
        return "HIGH", "Past-due/short exposure or a major actual staffing gap is present."
    if picking_short > 2 or net_gap < 0 or alert_count > 0:
        return "MEDIUM", "Execution is controllable, but staffing gaps or customer/load alerts require follow-up."
    return "LOW", "No major service risk detected from current pacing, staffing, or direct alerts."


def pdf_alert_table(data, col_widths=None, header_fill="#0F5B78", header_text="#000000"):
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_fill)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(header_text)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.2),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7B7B7")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def build_pdf_report(
    day, shift, total_cases, hours_remaining, total_outbound_loads_day,
    summary_table, present_recommendations, recommendations, board_text,
    board_analysis_text, oc_matches, oc_load_matches,
    crossdock_matches, tt4_matches, notes, override_mode=False,
    actual_counts=None, recommended_counts=None, deviation_reason=None,
    python_shift_goal_preview=None,
):
    """Create the final PDF bytes with facts + AI insights."""
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("ReportLab is not installed. Add reportlab to requirements.txt.")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.55 * inch,
    )

    base = getSampleStyleSheet()
    styles = {
        "Title": ParagraphStyle(
            "Title", parent=base["Title"], fontName="Helvetica-Bold", fontSize=18,
            leading=22, alignment=TA_CENTER, textColor=colors.HexColor("#0F5B78"), spaceAfter=8,
        ),
        "Subtitle": ParagraphStyle(
            "Subtitle", parent=base["Normal"], fontSize=9, leading=11,
            alignment=TA_CENTER, textColor=colors.HexColor("#555555"), spaceAfter=12,
        ),
        "Section": ParagraphStyle(
            "Section", parent=base["Heading1"], fontName="Helvetica-Bold", fontSize=14,
            leading=17, textColor=colors.white, backColor=colors.HexColor("#0F5B78"),
            borderPadding=6, spaceBefore=8, spaceAfter=8,
        ),
        "Subsection": ParagraphStyle(
            "Subsection", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=10.5,
            leading=13, textColor=colors.HexColor("#0F5B78"), spaceBefore=6, spaceAfter=4,
        ),
        "AlertTitle": ParagraphStyle(
            "AlertTitle", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=10.5,
            leading=13, textColor=colors.HexColor("#1F1F1F"), spaceBefore=8, spaceAfter=4,
        ),
        "Body": ParagraphStyle("Body", parent=base["Normal"], fontSize=9, leading=12, spaceAfter=4),
        "BodySmall": ParagraphStyle("BodySmall", parent=base["Normal"], fontSize=8.2, leading=10.5, spaceAfter=3),
        "ActionBody": ParagraphStyle("ActionBody", parent=base["Normal"], fontSize=10.2, leading=13.5, spaceAfter=4),
        "Tiny": ParagraphStyle("Tiny", parent=base["Normal"], fontSize=7.2, leading=9, spaceAfter=2),
        "Box": ParagraphStyle(
            "Box", parent=base["Normal"], fontSize=9, leading=12, borderWidth=0.5,
            borderColor=colors.HexColor("#B7B7B7"), borderPadding=6,
            backColor=colors.HexColor("#F7F9FB"), spaceAfter=6,
        ),
    }

    story = []
    story.append(Paragraph("Staffing + Board Full Report", styles["Title"]))
    story.append(Paragraph(
        pdf_safe(f"{day} {shift} Shift | Total cases: {total_cases:,} | Hours remaining: {hours_remaining} | Outbound loads input: {total_outbound_loads_day}"),
        styles["Subtitle"],
    ))

    try:
        payload = json.loads(board_text or "{}")
    except Exception:
        payload = {}
    py_out = payload.get("python_verified_outbound_summary", {})
    py_in = payload.get("python_verified_inbound_summary", {})
    py_today = payload.get("python_verified_today_totals", {})
    all_rows = payload.get("all_outbound_rows", [])
    selected_rows = rows_for_selected_day(all_rows, day)
    pacing = build_selected_day_pacing(all_rows, day, shift, hours_remaining) if all_rows else {}
    health = derive_shift_health(summary_table, pacing, py_out)
    cutoff = appointment_cutoff_from_rows(selected_rows)

    ai_shift_goal = ""
    if python_shift_goal_preview:
        ai_shift_goal = python_shift_goal_preview.get("goal", "")
    if not ai_shift_goal:
        ai_shift_goal = extract_ai_shift_goal(board_analysis_text)
    if not ai_shift_goal:
        ai_shift_goal = f"Have all selected-day loads through {cutoff} controlled, with picks/pulls protected and every OC, Cross Dock, and TT4 action verified before release."

    picks_left = py_today.get("picks_left_today", 0)
    pulls_left = py_today.get("pulls_left_today", 0)

    def _assigned_count(task):
        try:
            if summary_table is not None and task in summary_table.index:
                return int(summary_table.loc[task, "Assigned"])
        except Exception:
            pass
        return 0

    def _needed_count(task):
        try:
            if summary_table is not None and task in summary_table.index:
                return int(summary_table.loc[task, "Needed"])
        except Exception:
            pass
        return 0

    def _gap_count(task):
        try:
            if summary_table is not None and task in summary_table.index:
                return int(summary_table.loc[task, "Difference"])
        except Exception:
            pass
        return 0


    def _staffing_fact(task):
        assigned = _assigned_count(task)
        needed_val = _needed_count(task)
        gap_val = _gap_count(task)
        return f"{assigned} assigned / need {needed_val} / gap {gap_val:+d}"

    pickers = _assigned_count("Picking")
    taskers = _assigned_count("Tasking")
    loaders = _assigned_count("Loading")
    unloaders = _assigned_count("Unloading")
    receivers = _assigned_count("Receiving")

    picking_capacity = pickers * float(hours_remaining or 0) * PICK_RATE
    tasking_pull_capacity = max(0, taskers - TASK_FLOOR) * float(hours_remaining or 0) * PULL_RATE
    loading_capacity = loaders * float(hours_remaining or 0)
    net_gap = int(summary_table["Difference"].sum()) if summary_table is not None and "Difference" in summary_table else 0
    service_risk, service_risk_reason = derive_service_risk_level(
        summary_table, pacing, py_out, oc_load_matches, crossdock_matches, tt4_matches
    )

    # 1. Staffing + Board Summary
    story.append(Paragraph("1. Staffing + Board Summary (Workload and Capacity)", styles["Section"]))
    health_table = Table([
        [
            Paragraph("SHIFT HEALTH", styles["Tiny"]),
            Paragraph("SERVICE RISK", styles["Tiny"]),
            Paragraph("SHIFT GOAL - PYTHON SOURCE OF TRUTH", styles["Tiny"]),
        ],
        [
            Paragraph(f"<b>{pdf_safe(health)}</b>", styles["Body"]),
            Paragraph(f"<b>{pdf_safe(service_risk)}</b><br/>{pdf_safe(service_risk_reason)}", styles["BodySmall"]),
            Paragraph(pdf_safe(ai_shift_goal), styles["BodySmall"]),
        ],
    ], colWidths=[1.15 * inch, 2.0 * inch, 4.2 * inch])
    health_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (0, 1), pdf_status_color(health)),
        ("BACKGROUND", (1, 1), (1, 1), pdf_status_color(service_risk)),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7B7B7")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(health_table)
    story.append(Spacer(1, 6))

    fact_rows = [
        ["Selected-day loads", pacing.get("selected_day_total_loads", "not provided"), "Pacing", pacing.get("pacing", "not provided")],
        ["Completed", pacing.get("completed_count", 0), "Loaded", pacing.get("loaded_count", 0)],
        ["Due by now", pacing.get("due_by_now", 0), "Due now not RTL", pacing.get("due_not_RTL", 0)],
        ["Picks left", picks_left, "Pulls left", pulls_left],
        ["Picking staffing", _staffing_fact("Picking"), "Picking capacity", f"{picking_capacity:,.0f} cases"],
        ["Tasking staffing", _staffing_fact("Tasking"), "Tasking/pull capacity", f"{tasking_pull_capacity:,.0f} pallets"],
        ["Loading staffing", _staffing_fact("Loading"), "Loading capacity", f"{loading_capacity:,.1f} loads"],
        ["Unloading staffing", _staffing_fact("Unloading"), "Receiving staffing", _staffing_fact("Receiving")],
        ["Net staffing gap", f"{net_gap:+d}", "Total present", len(present_recommendations)],
        ["Inbound loads", py_in.get("loads_read_from_inbound", 0), "Inbound on lot/at door", py_in.get("on_lot", 0) + py_in.get("at_door", 0)],
    ]
    story.append(pdf_table([["Fact", "Value", "Fact", "Value"]] + [[pdf_safe(c) for c in r] for r in fact_rows], [1.35*inch, 2.05*inch, 1.45*inch, 2.5*inch]))
    story.append(Spacer(1, 5))

    handoff = build_second_shift_handoff_forecast(
        board_text=board_text,
        day=day,
        shift=shift,
        hours_remaining=hours_remaining,
        summary_table_or_counts=summary_table,
    )
    story.append(Paragraph("Second Shift Handoff Forecast", styles["Subsection"]))
    handoff_rows = [
        ["Loads projected controlled", f"{handoff.get('loads_controlled', 0)} / {handoff.get('selected_day_loads', 0)}", "Controlled through appt", handoff.get("controlled_through_appt", "n/a")],
        ["Loads left for 2nd shift", handoff.get("loads_left_for_next_shift", 0), "Carryover bottleneck", handoff.get("bottleneck", "Unknown")],
        ["Picks left now", f"{handoff.get('picks_left_now', 0):,}", "Est. picks left at handoff", f"{handoff.get('estimated_picks_left_at_handoff', 0):,}"],
        ["Pulls left now", f"{handoff.get('pulls_left_now', 0):,}", "Est. pulls left at handoff", f"{handoff.get('estimated_pulls_left_at_handoff', 0):,}"],
    ]
    story.append(pdf_table([["Handoff Fact", "Value", "Handoff Fact", "Value"]] + [[pdf_safe(c) for c in r] for r in handoff_rows], [1.65*inch, 1.25*inch, 1.85*inch, 2.6*inch]))
    story.append(Paragraph(pdf_safe(handoff.get("handoff_message", "")), styles["BodySmall"]))
    story.append(Spacer(1, 5))

    story.append(Paragraph(f"Board summary - selected day ({pdf_safe(day)})", styles["Subsection"]))
    board_summary_rows = build_pdf_board_summary_rows(selected_rows)
    board_summary_data = [["Status", "Count"]] + board_summary_rows[1:]
    story.append(pdf_table(board_summary_data, [2.45*inch, 1.05*inch]))
    story.append(Spacer(1, 4))

    # Staffing by function was folded into the first-page fact table above so the full
    # first-page summary stays on one PDF page.

    # Removed executive AI summary from the first page to avoid repeating the KPI facts.
    story.append(PageBreak())

    # 2. OC Loads and Actions — own page
    story.append(Paragraph("2. OC Loads and Actions", styles["Section"]))
    if oc_load_matches:
        data = [["Load", "Customer", "Time / Status", "Priority", "Required Actions"]]
        for m in oc_load_matches:
            actions = []
            if m.get("requirements"):
                actions.append(m.get("requirements"))
            if m.get("sign_off"):
                actions.append("DC supervisor sign-off required before ship.")
            if m.get("pictures"):
                actions.append("Photos required: 3 on dock + 3 during loading; email to manager.")
            data.append([
                pdf_safe(m.get("load", "")),
                pdf_safe(m.get("customer_on_board") or m.get("oc_name", "")),
                pdf_safe(f"{m.get('time','')} / {m.get('status','')}"),
                pdf_safe(m.get("priority", "")),
                Paragraph(pdf_safe(" ".join(actions) if actions else "Special handling required per OC list."), styles["Tiny"]),
            ])
        story.append(pdf_alert_table(data, [0.75*inch, 1.55*inch, 1.1*inch, 0.7*inch, 3.2*inch], header_fill="#FFD966", header_text="#000000"))
    elif oc_matches:
        story.append(Paragraph("OC customers were detected in the board data. Load-level detail was not available from the current OC matching function.", styles["BodySmall"]))
        for match in oc_matches:
            c = match["customer"]
            actions = []
            if c.get("requirements"):
                actions.append(c.get("requirements"))
            if c.get("sign_off"):
                actions.append("DC supervisor sign-off required before ship.")
            if c.get("pictures"):
                actions.append("Photos required: 3 on dock + 3 during loading; email to manager.")
            story.append(Paragraph(f"<b>{pdf_safe(c.get('name','').upper())}</b> - {pdf_safe('; '.join(actions))}", styles["BodySmall"]))
    else:
        story.append(Paragraph("No Opportunity Customer loads detected on today's board.", styles["BodySmall"]))

    story.append(PageBreak())

    # 3. Cross Dock — own page
    story.append(Paragraph("3. Cross Dock Loads, Pallets and Actions", styles["Section"]))
    if crossdock_matches:
        data = [["Load", "Board Customer", "Time / Status", "Cross Dock Customer", "Pallets", "Location", "Required Action"]]
        for m in crossdock_matches:
            action = f"Verify {m.get('pallets', 0)} pallet(s) at {m.get('location', '')} are 100% on load {m.get('load', '')} before shipping."
            data.append([
                pdf_safe(m.get("load", "")),
                pdf_safe(m.get("board_customer", "")),
                pdf_safe(f"{m.get('board_time','')} / {m.get('board_status','')}"),
                pdf_safe(m.get("customer", "")),
                pdf_safe(m.get("pallets", 0)),
                pdf_safe(m.get("location", "")),
                Paragraph(pdf_safe(action), styles["Tiny"]),
            ])
        story.append(pdf_alert_table(data, [0.62*inch, 1.2*inch, 0.88*inch, 1.2*inch, 0.5*inch, 0.7*inch, 2.2*inch], header_fill="#A9D18E", header_text="#000000"))
    else:
        story.append(Paragraph("No Cross Dock pallets matched today's board loads, or no Cross Dock sheet was uploaded.", styles["BodySmall"]))

    story.append(PageBreak())

    # 4. TT4 — own page
    story.append(Paragraph("4. Loads with TT4", styles["Section"]))
    if tt4_matches:
        tt4_devices = load_tt4_device_list()
        data = [["Load", "Customer", "Time / Status", "Required Action"]]
        for m in tt4_matches:
            action = f"Verify TT4 requirement is completed for load {m.get('load', '')} before this load ships."
            device_number, caution = match_tt4_device_for_customer(m.get("customer", ""), tt4_devices)
            if device_number:
                action = (
                    f"Use TT4 {device_number} for this customer. "
                    f"Verify it is on load {m.get('load', '')} before this load ships."
                )
                if caution:
                    action += f" Caution: {caution}"
            data.append([
                pdf_safe(m.get("load", "")),
                pdf_safe(m.get("customer", "")),
                pdf_safe(f"{m.get('time','')} / {m.get('status','')}"),
                Paragraph(pdf_safe(action), styles["Tiny"]),
            ])
        story.append(pdf_alert_table(data, [0.8*inch, 2.45*inch, 1.2*inch, 2.9*inch], header_fill="#C00000", header_text="#FFFFFF"))
    else:
        story.append(Paragraph("No TT4-required loads detected on today's board.", styles["BodySmall"]))

    story.append(PageBreak())

    # 5. Prioritization - AI only.
    story.append(Paragraph("5. Prioritization", styles["Section"]))
    priority_lines = extract_ai_prioritization_lines(board_analysis_text)
    if priority_lines:
        story.append(Paragraph("AI prioritization and board execution insight", styles["Subsection"]))
        story.extend(pdf_paragraph_list_large(priority_lines[:70], styles))
    else:
        story.append(Paragraph("AI prioritization was not generated.", styles["Body"]))

    # Keep Top Action Items on the same page when Prioritization is short.
    # ReportLab will naturally continue onto a new page if the content is too long.
    story.append(Spacer(1, 10))

    # 6. Top Action Items - AI only
    story.append(Paragraph("6. Top Action Items", styles["Section"]))
    story.append(Paragraph("AI next actions from the board analysis", styles["Subsection"]))
    top_action_lines = extract_ai_top_action_items_lines(board_analysis_text)
    if top_action_lines:
        story.extend(pdf_paragraph_list_large(top_action_lines[:80], styles))
    else:
        story.append(Paragraph("AI top action items were not generated.", styles["Body"]))

    # Removed possible outcomes / staffing-engine recommendation page per report cleanup.
    doc.build(story, onFirstPage=pdf_add_footer, onLaterPages=pdf_add_footer)
    buffer.seek(0)
    return buffer.getvalue()

def compute_recommended_allocation(
    day, shift, total_cases, hours_remaining, total_outbound_loads_day,
    crossroads_open, deer_creek_open, msb_open, present_workers, board_file=None,
):
    """Phase 1: compute the recommended placement only. No file writes, no AI."""
    total_outbound_loads_actual = total_outbound_loads_day * LOAD_TARGET_SHARE

    # NEEDED is derived from total cases via the per-shift shares in
    # calculate_input_values() -- NOT from the board's all-day remaining picks/pulls.
    # The board totals still drive the goal/handoff preview further below.
    needed, raw_needed, cases_to_pick, full_pallets, inbound_pallets = calculate_needed(
        day, shift, total_cases, hours_remaining, total_outbound_loads_actual,
        crossroads_open, deer_creek_open, msb_open,
    )

    # Read K2/L2 from the board now (before generate_recommendations) so the optimizer
    # can cap needed['Loading'] using real pick/pull throughput.
    # Without a board file, picks_left=0 is treated as unlimited — no cap benefit, no harm.
    _early_picks_left = 0
    _early_pulls_left = 0
    if board_file is not None:
        try:
            board_file.seek(0)
            _early_totals = read_board_today_totals_from_excel(board_file)
            _early_picks_left = pdf_number(_early_totals.get("picks_left_today", 0))
            _early_pulls_left = pdf_number(_early_totals.get("pulls_left_today", 0))
            board_file.seek(0)
        except Exception:
            pass

    # Run the optimizer with the actual present headcount to get the correct Loading cap.
    # This must happen before generate_recommendations so Loading is never overfilled.
    _early_present = len([n for n in present_workers if str(n).strip()])
    _early_optimal = compute_throughput_optimal_allocation(
        picks_left=_early_picks_left,
        pulls_left=_early_pulls_left,
        total_loads=int(total_outbound_loads_day),
        hours_remaining=hours_remaining,
        present_total=_early_present,
    )
    needed["Loading"] = min(needed["Loading"], _early_optimal.get("Loading", needed["Loading"]))

    staffing_sheet = "Staffing sheet 1ST Shift" if shift == "1st" else "Staffing Sheet 2nd Shift"
    staff = pd.read_excel(TEMPLATE_FILE, sheet_name=staffing_sheet, usecols="A,D,F,H,I")
    staff.columns = ["Name", "Skills", "Best Fit", "Present", "Recommended Task"]
    staff = staff[staff["Name"].notna()].copy()

    selected = {name.strip().lower() for name in present_workers}
    staff["Present"] = staff["Name"].astype(str).str.strip().str.lower().apply(
        lambda x: "x" if x in selected else ""
    )

    staff = generate_recommendations(staff, needed)
    present_recommendations, summary_table = build_summary(staff, needed)

    task_order = ["Picking", "Tasking", "Loading", "Unloading", "Receiving"]
    recommended_counts = {
        t: int(summary_table.loc[t, "Assigned"]) if t in summary_table.index else 0
        for t in task_order
    }
    total_present = len(present_recommendations)
    lead_extra = int(
        (present_recommendations["Recommended Task"].astype(str).str.strip() == "Lead/Extra").sum()
    )

    board_text_for_preview = ""
    python_shift_goal_preview = None
    if board_file is not None:
        try:
            board_file.seek(0)
            board_text_for_preview = read_board_file_to_text(board_file)

            # --- Throughput-optimal recommendation ------------------------
            _payload = json.loads(board_text_for_preview or "{}")
            _today = _payload.get("python_verified_today_totals", {}) or {}
            _sel = rows_for_selected_day(_payload.get("all_outbound_rows", []) or [], day)
            completed_or_loaded_now = sum(
                1 for r in _sel if status_is_completed_or_loaded(r.get("status"))
            )
            optimal = compute_throughput_optimal_allocation(
                picks_left=pdf_number(_today.get("picks_left_today", 0)),
                pulls_left=pdf_number(_today.get("pulls_left_today", 0)),
                total_loads=len(_sel),
                hours_remaining=hours_remaining,
                present_total=len(present_recommendations),
                completed_or_loaded_now=completed_or_loaded_now,
            )

            # Build two skill-grounded candidates and KEEP WHICHEVER CONTROLS MORE
            # selected-day loads, using the same appointment-control logic used
            # everywhere else. Both are realized by generate_recommendations(), which
            # only ever assigns a worker to a function they are actually skilled for,
            # so a candidate can never recommend a loader/picker/etc. we don't have.
            #   A) the per-shift staffing NEED from calculate_needed()
            #   B) the throughput-optimal split, capped to the skills actually present
            def _named_counts_for_targets(targets):
                trial = generate_recommendations(staff.copy(), targets)
                _present, _summary = build_summary(trial, needed)
                return trial, _present, _summary, assigned_counts_from_summary(_summary)

            optimal_capped = cap_allocation_to_available_skills(optimal, staff)

            # Cap needed['Loading'] at the optimizer's value before anything else runs.
            # The optimizer accounts for what Picking/Tasking can actually feed to loaders.
            # calculate_needed can overshoot (e.g. Loading=6 when hrs=8) because it doesn't
            # know pick throughput. Updating needed in-place here keeps generate_recommendations,
            # build_summary, and the summary table all consistent — no contradictory "Overstaffed"
            # next to Loading when Picking is the actual bottleneck.
            needed["Loading"] = min(needed["Loading"], optimal_capped.get("Loading", needed["Loading"]))

            need_staff, need_present, need_summary, need_counts = _named_counts_for_targets(needed)
            opt_staff, opt_present, opt_summary, opt_counts = _named_counts_for_targets(optimal_capped)

            def _loads_controlled(counts):
                return int(appointment_controlled_by_allocation(
                    board_text_for_preview, day, shift, hours_remaining, counts,
                ).get("loads_controlled", 0) or 0)

            need_controlled = _loads_controlled(need_counts)
            opt_controlled = _loads_controlled(opt_counts)

            # Pick the allocation that controls more loads. Tie -> fewer total workers
            # used, then fall back to the calculated need for stability.
            if (opt_controlled, -sum(opt_counts.values())) > (need_controlled, -sum(need_counts.values())):
                staff = opt_staff
                present_recommendations, summary_table = opt_present, opt_summary
                recommended_counts = opt_counts
            else:
                staff = need_staff
                present_recommendations, summary_table = need_present, need_summary
                recommended_counts = need_counts

            python_shift_goal_preview = compute_python_shift_goal_preview(
                board_text=board_text_for_preview,
                day=day,
                shift=shift,
                hours_remaining=hours_remaining,
                summary_table=summary_table,
            )
        except Exception:
            python_shift_goal_preview = None
        finally:
            try:
                board_file.seek(0)
            except Exception:
                pass

    total_present = len(present_recommendations)
    lead_extra = int(
        (present_recommendations["Recommended Task"].astype(str).str.strip() == "Lead/Extra").sum()
    )

    return {
        "needed": needed,
        "recommended_counts": recommended_counts,
        "recommended_summary_table": summary_table.copy(),
        "board_text_for_preview": board_text_for_preview,
        "python_shift_goal_preview": python_shift_goal_preview,
        "total_present": total_present,
        "total_recommended": sum(recommended_counts.values()),
        "short_by": max(0, int(pd.Series(needed).sum()) - total_present),
        "lead_extra": lead_extra,
    }

def run_full_generation(
    day, shift, total_cases, hours_remaining, total_outbound_loads_day,
    crossroads_open, deer_creek_open, msb_open, present_workers, notes, board_file,
    crossdock_file=None, override_mode=False, actual_counts=None,
    recommended_counts=None, deviation_reason=None,
):
    """Phase 2: full report: reads board, runs AI, and builds direct alerts."""
    total_outbound_loads_actual = total_outbound_loads_day * LOAD_TARGET_SHARE
    selected = {name.strip().lower() for name in present_workers}

    # NEEDED is derived from total cases via the per-shift shares in
    # calculate_input_values() -- NOT from the board's all-day remaining picks/pulls.
    # (cases_to_pick / full_pallets returned here are the formula values; B5/B6 above
    # still hold the true board remaining for the goal/handoff math.)
    needed, raw_needed, cases_to_pick, full_pallets, inbound_pallets = calculate_needed(
        day, shift, total_cases, hours_remaining, total_outbound_loads_actual,
        crossroads_open, deer_creek_open, msb_open,
    )

    # Same early optimizer cap as Phase 1 — applied before generate_recommendations.
    # board_text is read later, but we can read K2/L2 cheaply now just for this cap.
    _early_picks_left = 0
    _early_pulls_left = 0
    if board_file is not None:
        try:
            board_file.seek(0)
            _early_totals = read_board_today_totals_from_excel(board_file)
            _early_picks_left = pdf_number(_early_totals.get("picks_left_today", 0))
            _early_pulls_left = pdf_number(_early_totals.get("pulls_left_today", 0))
            board_file.seek(0)
        except Exception:
            pass
    _early_present = len([n for n in present_workers if str(n).strip()])
    _early_optimal = compute_throughput_optimal_allocation(
        picks_left=_early_picks_left,
        pulls_left=_early_pulls_left,
        total_loads=int(total_outbound_loads_day),
        hours_remaining=hours_remaining,
        present_total=_early_present,
    )
    needed["Loading"] = min(needed["Loading"], _early_optimal.get("Loading", needed["Loading"]))
    # Further refine using Phase 1 recommended_counts if available.
    if recommended_counts and "Loading" in recommended_counts:
        needed["Loading"] = min(needed["Loading"], recommended_counts["Loading"])

    staffing_sheet = "Staffing sheet 1ST Shift" if shift == "1st" else "Staffing Sheet 2nd Shift"
    staff = pd.read_excel(TEMPLATE_FILE, sheet_name=staffing_sheet, usecols="A,D,F,H,I")
    staff.columns = ["Name", "Skills", "Best Fit", "Present", "Recommended Task"]
    staff = staff[staff["Name"].notna()].copy()
    staff["Present"] = staff["Name"].astype(str).str.strip().str.lower().apply(
        lambda x: "x" if x in selected else ""
    )

    # Cap needed['Loading'] using the Phase 1 recommended_counts (which already had the
    # optimizer cap applied). Keeps the PDF summary table consistent with the Compute step.
    if recommended_counts and "Loading" in recommended_counts:
        needed["Loading"] = min(needed["Loading"], recommended_counts["Loading"])

    staff = generate_recommendations(staff, needed)
    present_recommendations, summary_table = build_summary(staff, needed)

    # Tool's recommended placement, captured BEFORE any supervisor override below
    # reassigns the same crew. This is shown on screen as the comparison board.
    # Use the load-maximizing allocation chosen in Phase 1 (recommended_counts) so
    # the Recommended Staffing Board matches the Recommended Allocation (e.g. shows
    # 3 loaders when 3 control more loads), not the raw per-shift need. Falls back
    # to the need-based board when no Phase-1 recommendation was passed in.
    if recommended_counts:
        rec_targets = cap_allocation_to_available_skills(recommended_counts, staff)
        rec_board_staff = generate_recommendations(staff.copy(), rec_targets)
        recommended_present_board, _ = build_summary(rec_board_staff, needed)
        recommended_present_board = recommended_present_board.copy()
    else:
        recommended_present_board = present_recommendations.copy()
    ai_recommended = None
    ai_reason = None

    if override_mode and actual_counts:
        # Assign named workers to the requested actual/recommended allocation, but still
        # enforce skills. If only 2 present workers have L, the named board and the
        # assigned Loading count will both show 2, not an impossible 4.
        actual_counts = cap_allocation_to_available_skills(actual_counts, staff)
        staff = generate_recommendations(staff, actual_counts)
        present_recommendations, summary_table = build_summary(staff, needed)

        # In override mode, every present worker must be assigned to one of the five functions.
        # Safe move sources are only the functions showing a positive gap.
        availability = compute_labor_availability(summary_table, present_recommendations, lead_extra_count=0)
        ai_recommended = recommended_counts
        ai_reason = deviation_reason
    else:
        availability = compute_labor_availability(summary_table, present_recommendations)

    recommendations = build_recommendations(
        summary_table, present_recommendations, raw_needed, hours_remaining, notes,
        availability=availability,
    )

    board_analysis_text = None
    python_shift_goal_preview = None
    oc_matches = []
    oc_load_matches = []
    crossdock_matches = []
    tt4_matches = []

    if board_file is not None:
        board_file.seek(0)
        board_text = read_board_file_to_text(board_file)

        # Direct Python alerts only. These do not go to AI.
        try:
            board_payload_for_alerts = json.loads(board_text)
            board_rows_for_alerts = board_payload_for_alerts.get("all_outbound_rows", [])

            oc_load_matches = find_oc_load_matches(board_rows_for_alerts, day)
            tt4_matches = find_tt4_required_loads(board_rows_for_alerts, day)

            if crossdock_file is not None:
                crossdock_rows = read_crossdock_rows(crossdock_file)
                crossdock_matches = find_crossdock_matches(crossdock_rows, board_rows_for_alerts)

        except Exception as e:
            st.error(f"Direct alert matching failed: {e}")
            st.exception(e)

        oc_matches = find_oc_customers_in_board(board_text)
        oc_alert_text = build_oc_alert_text(oc_matches)

        python_shift_goal_preview = compute_python_shift_goal_preview(
            board_text=board_text,
            day=day,
            shift=shift,
            hours_remaining=hours_remaining,
            summary_table=summary_table,
        )

        board_analysis_text = analyze_board_with_groq(
            board_text=board_text, day=day, shift=shift, total_cases=total_cases,
            hours_remaining=hours_remaining, total_outbound_loads=total_outbound_loads_day,
            crossroads_open=crossroads_open, deer_creek_open=deer_creek_open, msb_open=msb_open,
            needed=needed, summary_table=summary_table, cases_to_pick=cases_to_pick,
            inbound_pallets=inbound_pallets, notes=notes, oc_alert_text=oc_alert_text,
            recommended_allocation=ai_recommended, deviation_reason=ai_reason,
            python_shift_goal_preview=python_shift_goal_preview,
        )

        # --- Snapshot today's commitments + shift goal for end-of-shift closeout ---
        try:
            import shift_log
            operating_date = datetime.date.today().strftime("%m/%d/%Y")
            cpu_commitments = []
            for r in rows_for_selected_day(board_rows_for_alerts, day):
                if "CPU" in str(r.get("type", "")).upper():
                    cpu_commitments.append({
                        "load": r.get("load") or r.get("load_number", ""),
                        "customer": r.get("customer", ""),
                        "appt_time": r.get("time") or r.get("appt_time", ""),
                        "morning_status": r.get("status", ""),
                    })
            shift_goal_for_snapshot = (
                python_shift_goal_preview.get("goal", "")
                if python_shift_goal_preview else ""
            )
            shift_log.snapshot_commitments(
                operating_date, shift, oc_load_matches, cpu_commitments,
                shift_goal=shift_goal_for_snapshot, total_loads_for_day=int(total_outbound_loads_day or 0), 
            )
        except Exception as e:
            st.warning(f"Commitment snapshot skipped: {e}")

    else:
        board_text = ""

    email_subject, email_body = build_email_draft(
        day=day, shift=shift, total_cases=total_cases, hours_remaining=hours_remaining,
        total_outbound_loads_day=total_outbound_loads_day, summary_table=summary_table,
        present_recommendations=present_recommendations, recommendations=recommendations,
        board_analysis_text=board_analysis_text, oc_matches=oc_matches,
        board_text=board_text, shift_goal_preview=python_shift_goal_preview,
    )

    pdf_output_bytes = build_pdf_report(
        day=day, shift=shift, total_cases=total_cases, hours_remaining=hours_remaining,
        total_outbound_loads_day=total_outbound_loads_day, summary_table=summary_table,
        present_recommendations=present_recommendations, recommendations=recommendations,
        board_text=board_text, board_analysis_text=board_analysis_text,
        oc_matches=oc_matches,
        oc_load_matches=oc_load_matches, crossdock_matches=crossdock_matches,
        tt4_matches=tt4_matches, notes=notes, override_mode=override_mode,
        actual_counts=actual_counts, recommended_counts=recommended_counts,
        deviation_reason=deviation_reason,
        python_shift_goal_preview=python_shift_goal_preview,
    )

    return {
        "output_bytes": pdf_output_bytes,
        "summary_table": summary_table,
        "present_recommendations": present_recommendations,
        "recommended_present_board": recommended_present_board,
        "recommendations": recommendations,
        "board_analysis_text": board_analysis_text,
        "oc_matches": oc_matches,
        "oc_load_matches": oc_load_matches,
        "crossdock_matches": crossdock_matches,
        "tt4_matches": tt4_matches,
        "email_subject": email_subject,
        "email_body": email_body,
        "override_mode": override_mode,
        "actual_counts": actual_counts,
        "recommended_counts": recommended_counts,
        "deviation_reason": deviation_reason,
        "python_shift_goal_preview": python_shift_goal_preview,
    }


# ============================================================
#  STREAMLIT INTERFACE
# ============================================================

st.sidebar.header("Daily Inputs")

day = st.sidebar.selectbox(
    "Day",
    ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
)

shift = st.sidebar.selectbox("Shift", ["1st", "2nd"])

# Load names fresh every render — no cache — filtered by selected shift
names = load_names_for_shift(shift)

total_cases = st.sidebar.number_input("Total Cases for Today", min_value=0, step=1, value=0)

hours_remaining = st.sidebar.number_input("Hours Remaining in Shift", min_value=0.0, step=0.25, value=8.0)

total_outbound_loads_day = st.sidebar.number_input("Total Outbound Loads for the Day", min_value=0, step=1, value=0)

crossroads_open = st.sidebar.selectbox("Crossroads plant open?", ["YES", "NO"])
deer_creek_open = st.sidebar.selectbox("Deer Creek plant open?", ["YES", "NO"])
msb_open        = st.sidebar.selectbox("MSB plant open?", ["YES", "NO"])

present_workers = st.sidebar.multiselect("Who is present?", names)

notes = st.sidebar.text_area("Operations Notes")

st.markdown("---")
st.subheader("Outbound Board Excel / CSV")

board_file = st.file_uploader(
    "Upload the outbound load board Excel or CSV file",
    type=["xlsx", "xls", "csv"],
    help="Cell values and color flags (yellow = load check, light-blue = TT4, red font = Canadian) are read directly from the file.",
)

if board_file:
    st.success("Board file loaded — ready for analysis.")

st.subheader("Cross Dock Sheet Excel")
crossdock_file = st.file_uploader(
    "Upload the daily Cross Dock sheet Excel",
    type=["xlsx", "xls"],
    help="Matches Cross Dock Trip # against the board load # and alerts you when a pallet/location must be verified before shipping.",
)

if crossdock_file:
    st.success("Cross Dock sheet loaded — alerts will be matched against the board.")

if board_file:
    with st.expander("Preview: What Python parsed from the board (no AI tokens used)", expanded=False):
        try:
            board_file.seek(0)
            file_name_lower = board_file.name.lower()
            if file_name_lower.endswith(".csv"):
                preview_rows = board_records_from_csv(board_file)
            else:
                preview_rows = board_records_from_excel(board_file)

            if not preview_rows:
                st.warning("No load rows were parsed. Check that the Outbound sheet has load numbers in column A and the board values are saved/calculated in Excel.")
            else:
                total_staff_present = len(present_workers)
                st.metric("Staff Present Today", total_staff_present)

                if not file_name_lower.endswith(".csv"):
                    board_file.seek(0)
                    preview_today_totals = read_board_today_totals_from_excel(board_file)
                    pcol1, pcol2 = st.columns(2)
                    pcol1.metric("Pulls Left Today from K2", preview_today_totals.get("pulls_left_today", 0))
                    pcol2.metric("Picks Left Today from L2", preview_today_totals.get("picks_left_today", 0))

                st.markdown("---")

                preview_summary = build_python_board_summary(preview_rows)
                total = preview_summary["loads_read_from_board"]

                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Total Loads",   total)
                col2.metric("RTL",           preview_summary["rtl_loads"])
                col3.metric("Picking/Short", preview_summary["picking_short_loads"])
                col4.metric("R/S",           preview_summary["rs_loads"])
                col5.metric("Loaded Short",  preview_summary["loaded_short_loads"])

                col6, col7, col8, col9, col10 = st.columns(5)
                col6.metric("Picking",           preview_summary["picking_loads"])
                col7.metric("Blank/Not Started", preview_summary["blank_or_not_started_loads"])
                col8.metric("Live Loads",        preview_summary["live_loads"])
                col9.metric("CPU Loads",         preview_summary["cpu_loads"])
                col10.metric("Late",             preview_summary["late_loads"])

                st.caption(f"Outbound loads by day: {preview_summary['loads_by_day']}")

                board_file.seek(0)
                inbound_preview_rows = board_records_from_inbound_sheet(board_file)
                if inbound_preview_rows:
                    inbound_preview_summary = build_python_inbound_summary(inbound_preview_rows)
                    st.markdown("---")
                    st.markdown("**Inbound**")
                    ib1, ib2, ib3, ib4 = st.columns(4)
                    ib1.metric("Total Inbound",    inbound_preview_summary["loads_read_from_inbound"])
                    ib2.metric("Live",             inbound_preview_summary["live_loads"])
                    ib3.metric("Drop",             inbound_preview_summary["drop_loads"])
                    ib4.metric("On Lot / At Door", inbound_preview_summary["on_lot"] + inbound_preview_summary["at_door"])
                    st.caption(f"Inbound loads by day: {inbound_preview_summary['loads_by_day']}")
                    inbound_df = pd.DataFrame([
                        {
                            "Day":      r.get("day", ""),
                            "Load #":   r.get("load_number", ""),
                            "Carrier":  r.get("carrier", ""),
                            "Time":     r.get("appt_time", ""),
                            "Type":     r.get("type", ""),
                            "Trailer":  r.get("trailer", ""),
                            "Status":   r.get("status", ""),
                            "Receiver": r.get("receiver", ""),
                            "Origin":   r.get("origin", ""),
                            "Notes":    r.get("notes", ""),
                        }
                        for r in inbound_preview_rows
                    ])
                    st.dataframe(inbound_df, use_container_width=True, height=250)

                st.markdown("---")
                st.markdown("**Every outbound load row Python extracted from the file:**")
                preview_df = pd.DataFrame([
                    {
                        "Day":      r.get("day", ""),
                        "Date":     r.get("date", ""),
                        "Load #":   r.get("load_number", ""),
                        "Customer": r.get("customer", ""),
                        "Carrier":  r.get("carrier", ""),
                        "Time":     r.get("appt_time", ""),
                        "Door":     r.get("door", ""),
                        "Trailer":  r.get("trailer", ""),
                        "Status":   r.get("status", "") or "—",
                        "Type":     r.get("type", ""),
                        "TT4":      r.get("tt4", ""),
                        "Loader":   r.get("loader", ""),
                        "Picks":    r.get("picks", 0),
                        "Pulls":    r.get("pulls", 0),
                        "Flags":    ", ".join(r.get("flags", [])),
                        "Comments": r.get("comments", ""),
                    }
                    for r in preview_rows
                ])
                st.dataframe(preview_df, use_container_width=True, height=400)

                st.markdown("**Quick sanity checks:**")
                issues = []
                blank_time = [r["load_number"] for r in preview_rows if not r.get("appt_time")]
                if blank_time:
                    issues.append(f" {len(blank_time)} load(s) have no time parsed: {', '.join(blank_time[:5])}{'...' if len(blank_time) > 5 else ''}")
                blank_customer = [r["load_number"] for r in preview_rows if not r.get("customer")]
                if blank_customer:
                    issues.append(f" {len(blank_customer)} load(s) have no customer name: {', '.join(blank_customer[:5])}")
                no_day = [r["load_number"] for r in preview_rows if not r.get("day")]
                if no_day:
                    issues.append(f" {len(no_day)} load(s) have no day context (missing day header row?): {', '.join(no_day[:5])}")
                if issues:
                    for issue in issues:
                        st.warning(issue)
                else:
                    st.success("All loads have time, customer, and day context — parse looks clean.")

        except Exception as e:
            st.error(f"Preview failed: {e}")
            st.exception(e)

with st.expander("View Opportunity Customer List (from Excel file)"):
    oc_list_preview, oc_load_error = load_oc_customer_list()
    if oc_list_preview:
        oc_preview_rows = []
        for c in oc_list_preview:
            oc_preview_rows.append({
                "Customer":        c["name"].title(),
                "Customer #":      c["customer_number"] or "—",
                "Priority":        c["priority"],
                "Issue":           c["issue"],
                "DC Requirements": c["requirements"],
                "Sign-Off Required": "Yes" if c["sign_off"] else "No",
                "Photos Required":   "Yes" if c["pictures"] else "No",
            })
        st.dataframe(pd.DataFrame(oc_preview_rows), use_container_width=True)
        st.caption(f"Loaded {len(oc_list_preview)} customers from '{OC_FILE}'")
    elif oc_load_error:
        st.warning(oc_load_error)
    else:
        st.info(f"OC customer list not connected. Place '{OC_FILE}' in the app folder to enable this feature.")

st.markdown("---")

# ── Phase 1: compute the recommended allocation (no board read, no AI) ──────
if st.button("Compute Recommended Allocation"):
    if not present_workers:
        st.error("Select who is present first.")
    else:
        reco = compute_recommended_allocation(
            day, shift, total_cases, hours_remaining, total_outbound_loads_day,
            crossroads_open, deer_creek_open, msb_open, present_workers, board_file,
        )
        reco["snapshot"] = {
            "day": day,
            "shift": shift,
            "total_cases": total_cases,
            "hours_remaining": hours_remaining,
            "total_outbound_loads_day": total_outbound_loads_day,
            "crossroads_open": crossroads_open,
            "deer_creek_open": deer_creek_open,
            "msb_open": msb_open,
            "present_workers": sorted(present_workers),
            "notes": notes,
            "board_sig": (board_file.name + str(getattr(board_file, "size", ""))) if board_file else "none",
            "crossdock_sig": (crossdock_file.name + str(getattr(crossdock_file, "size", ""))) if crossdock_file else "none",
        }
        st.session_state["reco"] = reco
        st.session_state.pop("gen_result", None)
        st.session_state.pop("allocation_choice", None)

reco = st.session_state.get("reco")

if reco:
    task_order = ["Picking", "Tasking", "Loading", "Unloading", "Receiving"]
    label_map = {
        "Picking": "Pickers",
        "Tasking": "Taskers",
        "Loading": "Loaders",
        "Unloading": "Unloaders",
        "Receiving": "Receivers",
    }
    rc = reco["recommended_counts"]

    st.subheader("Recommended Allocation")
    st.caption(
        f"Built from {reco['total_present']} present. "
        f"Extra over today's workload (bench): {reco['lead_extra']}. "
        "If you change any input on the left, click Compute again to refresh."
    )
    reco_df = pd.DataFrame(
        [{"Position": label_map[t], "Recommended": int(rc.get(t, 0))} for t in task_order]
    )
    st.table(reco_df)

    if reco.get("recommended_summary_table") is not None:
        st.subheader("Staffing Summary Used for This Preview")
        st.dataframe(reco["recommended_summary_table"], use_container_width=True)

    render_python_shift_goal_preview(reco.get("python_shift_goal_preview"))
    render_allocation_controls_preview(
        "Recommended allocation",
        appointment_controlled_by_allocation(
            reco.get("board_text_for_preview", ""), day, shift, hours_remaining,
            reco["recommended_counts"],
        ),
    )

    choice = st.radio(
        "Are you running this recommended allocation?",
        [
            "(choose one)",
            "Yes - generate the report on these numbers",
            "No - I'm running a different allocation",
        ],
        index=0,
        key="allocation_choice",
    )

    desired = None  # becomes (mode, actual_counts, reason) when a request is active

    if choice.startswith("Yes"):
        # Running the recommended numbers AS the actual allocation — report on them as-is.
        desired = ("override", dict(reco["recommended_counts"]), "Running the tool's recommended allocation.")

    elif choice.startswith("No"):
        st.markdown("**Enter what you actually have on each position:**")
        present_total = int(reco["total_present"])

        def clamp_widget_value(key, max_value):
            """Keep override inputs from ever exceeding the remaining present headcount."""
            max_value = max(0, int(max_value))
            if key in st.session_state:
                try:
                    current_value = int(st.session_state[key])
                except Exception:
                    current_value = 0
                if current_value > max_value:
                    st.session_state[key] = max_value
                elif current_value < 0:
                    st.session_state[key] = 0

        c1, c2, c3 = st.columns(3)
        c4, c5, _ = st.columns(3)

        # These max values are dynamic by row order. The total entered allocation
        # cannot exceed the number of present workers. To give more to a later
        # area, lower an earlier area first.
        remaining_for_pick = present_total
        clamp_widget_value("act_pick", remaining_for_pick)
        a_pick = c1.number_input(
            "Pickers",
            min_value=0,
            max_value=remaining_for_pick,
            step=1,
            value=min(int(rc.get("Picking", 0)), remaining_for_pick),
            key="act_pick",
        )

        remaining_for_task = max(0, present_total - int(a_pick))
        clamp_widget_value("act_task", remaining_for_task)
        a_task = c2.number_input(
            "Taskers",
            min_value=0,
            max_value=remaining_for_task,
            step=1,
            value=min(int(rc.get("Tasking", 0)), remaining_for_task),
            key="act_task",
        )

        remaining_for_load = max(0, present_total - int(a_pick) - int(a_task))
        clamp_widget_value("act_load", remaining_for_load)
        a_load = c3.number_input(
            "Loaders",
            min_value=0,
            max_value=remaining_for_load,
            step=1,
            value=min(int(rc.get("Loading", 0)), remaining_for_load),
            key="act_load",
        )

        remaining_for_unload = max(0, present_total - int(a_pick) - int(a_task) - int(a_load))
        clamp_widget_value("act_unload", remaining_for_unload)
        a_unload = c4.number_input(
            "Unloaders",
            min_value=0,
            max_value=remaining_for_unload,
            step=1,
            value=min(int(rc.get("Unloading", 0)), remaining_for_unload),
            key="act_unload",
        )

        remaining_for_recv = max(0, present_total - int(a_pick) - int(a_task) - int(a_load) - int(a_unload))
        clamp_widget_value("act_recv", remaining_for_recv)
        a_recv = c5.number_input(
            "Receivers",
            min_value=0,
            max_value=remaining_for_recv,
            step=1,
            value=min(int(rc.get("Receiving", 0)), remaining_for_recv),
            key="act_recv",
        )

        actual_counts = {
            "Picking": int(a_pick),
            "Tasking": int(a_task),
            "Loading": int(a_load),
            "Unloading": int(a_unload),
            "Receiving": int(a_recv),
        }
        entered_total = sum(actual_counts.values())

        override_preview = None
        if reco.get("board_text_for_preview"):
            actual_summary_for_preview = build_summary_table_from_counts(reco.get("needed", {}), actual_counts)
            override_preview = compute_python_shift_goal_preview(
                board_text=reco.get("board_text_for_preview", ""),
                day=day,
                shift=shift,
                hours_remaining=hours_remaining,
                summary_table=actual_summary_for_preview,
            )
            render_python_shift_goal_preview(override_preview)
            render_allocation_controls_preview(
                "Your allocation",
                appointment_controlled_by_allocation(
                    reco.get("board_text_for_preview", ""), day, shift, hours_remaining,
                    actual_counts,
                ),
            )

        reason = st.text_area(
            "Why are you running it differently? (the AI uses this so it won't flag an intentional gap)",
            key="dev_reason",
            placeholder="e.g. Only 1 loader instead of 3 - nothing RTL yet, so 1 is fine for now.",
        )

        st.caption(f"Entered total: {entered_total}   |   Present: {present_total}")

        if st.button("Confirm & Generate Report"):
            if entered_total != present_total:
                st.error(
                    f"Your counts add up to {entered_total}, but you marked {present_total} present. "
                    "They have to match before I can generate - adjust the numbers (or who's present)."
                )
            else:
                desired = ("override", actual_counts, reason)

    # ── Generate only when a request is active; cache by signature so AI
    #    fires once per unique config (not on every Streamlit rerun) ─────────
    if desired is not None:
        mode, dz_actual_counts, dz_reason = desired
        signature = json.dumps(
            {
                "snapshot": reco["snapshot"],
                "mode": mode,
                "actual_counts": dz_actual_counts,
                "reason": dz_reason,
            },
            sort_keys=True,
            default=str,
        )

        prev = st.session_state.get("gen_result")
        if not prev or prev.get("signature") != signature:
            with st.spinner("Generating PDF report (reading board, running AI, building PDF)..."):
                result = run_full_generation(
                    day,
                    shift,
                    total_cases,
                    hours_remaining,
                    total_outbound_loads_day,
                    crossroads_open,
                    deer_creek_open,
                    msb_open,
                    present_workers,
                    notes,
                    board_file,
                    crossdock_file=crossdock_file,
                    override_mode=(mode == "override"),
                    actual_counts=dz_actual_counts,
                    recommended_counts=rc,
                    deviation_reason=dz_reason,
                )
            result["signature"] = signature
            st.session_state["gen_result"] = result

# ── Render the generated report ────────────────────────────────────────────
result = st.session_state.get("gen_result")
if result:
    summary_table = result["summary_table"]
    present_recommendations = result["present_recommendations"]
    recommendations = result["recommendations"]
    oc_matches = result["oc_matches"]
    crossdock_matches = result.get("crossdock_matches", [])
    tt4_matches = result.get("tt4_matches", [])
    board_analysis_text = result["board_analysis_text"]

    st.success("PDF staffing report generated successfully.")

    if result["override_mode"]:
        st.info(
            "This report reflects the ACTUAL allocation you entered, not the tool's recommendation. "
            "The 'Recommended Staffing Board' below is the tool's suggested placement, shown for comparison."
        )
        if result.get("deviation_reason"):
            st.caption(f"Your stated reason: {result['deviation_reason']}")

    if tt4_matches:
        st.markdown("---")
        st.subheader("TT4 Alerts")
        st.error(
            "The following loads on today's board require TT4. "
            "Verify TT4 is completed before these loads ship."
        )
        for match in tt4_matches:
            with st.expander(
                f"Load {match.get('load', '')} — {match.get('customer', '')} — {match.get('time', '')}",
                expanded=True,
            ):
                st.markdown(f"**Customer:** {match.get('customer', '')}")
                st.markdown(f"**Time / Status:** {match.get('time', '')} / {match.get('status', '')}")
                st.markdown(
                    f"**Required Action:** Verify TT4 requirement is completed for load "
                    f"{match.get('load', '')} before this load ships."
                )
    elif board_file is not None:
        st.info("No TT4-required loads detected on today's board.")

    if crossdock_matches:
        st.markdown("---")
        st.subheader("Cross Dock Alerts")
        st.error(
            "The following Cross Dock pallet(s) matched loads on today's board. "
            "Verify these pallet(s) are 100% on the correct load before shipping."
        )
        for match in crossdock_matches:
            with st.expander(
                f"Load {match.get('load', '')} — {match.get('pallets', 0)} pallet(s) at {match.get('location', '')}",
                expanded=True,
            ):
                st.markdown(f"**Board Customer:** {match.get('board_customer', '')}")
                st.markdown(
                    f"**Board Time / Door / Status:** "
                    f"{match.get('board_time', '')} / {match.get('board_door', '')} / {match.get('board_status', '')}"
                )
                st.markdown(f"**Cross Dock Customer:** {match.get('customer', '')}")
                st.markdown(f"**Order/PO:** {match.get('order_po', '')}")
                st.markdown(f"**Location:** {match.get('location', '')}")
                st.markdown(
                    f"**Required Action:** Verify {match.get('pallets', 0)} pallet(s) located at "
                    f"{match.get('location', '')} are 100% on load {match.get('load', '')} before this load ships."
                )
    elif crossdock_file is not None and board_file is not None:
        st.info("No Cross Dock pallets matched the load numbers on today's board.")

    if oc_matches:
        st.markdown("---")
        st.subheader("Opportunity Customer Alerts")
        st.error(
            "The following customers on today's board are on the Opportunity Customer List "
            "and require special DC actions before their loads ship."
        )
        for match in oc_matches:
            c = match["customer"]
            with st.expander(f"{c['name'].upper()}  -  Priority: {c['priority']}", expanded=True):
                st.markdown(f"**DC Requirements:** {c['requirements']}")
                if c["sign_off"]:
                    st.markdown("**DC Supervisor Sign-Off REQUIRED before this load ships.**")
                if c["pictures"]:
                    st.markdown("**Photos REQUIRED:** 3 on dock + 3 during loading (6 total). Email to manager.")
    elif board_file is not None:
        st.info("No Opportunity Customers detected on today's board.")

    st.subheader("Staffing Summary")
    st.dataframe(summary_table, use_container_width=True)

    st.subheader("Recommended Staffing Board")
    recommended_board_to_show = result.get("recommended_present_board")
    if recommended_board_to_show is None:
        recommended_board_to_show = present_recommendations
    if result.get("override_mode") and result.get("deviation_reason") not in (
        None, "", "Running the tool's recommended allocation."
    ):
        st.caption(
            "This is the tool's recommended placement, shown for comparison. "
            "The report above reflects your actual allocation."
        )
    st.dataframe(
        recommended_board_to_show[["Name", "Skills", "Best Fit", "Recommended Task"]].reset_index(drop=True),
        use_container_width=True,
    )

    st.subheader("Written Recommendations / What-Ifs")
    for rec in recommendations:
        st.write(f"• {rec}")

    if board_analysis_text:
        st.markdown("---")
        st.subheader("Board Excel Analysis - AI Insights")
        st.info(
            "The analysis below was generated by Groq AI reading the board Excel/CSV file directly "
            "from cell values, including color flags for load checks, TT4s, and Canadian loads."
        )
        st.markdown(board_analysis_text)

    st.download_button(
        label="Download PDF Staffing Report",
        data=result["output_bytes"],
        file_name="Staffing Board Full Report.pdf",
        mime="application/pdf",
    )

    st.markdown("---")
    st.subheader("Email this report")
    boss_email = "brianm@resers.com"

    st.link_button(
        "Email to manager",
        _build_mailto(boss_email, result["email_subject"], result["email_body"]),
    )
    st.caption(
        "Opens your mail app with the message prefilled to "
        f"{boss_email}. Attach the PDF you downloaded above, then send."
    )
