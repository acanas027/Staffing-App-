"""
Order vs Inventory Matcher — single-file Streamlit app.

Upload a sample-orders workbook and a QPORT inventory report.
The app matches every ordered item to its available quantity and locations,
one row per item, with every stocking location and its quantity listed together.

Run with:
    pip install -r requirements.txt
    streamlit run app.py
"""
import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Order vs Inventory Matcher", layout="wide")

# ---------------------------------------------------------------------------
# TKRESERVE_QPORT layout (0-based column positions, header row = row 1)
#   A-F : Location hierarchy (zone, aisle, rack, shelf, position, level)
#         -> combined into one "Location" string, e.g. "RC2A6X3"
#   H   : Sku Number prefix + 2-digit fraction (e.g. 71117.00)
#   I   : Sku Number last 3 digits of the suffix (e.g. 225)
#         Full SKU = f"{int(H)}.{frac:02d}{int(I):03d}" -> "71117.00225"
#   J   : QUANTITY (units available for that location/lot)
#   K-P : Previous location hierarchy (same shape as A-F)
#         -> combined into one "Previous Location" string per row, e.g. "RC2A1X1"
#         The single most recent one (by the Y/Z transaction date-time) is kept per item.
# ---------------------------------------------------------------------------
COL_LOC = [0, 1, 2, 3, 4, 5]
COL_SKU_H = 7
COL_SKU_I = 8
COL_QTY = 9
COL_PREV = [10, 11, 12, 13, 14, 15]
COL_DATE2 = 24   # Y - most recent transaction date
COL_TIME2 = 25   # Z - most recent transaction time


def _clean(v):
    if pd.isna(v):
        return ""
    return str(v).strip()


def build_sku(h, i):
    """Reconstruct the full SKU number from the H (prefix.fraction) and I (suffix) columns."""
    if pd.isna(h) or pd.isna(i):
        return None
    try:
        h = float(h)
        i = float(i)
    except (ValueError, TypeError):
        return None
    int_part = int(h)
    frac = int(round((h - int_part) * 100))
    suffix = int(round(i))
    return f"{int_part}.{frac:02d}{suffix:03d}"


def load_tkreserve(file) -> pd.DataFrame:
    """Read TKRESERVE_QPORT.xlsx -> ['SKU', 'Location', 'Quantity', 'Previous Location', 'TxTime']"""
    raw = pd.read_excel(file, sheet_name=0, header=0)

    loc = raw.iloc[:, COL_LOC].apply(
        lambda row: "".join([_clean(v) for v in row if _clean(v) != ""]), axis=1
    )
    prev_loc = raw.iloc[:, COL_PREV].apply(
        lambda row: "".join([_clean(v) for v in row if _clean(v) != ""]), axis=1
    )

    h = pd.to_numeric(raw.iloc[:, COL_SKU_H], errors="coerce")
    i = pd.to_numeric(raw.iloc[:, COL_SKU_I], errors="coerce")
    sku = [build_sku(a, b) for a, b in zip(h, i)]

    qty = pd.to_numeric(raw.iloc[:, COL_QTY], errors="coerce").fillna(0)

    date2 = pd.to_numeric(raw.iloc[:, COL_DATE2], errors="coerce").fillna(0)
    time2 = pd.to_numeric(raw.iloc[:, COL_TIME2], errors="coerce").fillna(0)
    tx_time = date2 * 1_000_000 + time2

    df = pd.DataFrame({
        "SKU": sku, "Location": loc, "Quantity": qty,
        "Previous Location": prev_loc, "TxTime": tx_time,
    })
    df = df.dropna(subset=["SKU"])
    return df[df["SKU"] != ""]


def aggregate_inventory(inv: pd.DataFrame) -> pd.DataFrame:
    """One row per SKU:
    - Quantity Available = sum of all quantities for that SKU
    - Current Location   = every distinct location with its quantity, e.g. 'RC2A6X3 (105); RC2F27A (47)'
    - Previous Location  = the single most recent previous location recorded for that SKU
    """
    def summarize(g):
        loc_qty = g.groupby("Location")["Quantity"].sum().sort_values(ascending=False)
        loc_str = "; ".join(f"{loc} ({int(q)})" for loc, q in loc_qty.items() if loc)
        latest = g.loc[g["TxTime"].idxmax(), "Previous Location"]
        return pd.Series({
            "Quantity Available": g["Quantity"].sum(),
            "Current Location": loc_str,
            "Previous Location": latest,
        })

    return inv.groupby("SKU").apply(summarize, include_groups=False).reset_index()


def format_qty(qty, um):
    qty = 0 if pd.isna(qty) else qty
    qty_str = str(int(qty)) if float(qty).is_integer() else str(qty)
    um = _clean(um)
    return f"{qty_str} {um}".strip()


def load_sample_orders(file) -> pd.DataFrame:
    """Stack every non-empty sheet -> Order Date, Order #, Item No, Item Name, Order Qty"""
    xl = pd.ExcelFile(file)
    frames = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet, header=0)
        if df.empty or "Item no" not in df.columns:
            continue
        df = df.dropna(subset=["Item no"])
        if df.empty:
            continue
        order_no = pd.to_numeric(df.get("CO no"), errors="coerce")
        qty_num = pd.to_numeric(df.get("Order qty"), errors="coerce").fillna(0)
        um = df.get("U/M")
        frames.append(pd.DataFrame({
            "Order Date": sheet,
            "Order #": order_no.apply(lambda v: str(int(v)) if pd.notna(v) else ""),
            "Item No": df["Item no"].astype(str).str.strip(),
            "Item Name": df.get("Name"),
            "Order Qty Num": qty_num,
            "Order Qty": [format_qty(q, u) for q, u in zip(qty_num, um)],
        }))
    if not frames:
        return pd.DataFrame(columns=["Order Date", "Order #", "Item No", "Item Name", "Order Qty Num", "Order Qty"])
    return pd.concat(frames, ignore_index=True)


def is_sample_item(item_no: str) -> bool:
    """Sample items are coded like 'S12150' (letter prefix) instead of a numeric SKU."""
    item_no = str(item_no).strip()
    return not item_no[:1].isdigit()


def match_orders_to_inventory(orders: pd.DataFrame, tk_file) -> pd.DataFrame:
    inv_raw = load_tkreserve(tk_file)
    inv = aggregate_inventory(inv_raw)

    result = orders.merge(inv, how="left", left_on="Item No", right_on="SKU").drop(columns=["SKU"])
    result["Quantity Available"] = result["Quantity Available"].fillna(0)
    result["Current Location"] = result["Current Location"].fillna("Not found in inventory")
    result["Previous Location"] = result["Previous Location"].fillna("")

    cols = ["Order Date", "Order #", "Item No", "Item Name", "Order Qty",
            "Current Location", "Previous Location", "Quantity Available", "Order Qty Num"]
    result = result[cols]

    result["_is_sample"] = result["Item No"].apply(is_sample_item)
    result = result.sort_values(
        ["_is_sample", "Order Date", "Item No"]
    ).drop(columns=["_is_sample"]).reset_index(drop=True)
    return result


def build_excel(result: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order vs Inventory"

    display_cols = [c for c in result.columns if c != "Order Qty Num"]
    headers = display_cols
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    zero_fill = PatternFill("solid", start_color="FFF2CC")
    shortage_fill = PatternFill("solid", start_color="FFE0E0")

    for _, row in result.iterrows():
        ws.append([row[c] for c in display_cols])

    qty_avail_col = headers.index("Quantity Available") + 1
    item_no_col = headers.index("Item No") + 1

    for offset, (_, row) in enumerate(result.iterrows()):
        r = offset + 2
        item_no = row["Item No"]
        qty_avail = row["Quantity Available"]
        order_qty_num = row["Order Qty Num"]
        is_sample = str(item_no)[:1].isalpha() if item_no is not None else False
        if qty_avail == 0:
            fill = None if is_sample else zero_fill
        elif qty_avail < order_qty_num:
            fill = shortage_fill
        else:
            fill = None
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill

    widths = {"Order Date": 11, "Order #": 13, "Item No": 13, "Item Name": 40,
              "Order Qty": 10, "Quantity Available": 16, "Current Location": 30,
              "Previous Location": 16}
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 14)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
st.title("Sample Orders to Inventory Matcher")
st.write(
    "Upload the **sample orders** workbook and the **TKRESERVE_QPORT** inventory "
    "report. The app matches every ordered item to its available quantity and "
    "every location it's currently stocked in."
)

col1, col2 = st.columns(2)
with col1:
    orders_file = st.file_uploader("Sample Orders (.xlsx)", type=["xlsx"], key="orders")
with col2:
    tk_file = st.file_uploader("QPORT Inventory Report (.xlsx)", type=["xlsx"], key="tk")

if orders_file and tk_file:
    try:
        with st.spinner("Reading and matching..."):
            orders = load_sample_orders(orders_file)
            if orders.empty:
                st.error("No order rows found (expected a column named 'Item no' in the sample orders sheets).")
                st.stop()
            result = match_orders_to_inventory(orders, tk_file)
    except Exception as e:
        st.error(f"Something went wrong while processing the files: {e}")
        st.stop()

    zero_avail = (result["Quantity Available"] == 0).sum()
    short = ((result["Quantity Available"] > 0) & (result["Quantity Available"] < result["Order Qty Num"])).sum()

    m1, m2, m3 = st.columns(3)
    m1.metric("Order lines", len(result))
    m2.metric("Zero available", int(zero_avail))
    m3.metric("Short on stock", int(short))

    display_df = result.drop(columns=["Order Qty Num"])
    n_cols = len(display_df.columns)
    row_styles = []
    for _, row in result.iterrows():
        is_sample = str(row["Item No"])[:1].isalpha()
        if row["Quantity Available"] == 0:
            style = [""] * n_cols if is_sample else ["background-color: #fff2cc"] * n_cols
        elif row["Quantity Available"] < row["Order Qty Num"]:
            style = ["background-color: #ffe0e0"] * n_cols
        else:
            style = [""] * n_cols
        row_styles.append(style)

    def highlight(row):
        return row_styles[row.name]

    st.dataframe(display_df.style.apply(highlight, axis=1), use_container_width=True, height=500)

    excel_buf = build_excel(result)
    st.download_button(
        "Download results as Excel",
        data=excel_buf,
        file_name="order_vs_inventory.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload both files to run the match.")
