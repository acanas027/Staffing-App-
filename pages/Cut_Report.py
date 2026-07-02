"""
Cuts / Shorts From Loads — Rep Email Generator
------------------------------------------------
Same pattern as the "Email this report" button on the Shift Closeout page:
a mailto: link that opens your default mail app (Outlook) with the message
pre-filled. No setup, no server-side sending, no download step — you just
click "Open email", review it in Outlook, and press Send yourself.

Upload the weekly "SHORTS FROM LOADS" workbook and this app will:
  1. Read the 1ST SHIFT CUTS and 2ND SHIFT CUTS tabs
  2. Match each customer to their CS Rep + email using the
     CUSTOMER SERVICE MASTER LIST tab
  3. Build ONE email per rep with a table of every affected item
  4. Give you an "Open email" button per rep that opens it in Outlook,
     addressed, subject filled in, table in the body — ready to press Send.

Note: mailto: links only support plain text (no HTML/colors), so the table
is a monospace, column-aligned text table rather than a styled grid — it
will still line up cleanly in Outlook. mailto links also have a practical
length limit (~2000 characters); reps with a very large number of items
may get a link that's too long for some mail clients to open reliably.

Run with:  streamlit run cuts_email_generator.py
"""

import io
import re
import urllib.parse
import zipfile

import openpyxl
import pandas as pd
import pdfplumber
import streamlit as st

# --------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------
SHIFT_SHEETS = ["1ST SHIFT CUTS", "2ND SHIFT CUTS"]
MASTER_SHEET = "CUSTOMER SERVICE MASTER LIST"
MAILTO_SAFE_LENGTH = 1800  # links longer than this may fail to open in some clients

st.set_page_config(page_title="Cuts / Shorts Rep Email Generator", layout="wide")
st.title("📧 Cuts / Shorts From Loads — Rep Email Generator")
st.caption(
    "Upload the workbook → one email per Customer Service rep is built. "
    "Click Open email, it opens in Outlook ready to send — just press Send."
)

uploaded = st.file_uploader("Upload the workbook (.xlsx)", type=["xlsx"])
manifest_files = st.file_uploader(
    "Upload today's shipping manifest(s) (.zip or .pdf) — optional",
    type=["pdf", "zip"],
    accept_multiple_files=True,
    help=(
        "Used to fill in the CUSTOMER column for line items where it's blank in "
        "the workbook. Matches by ORDER NUMBER, then uses whatever customer name "
        "the manifest shows for that order. Accepts the .zip file straight from "
        "the mass-print export (PDFs inside are found automatically), a raw "
        ".pdf, or several of either if you have more than one to cover. This "
        "only covers orders that are actually on the manifest(s) you upload -- "
        "it won't resolve every unknown row."
    ),
)

# --------------------------------------------------------------------------
# HEADER MATCHING
# --------------------------------------------------------------------------
# Your actual current headers, plus a few reasonable variants, so the app
# keeps working if wording/spacing/punctuation changes slightly. Matching is
# by normalized header text (lowercased, punctuation/spaces stripped), so
# "CUSTOMER ", "Customer", and "customer" all match the same field, and any
# EXTRA column you add (anywhere in the sheet) is simply ignored.
SHIFT_FIELD_ALIASES = {
    "CS_REP": ["csrep", "csrep.", "csrepresentative"],
    "CUSTOMER": ["customer", "customername"],
    "LOAD": ["triploadnum", "tripload", "loadnumber", "load", "loadnum"],
    "ORDER_NUMBER": ["ordernumber", "orderno", "order"],
    "ITEM_NUMBER": ["itemnumber", "itemno", "item"],
    "DESCRIPTION": ["description", "desc", "itemdescription"],
    "QUANTITY_CASES_CUT": ["quantitycasescut", "qtycasescut", "casescut", "quantitycut", "qtycut"],
    "REASON_CODE": ["reasoncode"],
    "REASON_DESCRIPTION": ["reasondescription", "reasondesc"],
}
SHIFT_REQUIRED_FIELDS = ["CUSTOMER", "ORDER_NUMBER", "ITEM_NUMBER"]

MASTER_FIELD_ALIASES = {
    "REP": ["rep", "csrep", "repname", "csrepresentative"],
    "CUSTOMER": ["customer", "customername"],
    "EMAIL": ["email", "emailaddress", "repemail"],
}

# "Daily Cuts" front sheet: a simpler raw log (no CUSTOMER, no formal reason
# code) that gets expanded into the full shift-cuts row format. TRIP NUMBER
# maps to LOAD, CASE COUNT maps to QUANTITY CASES CUT, and REASON CODE here
# is actually informal shorthand text (e.g. "no inv.", "qh hold") rather than
# a formal code -- handled separately, see match_reason_code.
DAILY_CUTS_FIELD_ALIASES = {
    "TRIP_NUMBER": ["tripnumber", "trip", "load", "loadnumber"],
    "ORDER_NUMBER": ["ordernumber", "orderno", "order"],
    "ITEM_NUMBER": ["itemnumber", "itemno", "item"],
    "CASE_COUNT": ["casecount", "cases", "quantitycasescut", "qtycut"],
    "REASON_TEXT": ["reasoncode", "reason"],
}
DAILY_CUTS_REQUIRED_FIELDS = ["ORDER_NUMBER", "ITEM_NUMBER"]

ITEM_MASTER_FIELD_ALIASES = {
    "ITEM_NUMBER": ["itemnumber"],
    "DESCRIPTION": ["itemdescription", "description"],
}

# Best-effort keyword -> formal cut code mapping for the informal shorthand
# notes in Daily Cuts' REASON CODE column. This is NOT a reliable 1:1
# mapping -- checked against real data, the same shorthand (e.g. "no inv.")
# has been used for genuinely different formal codes depending on the
# specific situation, which only a person reviewing it would know. Only
# keywords with a real, checked precedent or low ambiguity are mapped here;
# anything else (including "no inv." and "lost" -- deliberately NOT mapped,
# see above) is left unmapped and flagged for manual review rather than
# guessing. Checked in order; first match wins.
REASON_KEYWORD_MAP = [
    (["short packaging", "short pack", "short coded", "shortcoded"], "5F"),
    (["qh hold", "qa hold", "quality hold"], "1Q"),
    (["reformul"], "2Q"),
    (["sched"], "2F"),
    (["equipment", "breakdown"], "3F"),
    (["batch qty", "batch quantity"], "4F"),
    (["capacity exceeded"], "0F"),
    (["damage"], "3D"),
    (["picking", "pick error"], "4D"),
    (["distribution capacity"], "5D"),
    (["receiving error"], "1D"),
    (["cancel"], "4C"),
    (["order entry"], "2C"),
    (["deleted item"], "1C"),
    (["late order"], "3C"),
    (["transfer late", "trailer late"], "2Z"),
    (["routed early"], "1Z"),
    (["substitut"], "3S"),
    (["order volume", "unusual order"], "2S"),
    (["qty too small", "quantity too small"], "1S"),
    (["shipped complete"], "2Y"),
    (["sample"], "1Y"),
    (["raw material", "material shortage"], "1R"),
    (["pkg shortage", "packaging shortage"], "2R"),
    (["broker non", "broker short/delay"], "1B"),
    (["broker purchas"], "2B"),
    (["broker", "trans truck late"], "3B"),
    (["broker short coded"], "4B"),
]


def _normalize_header(text):
    """Lowercase and strip all punctuation/whitespace for robust header matching."""
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


def _find_header_row_and_columns(ws, field_aliases, required_fields, max_scan_rows=None):
    """
    Scan rows top-to-bottom looking for the header row: the first row where at
    least all required_fields can be matched by header text. Returns
    (header_row_idx, {field_name: column_index}) or (None, {}) if not found.
    """
    alias_lookup = {}
    for field, aliases in field_aliases.items():
        for alias in aliases:
            alias_lookup[alias] = field

    max_row = max_scan_rows or ws.max_row
    for r in range(1, max_row + 1):
        col_map = {}
        for c in range(1, ws.max_column + 1):
            header_val = ws.cell(row=r, column=c).value
            normalized = _normalize_header(header_val)
            field = alias_lookup.get(normalized)
            if field and field not in col_map:
                col_map[field] = c
        if all(f in col_map for f in required_fields):
            return r, col_map

    return None, {}


# --------------------------------------------------------------------------
# PARSING HELPERS
# --------------------------------------------------------------------------
# Matches a section-header cell like "HEIDE MASON - HeideM@resers.com":
# group(1) = rep name, group(2) = email. Some newer master lists embed the
# email directly in the header text like this instead of a separate column.
NAME_EMAIL_RE = re.compile(r"^\s*(.*?)\s*-\s*([^\s]+@[^\s]+)\s*$")


def build_rep_directory(wb):
    """
    Reads the CUSTOMER SERVICE MASTER LIST sheet.

    If the sheet has a header row with REP / CUSTOMER / EMAIL-style column
    titles, columns are matched by name (any extra columns are ignored, in
    any position).

    Your master list currently has no such header row -- it uses a section
    layout instead: col A holds a section header per rep (either just the
    rep name, or "REP NAME - email@domain" with the email merged in), col B
    repeats the rep name on every row underneath, col C holds the customer
    name. If a separate email column exists further right (the older format
    for this file), that's picked up too. Whichever format has the email,
    it gets matched to whichever rep name is on that same row or the block
    it starts.

    Returns (customer_to_rep dict, rep_to_email dict).
    """
    ws = wb[MASTER_SHEET]

    header_row_idx, col_map = _find_header_row_and_columns(
        ws, MASTER_FIELD_ALIASES, required_fields=["REP", "CUSTOMER"]
    )

    customer_to_rep = {}
    rep_to_email = {}
    current_rep = None

    if header_row_idx is not None:
        rep_col = col_map.get("REP")
        cust_col = col_map.get("CUSTOMER")
        email_col = col_map.get("EMAIL")
        start_row = header_row_idx + 1
    else:
        # No header row found -- fall back to the known positional layout.
        rep_col, cust_col, email_col = 2, 3, 4  # cols B, C, D (email col may be blank)
        start_row = 1

    for r in range(start_row, ws.max_row + 1):
        section_header = ws.cell(row=r, column=1).value  # col A: section header, sparse
        rep_cell = ws.cell(row=r, column=rep_col).value if rep_col else None
        cust_cell = ws.cell(row=r, column=cust_col).value if cust_col else None
        email_cell = ws.cell(row=r, column=email_col).value if email_col else None

        # Newer format: column A holds "REP NAME - email@domain" on the
        # section-header row. Parse both the rep name and email from it.
        header_rep_name = None
        header_email = None
        if section_header and "@" in str(section_header):
            m = NAME_EMAIL_RE.match(str(section_header))
            if m:
                header_rep_name = m.group(1).strip()
                header_email = m.group(2).strip()

        if header_rep_name:
            current_rep = header_rep_name
        elif rep_cell:
            current_rep = str(rep_cell).strip()

        if header_email and current_rep and current_rep not in rep_to_email:
            rep_to_email[current_rep] = header_email

        if cust_cell and current_rep:
            customer_to_rep[str(cust_cell).strip().upper()] = current_rep

        # Older format: a separate email column.
        if email_cell and current_rep and current_rep not in rep_to_email:
            rep_to_email[current_rep] = str(email_cell).strip()

    return customer_to_rep, rep_to_email


def extract_shift_rows(wb, sheet_name):
    """
    Reads one shift-cuts sheet and returns a list of line-item dicts.

    Columns are matched by header name (CS REP, CUSTOMER, TRIP // LOAD #,
    ORDER NUMBER, ITEM NUMBER, DESCRIPTION, QUANTITY CASES CUT, REASON CODE,
    REASON DESCRIPTION) rather than fixed position, so extra columns can be
    added anywhere in the sheet without breaking anything.

    LOAD# forward-fills down (a single physical load can carry several
    orders, and the load number is only typed once). CUSTOMER forward-
    fills ONLY while the ORDER NUMBER stays the same as the row above
    (true multi-line continuation of one order) -- it is NOT carried
    across a change in order number, since that would be guessing a
    customer that was never actually entered. Rows where a new order
    number appears with no customer typed in are labelled "UNKNOWN"
    and surfaced separately so they can be fixed in the source file.
    """
    ws = wb[sheet_name]

    header_row_idx, col_map = _find_header_row_and_columns(
        ws, SHIFT_FIELD_ALIASES, SHIFT_REQUIRED_FIELDS
    )
    if header_row_idx is None:
        return []

    def get(r, field):
        col = col_map.get(field)
        return ws.cell(row=r, column=col).value if col else None

    rows = []
    last_load = None
    last_customer = None
    last_order = None

    for r in range(header_row_idx + 1, ws.max_row + 1):
        customer = get(r, "CUSTOMER")
        load = get(r, "LOAD")
        order_no = get(r, "ORDER_NUMBER")
        item_no = get(r, "ITEM_NUMBER")
        desc = get(r, "DESCRIPTION")
        qty = get(r, "QUANTITY_CASES_CUT")
        reason_code = get(r, "REASON_CODE")
        reason_desc = get(r, "REASON_DESCRIPTION")

        # Skip fully blank / padding rows
        if order_no in (None, "") and item_no in (None, "") and not customer and not load:
            continue

        if load:
            last_load = load

        if customer:
            last_customer = str(customer).strip()
        elif order_no != last_order:
            # New order number with no customer typed in -> genuinely unknown
            last_customer = None

        last_order = order_no

        if order_no in (None, ""):
            continue  # not a real line item

        rows.append(
            {
                "Shift": sheet_name.replace(" CUTS", "").title(),
                "CUSTOMER": last_customer or "UNKNOWN",
                "LOAD": last_load,
                "ORDER NUMBER": order_no,
                "ITEM NUMBER": item_no,
                "DESCRIPTION": desc if desc not in (None, "") else "",
                "QUANTITY CASES CUT": qty,
                "REASON CODE": reason_code,
                "REASON DESCRIPTION": reason_desc,
            }
        )

    return rows


def _is_blank_cell(value):
    """True for None or a string that's empty/whitespace-only (Daily Cuts uses ' ' as a spacer)."""
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_item_number(value):
    """Strips a leading/trailing '*' marker and whitespace, preserving the rest as-is."""
    if _is_blank_cell(value):
        return ""
    return str(value).strip().strip("*").strip()


def normalize_item_key(value):
    """
    Normalizes an item number for matching against the Item Master List.
    Both sheets may store the same item as a float (e.g. 71117.1482) or a
    string with a leading zero / trailing zero / asterisk (e.g.
    "*06795.28681" vs 6795.28681) -- parsing as a float and formatting with
    fixed precision makes those line up. Falls back to an uppercased string
    if it isn't numeric at all.
    """
    text = clean_item_number(value)
    if not text:
        return ""
    try:
        return f"{float(text):.5f}"
    except (ValueError, TypeError):
        return text.upper()


def build_item_description_lookup(wb, sheet_name="ITEM MASTER LIST"):
    """Reads the Item Master List sheet and returns {normalized_item_key: description}."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header_row_idx, col_map = _find_header_row_and_columns(
        ws, ITEM_MASTER_FIELD_ALIASES, ["ITEM_NUMBER", "DESCRIPTION"]
    )
    if header_row_idx is None:
        return {}

    lookup = {}
    item_col = col_map["ITEM_NUMBER"]
    desc_col = col_map["DESCRIPTION"]
    for r in range(header_row_idx + 1, ws.max_row + 1):
        item_val = ws.cell(row=r, column=item_col).value
        desc_val = ws.cell(row=r, column=desc_col).value
        key = normalize_item_key(item_val)
        if key and desc_val not in (None, ""):
            lookup.setdefault(key, str(desc_val).strip())
    return lookup


def build_cut_codes_lookup(wb, sheet_name="CUT CODES MASTER LIST"):
    """
    Reads the Cut Codes Master List sheet and returns {code: description}.
    This sheet has no header row -- it's laid out as section labels (e.g.
    "PLANT") followed by rows of (code, description) in two adjacent
    columns, so this scans for the two columns holding that pattern rather
    than matching a header.
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    code_re = re.compile(r"^\d[A-Z]$|^[A-Z]\d$")
    lookup = {}
    for row in ws.iter_rows():
        for i in range(len(row) - 1):
            code_val = row[i].value
            desc_val = row[i + 1].value
            if (
                isinstance(code_val, str)
                and code_re.match(code_val.strip())
                and isinstance(desc_val, str)
                and desc_val.strip()
            ):
                lookup.setdefault(code_val.strip().upper(), desc_val.strip())
    return lookup


def match_reason_code(raw_text, cut_codes_lookup):
    """
    Best-effort match of Daily Cuts' informal shorthand reason text (e.g.
    "qh hold", "short packaging") to a formal cut code, using
    REASON_KEYWORD_MAP. This is NOT guaranteed accurate -- see the caveat on
    REASON_KEYWORD_MAP. Returns (code, description, matched: bool). When
    unmatched, code/description are ("", "") and the raw text should be
    preserved separately so nothing is lost.
    """
    text = (raw_text or "").strip().lower()
    if not text:
        return "", "", False
    for keywords, code in REASON_KEYWORD_MAP:
        if any(kw in text for kw in keywords):
            desc = cut_codes_lookup.get(code, "")
            return code, desc, True
    return "", "", False


def parse_daily_cuts_sheet(wb):
    """
    Parses the "Daily Cuts" front sheet (always the FIRST worksheet in the
    workbook, whatever it's named) into the same row-dict shape that
    extract_shift_rows produces for the 1ST/2ND SHIFT CUTS sheets, so the
    rest of the pipeline (manifest matching, rep lookup, email generation)
    doesn't need to know which source it came from.

    Layout: TRIP NUMBER, ORDER NUMBER, ITEM NUMBER, CASE COUNT, REASON CODE
    (informal shorthand, not a formal code -- see match_reason_code). Blank
    rows and single-space "spacer" rows separate order groups; TRIP NUMBER
    forward-fills within a group the same way LOAD does elsewhere. There's
    no shift column -- everything is 1st shift until a row containing a
    cell that starts with "2ND" (case-insensitive) is found, after which
    everything is 2nd shift. CUSTOMER isn't available here at all yet --
    every row starts as UNKNOWN, to be filled in later by the shipping
    manifest step exactly like the other source format.

    Returns None if the first sheet doesn't look like this format at all
    (missing ORDER NUMBER / ITEM NUMBER columns), so the caller can fall
    back to reading 1ST/2ND SHIFT CUTS sheets directly for older workbooks.
    """
    ws = wb.worksheets[0]
    header_row_idx, col_map = _find_header_row_and_columns(
        ws, DAILY_CUTS_FIELD_ALIASES, DAILY_CUTS_REQUIRED_FIELDS
    )
    if header_row_idx is None:
        return None

    item_desc_lookup = build_item_description_lookup(wb)
    cut_codes_lookup = build_cut_codes_lookup(wb)

    trip_col = col_map.get("TRIP_NUMBER")
    order_col = col_map["ORDER_NUMBER"]
    item_col = col_map["ITEM_NUMBER"]
    case_col = col_map.get("CASE_COUNT")
    reason_col = col_map.get("REASON_TEXT")

    rows = []
    last_load = None
    last_order = None
    current_shift = "1St Shift"

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(
            isinstance(v, str) and re.match(r"^\s*2nd\b", v, re.IGNORECASE)
            for v in row_values
        ):
            current_shift = "2Nd Shift"
            continue

        trip = ws.cell(row=r, column=trip_col).value if trip_col else None
        order_no = ws.cell(row=r, column=order_col).value
        item_no_raw = ws.cell(row=r, column=item_col).value
        case_count = ws.cell(row=r, column=case_col).value if case_col else None
        reason_text = ws.cell(row=r, column=reason_col).value if reason_col else None

        if all(_is_blank_cell(v) for v in (trip, order_no, item_no_raw, case_count, reason_text)):
            last_order = None  # a spacer row ends the current order group
            continue

        if not _is_blank_cell(trip):
            last_load = trip

        if not _is_blank_cell(order_no):
            last_order = order_no

        if _is_blank_cell(item_no_raw) or last_order is None:
            continue  # not a real line item

        item_no = clean_item_number(item_no_raw)
        item_key = normalize_item_key(item_no)
        description = item_desc_lookup.get(item_key, "Not Found")

        raw_reason = "" if _is_blank_cell(reason_text) else str(reason_text).strip()
        code, code_desc, matched = match_reason_code(raw_reason, cut_codes_lookup)
        if matched:
            reason_code_out = code
            reason_description_out = code_desc
        else:
            reason_code_out = ""
            reason_description_out = f"NEEDS REVIEW - raw note: {raw_reason}" if raw_reason else ""

        rows.append(
            {
                "Shift": current_shift,
                "CUSTOMER": "UNKNOWN",
                "LOAD": last_load,
                "ORDER NUMBER": last_order,
                "ITEM NUMBER": item_no,
                "DESCRIPTION": description,
                "QUANTITY CASES CUT": case_count,
                "REASON CODE": reason_code_out,
                "REASON DESCRIPTION": reason_description_out,
            }
        )

    return rows


# --------------------------------------------------------------------------
# SHIPPING MANIFEST (PDF) PARSING
# --------------------------------------------------------------------------
# Matches a stop-header line like:
#   "2           213962136                                   F AND A FOOD SALES INC"
# group(1) = stop number, group(2) = a numeric CUSTID or a short facility
# code (e.g. "TK", "NC", "DC" -- our own warehouses use these instead of a
# CUSTID), group(3) = the customer/location name.
#
# Internal transfers between our own warehouses show up as a stop whose name
# starts with "RESER'S -" (e.g. "RESER'S - TK - TOPEKA DISTRIBUTION CENTER",
# "RESER'S - NC - HALIFAX DISTRIBUTION CENTER") regardless of which facility
# code is used, and are treated as non-customer stops (skipped for matching).
#
# Some stops are also the CARRIER's own cross-dock/relay terminal rather than
# a real customer (e.g. "FRZF - 4   FFE - Butler, MO", where "FRZF" is also
# the load's own carrier code from the CARR/SCT TR line) -- these are
# likewise skipped once the current load's carrier code is known.
MANIFEST_STOP_HEADER_RE = re.compile(
    r"^\s*(\d{1,2})\s+(\d{4,}|[A-Z]{2,4})\s+(?:-\s*\d+\s+)?(.+?)\s*$"
)
MANIFEST_CARRIER_RE = re.compile(r"^\s*CARR/SCT\s+TR:\s*([A-Z0-9]{2,6})\b", re.IGNORECASE)

# Matches an order number with its trailing load-sequence suffix, e.g.
# "3600012594-520" -> captures "3600012594".
MANIFEST_ORDER_RE = re.compile(r"\b(\d{6,12})-\d+\b")
MANIFEST_ORDER_ANCHOR_RE = re.compile(r"^(\d{6,12})-\d+$")

# Column x0 boundaries (points) for the CUSTID / CUST NAME / CUST PO block on
# the per-order detail line, calibrated against the actual report layout.
# Used only as a fallback (see parse_manifest_inline_fallback below) for
# loads where the true customer never gets its own clean stop-header line --
# e.g. a load that transfers through one of our own warehouses on the way,
# where the only place the ultimate customer is mentioned at all is this
# narrow, easily-wrapped column on the pickup stop's own line.
INLINE_CUSTID_X_RANGE = (395, 445)
INLINE_NAME_X_RANGE = (445, 505)
INLINE_ORDER_COL_MAX_X = 200

# The CUST NAME column hard-wraps at exactly this many characters per physical
# line (calibrated from the actual report: "DISTRIBUTO" and "PERISHABLE" both
# hit exactly 10 and are the longest fragments observed). A fragment that
# reaches this length, followed immediately by another alphabetic fragment,
# is almost certainly one word split mid-way by the wrap (e.g. "DISTRIBUTO"
# + "RS" -> "DISTRIBUTORS") rather than two separate words, so those get
# joined with no space instead of the usual space-separated join.
NAME_WRAP_CHAR_LIMIT = 10


def normalize_order_id(value):
    """
    Normalize an order number for matching between the workbook (numbers,
    sometimes floats like 3000131623.0) and the manifest (plain digit
    strings). Returns '' if it can't be parsed as a number at all.
    """
    if value is None or value == "":
        return ""
    try:
        f = float(value)
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    digits = re.sub(r"[^0-9]", "", str(value))
    return digits


def clean_customer_name(name):
    """
    Strips numbers out of a resolved customer name -- store numbers, or
    stray CUSTID/PO digits that leaked in from a neighboring column during
    reconstruction. Removes any token that contains a digit at all (covers
    plain numbers like "6072" as well as store/location codes like "T3727"
    or "DC5"), then tidies up any leftover dangling dashes/punctuation and
    extra whitespace.
    """
    if not name:
        return name
    tokens = name.split()
    kept = [t for t in tokens if not any(ch.isdigit() for ch in t)]
    cleaned = " ".join(kept)
    cleaned = re.sub(r"[-,]\s*$", "", cleaned).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned or name  # never return an empty string -- fall back to original


def join_name_fragments(fragments):
    """
    Joins wrapped CUST NAME fragments (each already known to be pure text,
    sorted top-to-bottom) into one name. A fragment that hit the column's
    hard-wrap character limit, followed by another alphabetic fragment, is
    treated as a mid-word split and joined with no space; otherwise a space
    is inserted, matching normal word-by-word wrapping.

    Checks each fragment's OWN original length (not the cumulative merged
    result) against the wrap limit -- otherwise a short fragment appended
    onto an already-long merged word (e.g. "SPRINGFIEL" + "D" -> "SPRINGFIELD",
    now 11 characters) would incorrectly look "at the limit" again and
    swallow the next real word too (producing "SPRINGFIELDDELI").
    """
    if not fragments:
        return ""
    result = fragments[0]
    prev_len = len(fragments[0])
    for frag in fragments[1:]:
        if prev_len >= NAME_WRAP_CHAR_LIMIT and frag.isalpha():
            result += frag  # mid-word wrap -- no space
        else:
            result += " " + frag
        prev_len = len(frag)
    return result


def parse_manifest_inline_fallback(pdf_file):
    """
    Fallback parser for orders that never get a clean stop-header line --
    typically loads that transfer through one of our own warehouses (or a
    carrier's own relay point) en route, where the ultimate customer is only
    ever mentioned inline on the per-order detail line's narrow CUST NAME
    column (which wraps across several short physical lines, occasionally
    splitting a single word mid-way -- see join_name_fragments).

    Uses word x/y positions (not flattened text) to reconstruct each order's
    CUST NAME column specifically, ignoring what's happening in neighboring
    columns on the same wrapped lines. Numbers (store numbers, or stray
    CUSTID/PO digits) are stripped from the result.

    Returns {order_number_str: reconstructed_name}.
    """
    pdf_file.seek(0)
    order_to_name = {}

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))

            anchors = [
                (w["top"], w["text"])
                for w in words_sorted
                if w["x0"] < INLINE_ORDER_COL_MAX_X and MANIFEST_ORDER_ANCHOR_RE.match(w["text"])
            ]

            for idx, (top, order_text) in enumerate(anchors):
                order_no = order_text.split("-")[0]
                next_top = anchors[idx + 1][0] if idx + 1 < len(anchors) else top + 40
                block = [w for w in words_sorted if top - 0.5 <= w["top"] < next_top - 0.5]

                name_words = [
                    w for w in block
                    if INLINE_NAME_X_RANGE[0] <= w["x0"] < INLINE_NAME_X_RANGE[1]
                ]
                if name_words:
                    fragments = [w["text"] for w in sorted(name_words, key=lambda w: w["top"])]
                    name = clean_customer_name(join_name_fragments(fragments))
                    if name:
                        order_to_name.setdefault(order_no, name)

    return order_to_name


def parse_shipping_manifest(pdf_file):
    """
    Parses a Resers-format shipping manifest PDF and returns
    (order_to_customer, order_confidence):
      order_to_customer: {order_number_str: customer_name}
      order_confidence:  {order_number_str: "exact" | "inline"}

    Primary method: each load's stops are read in order, tracking the load's
    own carrier code from its "CARR/SCT TR:" line. A stop is treated as a
    non-customer (skipped) if its name starts with "RESER'S -" (one of our
    own warehouses) or if its facility code matches the load's own carrier
    code (the carrier's own cross-dock/relay terminal, not a real customer).
    Any other stop's header line gives the real customer name (marked
    "exact"); every order number appearing before the next stop header is
    attributed to it.

    Fallback: some loads never have a real external stop at all in this
    manifest (the actual delivery is a separate, later leg not included
    here). For orders left unresolved by the primary method,
    parse_manifest_inline_fallback recovers a best-effort name from the
    per-order line's own narrow CUST NAME column (marked "inline" -- see
    that function's docstring for the accuracy trade-off).
    """
    pdf_file.seek(0)
    order_to_customer = {}
    order_confidence = {}
    current_customer = None
    current_carrier_code = None

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text(layout=True) or ""
            for line in text.split("\n"):
                carrier_m = MANIFEST_CARRIER_RE.match(line)
                if carrier_m:
                    current_carrier_code = carrier_m.group(1).upper()
                    continue

                m = MANIFEST_STOP_HEADER_RE.match(line)
                if m:
                    stop_code, rest = m.group(2), m.group(3)
                    rest_clean = re.sub(r"\s+", " ", rest.strip())
                    is_internal = rest_clean.upper().startswith("RESER'S")
                    is_carrier_relay = (
                        current_carrier_code is not None
                        and stop_code.upper() == current_carrier_code
                    )
                    if is_internal or is_carrier_relay:
                        current_customer = None
                    else:
                        current_customer = clean_customer_name(rest_clean)
                    continue

                for order_match in MANIFEST_ORDER_RE.finditer(line):
                    order_no = order_match.group(1)
                    if current_customer:
                        order_to_customer.setdefault(order_no, current_customer)
                        order_confidence.setdefault(order_no, "exact")

    pdf_file.seek(0)
    inline_map = parse_manifest_inline_fallback(pdf_file)
    for order_no, name in inline_map.items():
        if order_no not in order_to_customer:
            order_to_customer[order_no] = name
            order_confidence[order_no] = "inline"

    return order_to_customer, order_confidence


def load_manifest_maps(uploaded_files):
    """
    Takes the list of files from the manifest uploader (each may be a .pdf or
    a .zip containing one or more PDFs) and returns:
      (combined_order_to_customer_map, combined_order_confidence,
       list_of_per_file_summaries, list_of_errors)

    Each summary is a dict: {"name": filename, "pdf_count": n, "orders_found": n}.
    Each error is a string naming the file and what went wrong -- one bad file
    doesn't stop the others from being processed.
    """
    combined_map = {}
    combined_confidence = {}
    summaries = []
    errors = []

    def merge(sub_map, sub_confidence):
        for k, v in sub_map.items():
            if k not in combined_map:
                combined_map[k] = v
                combined_confidence[k] = sub_confidence.get(k, "exact")

    for uploaded_file in uploaded_files or []:
        name = uploaded_file.name
        try:
            uploaded_file.seek(0)
            if name.lower().endswith(".zip"):
                pdf_count = 0
                file_orders_found = 0
                with zipfile.ZipFile(uploaded_file) as zf:
                    pdf_names = [n for n in zf.namelist() if n.lower().endswith(".pdf")]
                    if not pdf_names:
                        errors.append(f"{name}: no PDF files found inside this zip.")
                        continue
                    for pdf_name in pdf_names:
                        pdf_bytes = zf.read(pdf_name)
                        sub_map, sub_confidence = parse_shipping_manifest(io.BytesIO(pdf_bytes))
                        pdf_count += 1
                        file_orders_found += len(sub_map)
                        merge(sub_map, sub_confidence)
                summaries.append(
                    {"name": name, "pdf_count": pdf_count, "orders_found": file_orders_found}
                )
            else:
                # Treat anything else (e.g. .pdf) as a single PDF.
                sub_map, sub_confidence = parse_shipping_manifest(uploaded_file)
                merge(sub_map, sub_confidence)
                summaries.append(
                    {"name": name, "pdf_count": 1, "orders_found": len(sub_map)}
                )
        except Exception as e:
            errors.append(f"{name}: {e}")

    return combined_map, combined_confidence, summaries, errors


def apply_manifest_customers(df, manifest_map, manifest_confidence=None):
    """
    Fills in CUSTOMER for rows currently marked UNKNOWN, using the manifest's
    order-number -> customer-name mapping. Adds two columns so the UI can
    show what happened:
      FROM_MANIFEST: True if this row's customer came from the manifest
      MANIFEST_CONFIDENCE: "exact" (from a clean stop-header line) or
        "inline" (best-effort reconstruction -- see
        parse_manifest_inline_fallback docstring) for manifest-filled rows,
        "" otherwise.
    Rows that already have a customer name are left untouched.
    """
    manifest_confidence = manifest_confidence or {}
    df = df.copy()
    from_manifest = []
    confidence = []
    new_customers = []

    for _, row in df.iterrows():
        if row["CUSTOMER"] != "UNKNOWN":
            new_customers.append(row["CUSTOMER"])
            from_manifest.append(False)
            confidence.append("")
            continue
        key = normalize_order_id(row["ORDER NUMBER"])
        match = manifest_map.get(key) if key else None
        if match:
            new_customers.append(match)
            from_manifest.append(True)
            confidence.append(manifest_confidence.get(key, "exact"))
        else:
            new_customers.append(row["CUSTOMER"])
            from_manifest.append(False)
            confidence.append("")

    df["CUSTOMER"] = new_customers
    df["FROM_MANIFEST"] = from_manifest
    df["MANIFEST_CONFIDENCE"] = confidence
    return df


# --------------------------------------------------------------------------
# EMAIL BUILDING
# --------------------------------------------------------------------------
TABLE_COLS = [
    "CUSTOMER", "LOAD", "ORDER NUMBER", "ITEM NUMBER", "DESCRIPTION",
    "QUANTITY CASES CUT", "REASON CODE", "REASON DESCRIPTION",
]


def build_plain_text_table(df):
    """
    Plain-text item list for the mailto body. Customer/Load/Order repeat across
    multiple line items in the same shipment, so instead of repeating that on
    every item (harder to scan), each distinct Customer/Load/Order combination
    gets one header, with its items listed underneath. Outlook's compose window
    uses a proportional font, so this avoids relying on column-width alignment
    (which drifts) and stays compact enough to fit more items under the mailto
    link's length limit.
    """
    def val(row, col):
        v = row[col]
        return "" if pd.isna(v) else str(v)

    lines = []
    grouped = df.groupby(["CUSTOMER", "LOAD", "ORDER NUMBER"], sort=False, dropna=False)
    for (customer, load, order), sub in grouped:
        header = f"{customer}   |   Load #: {load}   |   Order #: {order}"
        rule = "=" * min(len(header), 60)
        lines.append(rule)
        lines.append(header)
        lines.append(rule)

        for _, row in sub.iterrows():
            reason_code = val(row, "REASON CODE")
            reason_desc = val(row, "REASON DESCRIPTION")
            reason = f"{reason_code} - {reason_desc}" if reason_code and reason_desc else (reason_code or reason_desc)

            lines.append(f"  Item #:      {val(row, 'ITEM NUMBER')}")
            lines.append(f"  Description: {val(row, 'DESCRIPTION')}")
            lines.append(f"  Qty Cut:     {val(row, 'QUANTITY CASES CUT')}")
            lines.append(f"  Reason:      {reason}")
            lines.append("")

    return "\n".join(lines).rstrip()


def build_subject(group):
    order_numbers = sorted({str(x) for x in group["ORDER NUMBER"]})
    customers = sorted({str(x) for x in group["CUSTOMER"]})
    loads = sorted({str(x) for x in group["LOAD"] if x not in (None, "")})
    return (
        f"{', '.join(order_numbers)} - CUT - {', '.join(customers)} "
        f"- Load# {', '.join(loads) if loads else 'N/A'}"
    )


def build_mailto(to_addr, subject, body):
    """Same helper as the Shift Closeout page: builds a mailto: link."""
    params = urllib.parse.urlencode(
        {"subject": subject, "body": body},
        quote_via=urllib.parse.quote,
    )
    return f"mailto:{(to_addr or '').strip()}?{params}"


# --------------------------------------------------------------------------
# MAIN APP
# --------------------------------------------------------------------------
if uploaded:
    wb = openpyxl.load_workbook(uploaded, data_only=True)

    if MASTER_SHEET not in wb.sheetnames:
        st.error(f"Missing expected sheet in this workbook: {MASTER_SHEET}")
        st.stop()

    customer_to_rep, rep_to_email = build_rep_directory(wb)

    all_rows = parse_daily_cuts_sheet(wb)
    if all_rows is not None:
        st.caption(
            f"Read line items from the '{wb.worksheets[0].title}' sheet (first tab) "
            f"— {len(all_rows)} item(s) found, expanded with item descriptions and "
            "cut-code lookups."
        )
    else:
        missing = [s for s in SHIFT_SHEETS if s not in wb.sheetnames]
        if missing:
            st.error(
                f"The first sheet isn't in the expected Daily Cuts format, and this "
                f"workbook is also missing: {missing}. Can't find any line items to process."
            )
            st.stop()
        all_rows = []
        for sheet in SHIFT_SHEETS:
            all_rows.extend(extract_shift_rows(wb, sheet))

    if not all_rows:
        st.warning("No line items found.")
        st.stop()

    df = pd.DataFrame(all_rows)
    df["FROM_MANIFEST"] = False
    df["MANIFEST_CONFIDENCE"] = ""

    manifest_map, manifest_confidence, manifest_summaries, manifest_errors = load_manifest_maps(
        manifest_files
    )

    for err in manifest_errors:
        st.error(f"Could not read manifest file — {err}")

    if manifest_summaries:
        detail = "; ".join(
            f"{s['name']} ({s['pdf_count']} PDF{'s' if s['pdf_count'] != 1 else ''}, "
            f"{s['orders_found']} order(s))"
            for s in manifest_summaries
        )
        st.caption(f"Manifest file(s) read: {detail}")

    if manifest_map:
        unknown_before = int((df["CUSTOMER"] == "UNKNOWN").sum())
        df = apply_manifest_customers(df, manifest_map, manifest_confidence)
        resolved = int(df["FROM_MANIFEST"].sum())
        inline_count = int((df["MANIFEST_CONFIDENCE"] == "inline").sum())
        msg = (
            f"Shipping manifest(s) loaded: {len(manifest_map)} unique order(s) found "
            f"across all of them. Filled in the customer for {resolved} of "
            f"{unknown_before} previously unknown line item(s)."
        )
        if inline_count:
            msg += (
                f" {inline_count} of those came from a best-effort reconstruction "
                "(orders that transfer through one of our own warehouses with no "
                "external stop in this manifest) — double-check those before "
                "trusting the auto-match, spacing can be slightly off."
            )
        st.success(msg)
    elif manifest_files and not manifest_errors:
        st.warning(
            "The shipping manifest(s) were read, but no order numbers could be "
            "parsed — double check they're the standard Resers shipping manifest "
            "format."
        )

    df["REP"] = df["CUSTOMER"].apply(
        lambda c: customer_to_rep.get(str(c).strip().upper())
    )
    df["EMAIL"] = df["REP"].map(rep_to_email)

    matched = df[df["REP"].notna()].copy()
    unmatched = df[df["REP"].isna()].copy()

    st.subheader(
        f"Found {len(df)} affected line item(s) — "
        f"{len(matched)} matched to a rep, {len(unmatched)} need review"
    )

    if not unmatched.empty:
        st.warning(
            f"{len(unmatched)} line item(s) couldn't be matched to a rep "
            "(customer is UNKNOWN, or the customer name doesn't exactly match "
            "an entry in the master list). No email is generated for these — "
            "fix the CUSTOMER column (or the master list) and re-upload, or "
            "review manually below."
        )
        st.dataframe(
            unmatched[
                ["Shift", "CUSTOMER", "LOAD", "ORDER NUMBER", "ITEM NUMBER",
                 "DESCRIPTION", "FROM_MANIFEST", "MANIFEST_CONFIDENCE"]
            ],
            use_container_width=True,
        )

    st.markdown("---")
    st.subheader("Emails")

    if matched.empty:
        st.info("No rows matched a rep yet — nothing to generate.")
    else:
        for rep, group in matched.groupby("REP"):
            email_addr = rep_to_email.get(rep, "")
            subject = build_subject(group)
            plain_table = build_plain_text_table(group)
            body = f"{plain_table}\n\nThank you."
            mailto_link = build_mailto(email_addr, subject, body)
            too_long = len(mailto_link) > MAILTO_SAFE_LENGTH

            with st.expander(
                f"✉️  {rep}  —  {email_addr or '⚠️ NO EMAIL ON FILE'}  "
                f"({len(group)} item{'s' if len(group) != 1 else ''})"
            ):
                st.text_input("To", value=email_addr, disabled=True, key=f"to_{rep}")
                st.text_input("Subject", value=subject, disabled=True, key=f"subj_{rep}")
                st.code(body, language=None)

                if not email_addr:
                    st.error("No email on file for this rep — add one to the master list.")
                else:
                    if too_long:
                        st.warning(
                            f"This email has {len(group)} items — the link is "
                            f"{len(mailto_link)} characters, above the ~1800-character "
                            "range some mail clients handle reliably. Try opening it "
                            "anyway; if Outlook doesn't open or the body looks cut off, "
                            "let me know and I'll add a way to split large reps into "
                            "multiple emails."
                        )
                    st.link_button("📤 Open email (ready to send in Outlook)", mailto_link)
