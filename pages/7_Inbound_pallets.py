import streamlit as st
import pandas as pd
from io import BytesIO
from copy import copy
import base64
import re
from openpyxl import load_workbook

st.set_page_config(page_title="Inbound Pallets", layout="wide")

st.title("📦 Pallets per Trailer")

st.write(
    "Upload your inbound report. Each **LPN** counts as one pallet, and the "
    "**trailer number** is columns C, D, E, F and G combined. The app also "
    "creates a Transfer Trailer Log using the embedded template."
)

uploaded = st.file_uploader("Upload your Excel file", type=["xlsx", "xlsm"])

# -------------------------------------------------------------------
# Embedded Transfer Trailer Log template
# -------------------------------------------------------------------
# This is the uploaded file: Transfer Log New 8-2025.xlsx
TRANSFER_LOG_TEMPLATE_BASE64 = """
UEsDBBQABgAIAAAAIQDTXImligEAAK8FAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsVMluwjAQvVfqP0S+VsTQQ1VVBA5dji0S
9ANMPCSGxLY8w/b3nZhFFWIRgksSx5733jzPTLe/qqtkAQGNs5nopG2RgM2dNrbIxO/oq/UqEiRltaqchUysAUW/
9/jQHa09YMLRFjNREvk3KTEvoVaYOg+WdyYu1Ip4GQrpVT5TBcjndvtF5s4SWGpRgyF63Q+YqHlFyeeKf2+UjI0V
yfvmXEOVCeV9ZXJFLFQurD4gabnJxOSgXT6vGTpFH0BpLAGorlIfDDOGIRBxYijkUc6pLw44Td1onnooToQEqPA6
nVsjUo6MuWBpPD6xWycYmp3TRmzjfvgGg9GQDFSgb1WzXXJVyaULs7Fzs/Q8yLVuRlfTWhm7032GPx5GGV+dOwtp
8ovAF3QQlyXI+LxdQoS5QIi0rgDvbXsEvcRcqgB6SFzwxd0F/Me+oEMHtWwkyO3H7b5vgc7xcvcPgvPIAybA9e7v
WrOJbnkGgkAG9s15rMj3jDydbr5uaMafBn2EW8Zx2/sDAAD//wMAUEsDBBQABgAIAAAAIQC1VTAj9AAAAEwCAAAL
AAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAArJJNT8MwDIbvSPyHyPfV3ZAQQkt3QUi7IVR+gEncD7WNoyQb3b8nHBBUGoMDR3+9fvzK2908jerIIfbi
NKyLEhQ7I7Z3rYaX+nF1ByomcpZGcazhxBF21fXV9plHSnkodr2PKqu4qKFLyd8jRtPxRLEQzy5XGgkTpRyGFj2Z
gVrGTVneYviuAdVCU+2thrC3N6Dqk8+bf9eWpukNP4g5TOzSmRXIc2Jn2a58yGwh9fkaVVNoOWmwYp5yOiJ5X2Rs
wPNEm78T/XwtTpzIUiI0Evgyz0fHJaD1f1q0NPHLnXnENwnDq8jwyYKLH6jeAQAA//8DAFBLAwQUAAYACAAAACEA
q29rIHwDAADGCAAADwAAAHhsL3dvcmtib29rLnhtbKxVbW+jOBD+ftL9B8R3ik0wBNR0Fd50ldpVlWbbO6lS5YIp
VgHnjGlSVfvfd0xC2m5Op1z3osSOPcPjZ2aeMadfNk1tPDPZcdHOTHyCTIO1uSh4+zgzvy0za2oanaJtQWvRspn5
wjrzy9nvv52uhXx6EOLJAIC2m5mVUqvQtru8Yg3tTsSKtWAphWyogqV8tLuVZLToKsZUU9sOQp7dUN6aW4RQHoMh
ypLnLBF537BWbUEkq6kC+l3FV92I1uTHwDVUPvUrKxfNCiAeeM3VywBqGk0enj+2QtKHGsLeYGJsJHw9+GEEgzOe
BKaDoxqeS9GJUp0AtL0lfRA/RjbGH1KwOczBcUiuLdkz1zXcs5LeJ1l5eyzvDQyjX0bDIK1BKyEk75NoZM/NMc9O
S16zm610DbpafaWNrlRtGjXtVFpwxYqZ6cNSrNmHDdmvop7XYHUC7PimfbaX85U0ClbSvlZLEPIID47ImSCkPUEY
81ox2VLFYtEq0OEurl/V3IAdVwIUbizY3z2XDBoL9AWxwkjzkD50V1RVRi/rmRmHd986CP9u3hbgGdP2LhHrthbQ
ZHfv1EkPW+E/6JPmOmgbot4y2/7/OQNAUIajBq+UNOD/eXIBdbimz1AVqH2xa9pzSDue3Le5DPH9a5zOSRxMYitD
sW+5KcHWlJCp5fh+RLI0dgn2vkMw0gtzQXtV7QquoWemC9U9MF3SzWjBKOx58UbjFe0+lp5/Gkbbdx2wvtpuOFt3
b9LQS2Nzy9tCrCEEAjG9jKuJS0xjPZhueaEqrSykFb/d+4Pxxwr4YoIROEIDaF4z8zXxPZcEMbL8KQks13UcaxrE
U8snXoK9mERo4gx87HeEhisUiA2z0Q6yv9bXKoa7Ws9Dik1DhvoMeV7goYTjYzmtc5C5ngbHACMn0B5soy46Ncyg
MA70sIvmPgpcC6UTYrnTAOi5E8eK3cRJiZ8maUR0dfQrIPw/LsJB6OH4btEsKyrVUtL8Cd5IC1ZGtAM5bQMCvu/J
RmQK6QKKboYzy8UBsqLIcy2SZBPi4yROSfZGVodffvIamtrD04yqHhpPd+ewDvWY7Xb3m+V2Y1enD50XLhKd993T
/+Z4DdHX7Ejn7OZIx/jr5fLySN+LdHl/mx3rPL+Mkvnx/vPFYv7XMv1zPML+x4TaQ8H1OMjUHmVy9gMAAP//AwBQ
SwMEFAAGAAgAAAAhAIE+lJfzAAAAugIAABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKxSTUvEMBC9C/6H
MHebdhUR2XQvIuxV6w8IybQp2yYhM3703xsqul1Y1ksvA2+Gee/Nx3b3NQ7iAxP1wSuoihIEehNs7zsFb83zzQMI
Yu2tHoJHBRMS7Orrq+0LDppzE7k+ksgsnhQ45vgoJRmHo6YiRPS50oY0as4wdTJqc9Adyk1Z3su05ID6hFPsrYK0
t7cgmilm5f+5Q9v2Bp+CeR/R8xkJSTwNeQDR6NQhK/jBRfYI8rz8Zk15zmvBo/oM5RyrSx6qNT18hnQgh8hHH38p
knPlopm7Ve/hdEL7yim/2/Isy/TvZuTJx9XfAAAA//8DAFBLAwQUAAYACAAAACEAJm/Cq3AJAADEMwAAGAAAAHhs
L3dvcmtzaGVldHMvc2hlZXQxLnhtbJxUa2/aMBT9Pmn/IfJ3kjjkiaAVg6FVmqZpz8/GccAijjPbFKpq/33XJoFO
nUZaCbg3ic8593HC9PYoau+eKc1lM0PYD5HHGipL3mxm6Pu31ShHnjakKUktGzZDD0yj25u3b6YHqXZ6y5jxgKHR
M7Q1pp0EgaZbJoj2ZcsaeFJJJYiBS7UJdKsYKR1I1EEUhmkgCG/QiWGihnDIquKULSXdC9aYE4liNTFQv97yVvds
gg6hE0Tt9u2IStECxZrX3Dw4UuQJOrnbNFKRdQ19H3FMqHdU8IngO+5l3P1nSoJTJbWsjA/Mwanm5+0XQREQemZ6
3v8gGhwHit1zu8ALVfS6knBy5oouZONXkqVnMjsuNdnzcoYel/NiOY/jYhTnq3AUL4t3o3y1gMsIZ+Nwni7fp9lv
dDMtOWzYduUpVs3QHE8+jTMU3EydgX5wdtBPcs+Q9VdWM2oYiGDkGdl+ZJVZsLoGMBRgDbuWcmeRd3AmBA3tEFaD
UMPv2en0wnr+l1OFFBSDs+TTvJdfOYt/Vl7JKrKvzRd5+MD4ZmugjgQat86ZlA9LpilYFoT9KLGsVNZAAb+e4Pbd
A8uRo4sHXpotZLGP4zCF0x7dayPFz+5+hz7hYFEOB7HH4SE4GInD2dF0vJGfZOEYXxGMOyDEHhj6GQ6Lcfb/SuGp
U4TYAQs/T5I4za8A0w4I8UUtZh0O4qXSASOF/btCIV704ijJ8muzKTokxItilCc4ubZGDP+8p/1D8rKF4LN1/vLA
oIJxbwObdLJRMsgHuDeCTc5z+pdXA2f1PwAAAP//AAAA//+smutu4zYQhV8l8ANsItm5LZIAtUXS4kXvELhBsyi6
Kdbptn37ktZIIuewMhPo3+Lbo+F4DklxIj4cX19e3pvn9+enhx9vf1/8eFxVq4vjn8/fj/5fX6vN6uKfavN8+Prr
v83L8fDy/f1xdfWlvl49PRyC+JegflzV96sL/z9Hj38+XT1c/nx6uDyQZDtKLonsgDRABBAJRAHZA2mBaCAGiAXi
YnLpSzXWq/5QvYL6VK+hGFsgOyANEAFEAlFA9kBaIBqIAWKBuJgk5Vmz8rz6eVLffbnOTqzNNLHCc15ZrcZCAdkB
aYAIIvEErTbpDJWjZhhLAdkDaYFoGN0AsUTuxt/lYk1SOr/4kpV4Kl315dbX7v312+H37ZuvpZ+XmRW6GRdoCOLr
eBst0Iot0F5S3Ywp7YDITJg6DaPgoT0QmwmzTsM4eug2ZJOU4yZTjuq6qBzrEIz2N18JqGoVNrrDX8f3tz/2L99+
C4X2dT1X5PVY5BDTF9lnOO6C16zIvSSMM0qqm1Szy2lYnIY0ftKMcdh0FhkJG0mSxK/aMcots5MkoQ7Dzn6XSvYZ
yX0qaTOSir0gdO5nszlqqMJx+dZMY3MaNkkdjRWW0PCrKqbpKE6imWZpMifvMpNp46sKc2n+LRrCPK5OU/70Wt32
YNokdhw0Pbgf16zgCsmB6kEdm14z1/cFmrZAo/nghgPLgeOg60F1BTuB38xhCX+i6iGMX7XTG6YH0ytnx0HDgeBA
cqBolNmqF2jaAo3mgxsOLAeOg64Hpx0zmereh0WqfoqTlJ1IVHcgDRABRAJRw1izxS8RtSUiDQkYIBaIA9IRyXgQ
TqHx4Ty8qT4x8ys6zU5Tn0jsQa+ZSAMaAUQCUcNY8x5QQnOitiSShgQMEAvEAemIZDwIR90lPKAjc+RBT2IPOGkq
TgQQCUQRmd/6S0RtiUhDAgaIBeKAdEQyHuS6is+sAzp7Rx70JPaAk6biRACRQBSRMx5QQvProECkIQEDxAJxQDoi
GQ9y7clnPKCeIPKAWoCpHwnH5vAHiZE0QAQQCUQROeMBJTTvQYFIQwIGiAXigHREMh74M+oie1GIk76TexKvA06a
ihMBRAJRRM54QAnNe1Ag0pCAAWKBOCAdkYwHub70M+sgxEk96EnsASeN79rTlSGASCCKyBkPKKF5DwpEGhIwQCwQ
B6QjkvEg19R/xgNq/aK9iBrGaC/ipKk4EUAkEEXkjAeU0LwHBSINCRggFogD0hHJeLBQL1xRexl5QI1g5AEnDT01
rRUBRAJRw1jzZ9OSnrgkkoYEDBALxAHpiGQ8WKgzrqA1JhLvRdAcg0YAkUDUMNa8ByUdckkkDQkYIBaIA9IRQQ/q
hfrkU5zkfUAk8gBIA0QAkUDUMNasByWitkSkIQEDxAJxQDoiGQ8W6pP95wf2TiYSewB9MmgEEAlEDWPNe1DSJ5dE
0pCAAWKBOCAdkYwHC/XJ4QWZnouIxB5AnwwaAUQCUcNY8x5QQrPv5JJIGhIwQCwQB6QjkvFgoT65hj6ZSOwB9Mmg
EUAkEDWMNe9BSZ9cEklDAgaIBeKAdEQyHizUJ9fQJxOJPYA+GTQCiASihrHmPSjpk0siaUjAALFAHJCOSMaDhfpk
f/GB70XQJ5Mm+lsFEAFEAlHDWPMelPTJJZE0JGCAWCAOSEck48FCfXL4ysreB9Ankyb2APpk0Eggahhr3oOSPrkk
koYEDBALxAHpiGQ8KO6Ty799h6sF3o/N1K8RuZ76NSANEAFEAlFEwl47fr6FT5glorZEpCEBA8QCcUA6IqcLFund
oVzPXN2H2zEf/IJcU0c8fTLeDsi3TmOx1uzb+w6faxAJRDIXvWbRFT63R9RmQ7GP+xqfM4gsIpeNzq4pdMlzqUW5
lnrtl/yHr4tEl5vCfbmwh/lWcZrH7BrClkS+2R8vziFqBjTd3hGIJCKFaI/hW0Q6mzy7xWPwOYvIYQ5dgtKbZLnG
uj7dRvzoalmHUP57QrRaEO0QNYgEIolIIdojahFpRAaRReQQdQlKa8sb5vn7Kuv4VlA93T9KY/IG8EzM5CU7XTVK
YyYNzbnLqev4PVFNV13SkMn5/GzI5ALXtI2kIZPj5tmQyUWj//vhyenpbMj4Jlo9XSFLs0wOA2dDevW0V033yfqQ
l9M94v8AAAD//wAAAP//ZI9BDoJADEWvMukBREFdEIbEGBcsWHmCEcowEaaklHB9Z0iUhcv3mv7fFiOyxTsOw6wa
WrxoSKEsflYxdhpup7xOIfnzjyyvsuiTPaYsJmOxNmydn9WAXYg8Hs4XUOxsv0EKSmiKOg36RSI0BgLVo2mR4yAD
1RHJF0JFTH2iLJOaGzOghmtYIHboxYgjr2EiFjZOQlPuWg1ctad4XMtmdd7udnslWYnfc48o5QcAAP//AwBQSwME
FAAGAAgAAAAhAPZgtEG4BwAAESIAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7FrNjxu3Fb8HyP9AzF3WzOh7YTnQ
pzf27nrhlV3kSEmUhl7OcEBSuysUAQrn1EuBAmnRS4HeeiiKBmiABrnkjzFgI03/iDxyRprhioq9/kCSYncvM9Tv
Pf7mvcfHN49z95OrmKELIiTlSdcL7vgeIsmMz2my7HpPJuNK20NS4WSOGU9I11sT6X1y7+OP7uIDFZGYIJBP5AHu
epFS6UG1KmcwjOUdnpIEfltwEWMFt2JZnQt8CXpjVg19v1mNMU08lOAY1D5aLOiMoIlW6d3bKB8xuE2U1AMzJs60
amJJGOz8PNAIuZYDJtAFZl0P5pnzywm5Uh5iWCr4oev55s+r3rtbxQe5EFN7ZEtyY/OXy+UC8/PQzCmW0+2k/ihs
14OtfgNgahc3auv/rT4DwLMZPGnGpawzaDT9dphjS6Ds0qG70wpqNr6kv7bDOeg0+2Hd0m9Amf767jOOO6Nhw8Ib
UIZv7OB7ftjv1Cy8AWX45g6+Puq1wpGFN6CI0eR8F91stdvNHL2FLDg7dMI7zabfGubwAgXRsI0uPcWCJ2pfrMX4
GRdjAGggw4omSK1TssAziOJeqrhEQypThtceSnHCJQz7YRBA6NX9cPtvLI4PCC5Ja17ARO4MaT5IzgRNVdd7AFq9
EuTlN9+8eP71i+f/efHFFy+e/wsd0WWkMlWW3CFOlmW5H/7+x//99Xfov//+2w9f/smNl2X8q3/+/tW33/2Uelhq
hSle/vmrV19/9fIvf/j+H186tPcEnpbhExoTiU7IJXrMY3hAYwqbP5mKm0lMIkwtCRyBbofqkYos4MkaMxeuT2wT
PhWQZVzA+6tnFtezSKwUdcz8MIot4DHnrM+F0wAP9VwlC09WydI9uViVcY8xvnDNPcCJ5eDRKoX0Sl0qBxGxaJ4y
nCi8JAlRSP/GzwlxPN1nlFp2PaYzwSVfKPQZRX1MnSaZ0KkVSIXQIY3BL2sXQXC1ZZvjp6jPmeuph+TCRsKywMxB
fkKYZcb7eKVw7FI5wTErG/wIq8hF8mwtZmXcSCrw9JIwjkZzIqVL5pGA5y05/SGGxOZ0+zFbxzZSKHru0nmEOS8j
h/x8EOE4dXKmSVTGfirPIUQxOuXKBT/m9grR9+AHnOx191NKLHe/PhE8gQRXplQEiP5lJRy+vE+4vR7XbIGJK8v0
RGxl156gzujor5ZWaB8RwvAlnhOCnnzqYNDnqWXzgvSDCLLKIXEF1gNsx6q+T4iEMknXNbsp8ohKK2TPyJLv4XO8
vpZ41jiJsdin+QS8boXuVMBidFB4xGbnZeAJhfIP4sVplEcSdJSCe7RP62mErb1L30t3vK6F5b83WWOwLp/ddF2C
DLmxDCT2N7bNBDNrgiJgJpiiI1e6BRHL/YWI3leN2Mopt7AXbeEGKIyseiemyeuKnxMsBL/8eWqfD1b1uBW/S72z
L68cXqty9uF+hbXNEK+SUwLbyW7iui1tbksb7/++tNm3lm8LmtuC5ragcb2CfZCCpqhhoLwpWj2m8RPv7fssKGNn
as3IkTStHwmvNfMxDJqelGlMbvuAaQSX+nlgAgu3FNjIIMHVb6iKziKcQn8oMF3MpcxVLyVKuYS2kRk2/VRyTbdp
Pq3iYz7P2p2mv+RnJpRYFeN+AxpP2Ti0qlSGbrbyQc1vQ92wXZpW64aAlr0JidJkNomag0RrM/gaErpz9n5YdBws
2lr9xlU7pgBqW6/AezeCt/Wu16hnjKAjBzX6XPspc/XGu9o579XT+4zJyhEArcVdT3c0172Pp58uC7U38LRFwjgl
CyubhPGVKfBkBG/DeXSW++4/FXA39XWncKlFT5tisxoKGq32h/C1TiLXcgNLypmCJegS1ngIi85DM5x2vQX0jeEy
TiF4pH73wmwJhy8zJbIV/zapJRVSDbGMMoubrJP5J6aKCMRo3PX082/DgSUmiWTkOrB0f6nkQr3gfmnkwOu2l8li
QWaq7PfSiLZ0dgspPksWzl+N+NuDtSRfgbvPovklmrKVeIwhxBqtQHt3TiUcHwSZq+cUzsO2mayIv2s7U579rUOu
Ih9jlkY431LK2TyDmw1lS8fcbW1QusufGQy6a8LpUu+w77ztvn6v1pYr9sdOsWlaaUVvm+5s+uF2+RKrYhe1WGW5
+3rO7WySHQSqc5t4972/RK2YzKKmGe/mYZ2081Gb2nusCEq7T3OP3babhNMSb7v1g9z1qNU7xKawNIFvDs7LZ9t8
+gySxxBOEVcsO+1mCdyZ0jI9Fca3Uz5f55dMZokm87kuSrNU/pgsEJ1fdb3QVTnmh8d5NcASQJuaF1bYVtBZ7dmC
erPLRbMFuxXOythr9aotvJXYHLNuhU1r0UVbXW1O1HWtbmbWDsue2qRhYym42rUitMkFhtI5O8zNci/kmSuVV9pw
hVaCdr3f+o1efRA2BhW/3RhV6rW6X2k3erVKr9GoBaNG4A/74edAT0Vx0Mi+fBjDaRBb598/mPGdbyDizYHXnRmP
q9x841A13jffQATh/m8gwJFAKxwF9bAXDiqDYdCs1MNhs9Ju1XqVQdgchj3YtJvj3uceujDgoD8cjseNsNIcAK7u
9xqVXr82qDTbo344Dkb1oQ/gfPu5grcYnXNzW8Cl4XXvRwAAAP//AwBQSwMEFAAGAAgAAAAhALk2x3+YBAAA5SEA
AA0AAAB4bC9zdHlsZXMueG1s7Frdb+I4EH8/6f6HKO9pPiAsIMKqbIu00t7qdO1J92oSB6x17MgxbdjT/e83dhJI
l1IgDZSu9gUSx5n5jefTnow+5gk1HrDICGeB6V45poFZyCPC5oH59/3U6ptGJhGLEOUMB+YKZ+bH8e+/jTK5ovhu
gbE0gATLAnMhZTq07Sxc4ARlVzzFDJ7EXCRIwq2Y21kqMIoy9VJCbc9xenaCCDMLCsMkPIRIgsS3ZWqFPEmRJDNC
iVxpWqaRhMPPc8YFmlGAmrtdFBq52xOekYuKiR7d4pOQUPCMx/IK6No8jkmIt+EO7IGNwg0loNyMkuvbjvdE9lw0
pNS1BX4gSn3meBRzJjMj5EsmAxM0p4UdfmP8kU3VI9BwOWs8yr4bD4jCiGva41HIKReGBNXByukRhhJczLhOJc+M
r0gI/qjmxighdFU889SAVnk5OSGgADVoKzAFpPFopmadk+GamQZ4YunWzLpnWMo1s945mfXPyMxzzsnstAaivSAD
NyCUrj3TU04IA+MRhDCJBZvCjVFe369ScEEG0bbwIj1vz+y5QCvX82sv2JohOB4XEUT3Kia4XWBdjI1HFMcS/FKQ
+UL9S57C74xLCSFwPIoImnOGqHLl6o36m5AWIAMEplxABK8CCGERznEUmD3tCrZiUXI4aL7GoqEcNB0gV4gPml8I
d7hsCY7IMvlZpTtWc3tW45fuaob/jGXuWb5jbfPC/O6NpTu5bbYs38kiyzqEnyQgnsjoLgv0wbZ0WtgnTz57GLz7
EPbu5Ds6iL2ZhDvKxirmvCJYlrUmlMwhpvRO1Zj/xJv6FWqxPDbYMpkm8jPUmXBeofaZ1SUUy+VlUaoWN6qErVMr
aNfIdpxGdI08XjM4FpULu/PybQOlKV1NdKVebr6Ppea1Sq3TKrVuq9T8Vqn1WqX2oVVq6vymPQt5gVphf9eUzFmC
1cmROiRC1a2x4IJ8Bx9Th0UhPMdwliZxLv/iEg7g1LHhALzwUaD0HkZLA7bz+HhPHTSSGOxLbXa3/B4Otn5YQHUG
pvD96HBnXwBwsAuArMLiS2p6k4WFs2hJwvZsDeLF+Zd6W4hj3AOC0nuDrNLwsz7YLJXs9I8tcrtcGmz7+DqhSbyF
zPus5LAiT6N3A6S7bLcN2pCrmuA+Jkvsiy+7xNtKVS3G7S3PbApyqw65RJAnTYFni9Rbxe0pl/qVpc3OdHOwzzYo
Rl6Xbnal+T2ID4oE+kxYtyVeKjQuIBadEYLeisLms7bDfbK/Xe9UDdWCDcyvqoNOa+lktiRUEvbM3hZoRvlmt6z7
aFJ1w/U+es0FNBvhGC2pvF8/DMzN9R+67QFuV876kzxwqUkE5ub6i+obuboLCXX/lwwaPfBvLAUJzH9vJx8GN7dT
z+o7k77V7WDfGviTG8vvfprc3EwHjud8+q/Wk39FR15/QgDp3u0OMwp9e1EKW4K/24wFZu2mgK9tE2DXsQ+8nnPt
u4417Tiu1e2hvtXvdXxr6rveTa87ufWnfg2737Bz79iuW3wDoMD7Q0kSTAmrdFVpqD4KSoLbF4SwK03Ym+8zxv8D
AAD//wMAUEsDBBQABgAIAAAAIQC6TKIfUwMAAJAHAAAUAAAAeGwvc2hhcmVkU3RyaW5ncy54bWy8VU1v20YQvQfo
fxjw0CaoI8p2mraOKMORokKIYQu2jCDHFTmUFtoPZncp2/31eUtKtSw6QE4VBBLcnZmdefPm7eD8QSvasPPSmiw5
7vUTYpPbQpplltzNJ2//SsgHYQqhrOEseWSfnA9/eTXwPhB8jc+SVQjVWZr6fMVa+J6t2GCntE6LgE+3TH3lWBR+
xRy0Sk/6/fepFtIklNvahCz5+ySh2shvNY/ahdPTZDjwcjgIwxv2SO83TxNpmCbWFv4tzXHIWtBY+uDkog7InkZs
ArtBGoaDNLq27mMR+HDtK4uO3e1KluHs0HJqFkiooGxK7e+6DtuV60PbC+fkRiiaS905caaECTGCdfH5RThe2dp3
7OZOSMWO5qwrdiLUjun15M3hUTuzqfEV5031RDPhPWUzogliEGWTQ6+IHY25ZOOZRivO19uidq/nETLqRJg5W9R5
oCmQXjoZHuHZOtHu3Be8pkYGKZQ/zGdktUbPOutzGwAjEqwQXCl+wcIJ48uI0xavS7v8EUZXte51SGHRh5c2Tnv/
hdQ1GL5gyhULc0SlYyZbUsELJ/0RGRuotLUiW1jnY18rhoNANzYApnPkHz2g1nQLbo4KocWSC6paRI8Ip6wRNgbB
i0Peo1tWaG4LMFxiWzth3yGsp7BiClsoPAvFxTn9lPt8/DsYASBHKG9NGf2Dt+lA+TmDgQW9bizGGGZfVrI7Vrc1
OLuRHrneyqVp2NuZqBveSL6nOJadvWOgj+EhaZqKCnTJE0aFoqQQRIjC3lxI1C3WbDqYvEcY8OaxhaUhk6n1AlWi
f1XLKIJg7KHWifFnjyAykYdPrvuM7DjMKD2cFwf5ihp55iuRQzshgpCyDSfDkx7dYSiejbnnSLdwD/ypT20LECH+
Z3gsUijavwSBgUwfJ/jKrQLSKEIjdrPiJhYq05hcVMF6uhLO2ftoXAotgUizdxIXGrXmdkFLY11cTJujwnBCvwpd
faB3/V1Ne5n8b1l8cJxbV9DCPjxrfOmsjtdFwPTICIElBV7iytp2Bcnu3QDzjxl9BFGg4vSVlbL3B/x+uUeX0a2G
piN8tu3H07XScmMrT61+3l1dXl+MP40709PQaGfacBvU3SrXk3GKG3X4HQAA//8DAFBLAwQUAAYACAAAACEAd0Sh
dLICAAB7BQAAGAAAAHhsL2RyYXdpbmdzL2RyYXdpbmcxLnhtbJxU226cMBB9r9R/sPxOsIHlpmWjvVFFitqoaj/A
MSaLChjZ3ksU5d87NuxuozZS2yfGnvGZmXNmmN+euhYdhNKN7AtMbwhGoueyavqnAn//VnopRtqwvmKt7EWBn4XG
t4uPH+anSuVHvVEIAHqdw7HAO2OG3Pc134mO6Rs5iB68tVQdM3BUT36l2BGgu9YPCIl9PSjBKr0TwmxGD57w2H+g
dazp8cJVZo5yLdp22fOdVEhUjVnqAkMH9naKqZXsxmgu2wUN5r7tydoOAowvdb0IgnSWZRefvXJuJY+LkIz31j5f
2gAahGkQXXzujQO/pjTyl9Thn1PHJIvJlGKq5po6fid1FKZkagXKuqY+JxwaPmbuDw8Nf1BTGZ8PDwo1VYFDjHrW
gdLgNXslUAB0sVyczL02k4X2qinwS1kGq9m2jLwSLC8iq8hbbaPMK6H9bZCU6yCMX+1rGucchDYwY3fVWWAa/yZx
13AltazNDZedL+u64eI8MjAwNPKdxK7Ol3g7S2ZxmXhpkqRemKaht1rRlbdJs4xuY5rQbPaK/cXcd9Wfv66LUWnb
87X9kQyWA0H3kv/QqJfrHeufxFIPghtYDgfmZgRejuEO6A2Tj20zlE0LQ8Rya0/t/tV2jB1vJN93ojfjiijROuL0
rhk0RioX3aMAndRdRd9VJkiXhGTBylvPyBqUSbbeMosSLyHbJCJRStd0PSoT5XstoF/WbobmIk30z9KQSZoDawtM
3qN9pMRSo40Shu+sWQNbX4HhUaqLw1F7ZdPyrgcrGMtPtYLNZTnwhU4FhgWlsyjD6BlEimJKQzgAmtMdcYgICMQk
MUbchtA4yJJ0qvGMNShtPgnZIWsAvVCQo5cdYGLG0s4hMErXapx5WSreNqDchhlmn9ioN3+i6c7+Nxc/AQAA//8D
AFBLAwQKAAAAAAAAACEA7wWO7dYUAADWFAAAEwAAAHhsL21lZGlhL2ltYWdlMS5qcGf/2P/gABBKRklGAAEBAQB4
AHgAAP/bAEMAAgEBAgEBAgICAgICAgIDBQMDAwMDBgQEAwUHBgcHBwYHBwgJCwkICAoIBwcKDQoKCwwMDAwHCQ4P
DQwOCwwMDP/bAEMBAgICAwMDBgMDBgwIBwgMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwM
DAwMDAwMDAwMDP/AABEIAD8AcQMBIgACEQEDEQH/xAAfAAABBQEBAQEBAQAAAAAAAAAAAQIDBAUGBwgJCgv/xAC1
EAACAQMDAgQDBQUEBAAAAX0BAgMABBEFEiExQQYTUWEHInEUMoGRoQgjQrHBFVLR8CQzYnKCCQoWFxgZGiUmJygp
KjQ1Njc4OTpDREVGR0hJSlNUVVZXWFlaY2RlZmdoaWpzdHV2d3h5eoOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmq
srO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4eLj5OXm5+jp6vHy8/T19vf4+fr/xAAfAQADAQEBAQEBAQEBAAAA
AAAAAQIDBAUGBwgJCgv/xAC1EQACAQIEBAMEBwUEBAABAncAAQIDEQQFITEGEkFRB2FxEyIygQgUQpGhscEJIzNS
8BVictEKFiQ04SXxFxgZGiYnKCkqNTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqCg4SFhoeI
iYqSk5SVlpeYmZqio6Slpqeoqaqys7S1tre4ubrCw8TFxsfIycrS09TV1tfY2dri4+Tl5ufo6ery8/T19vf4+fr/
2gAMAwEAAhEDEQA/AP38or85x/wc6/AK81rV7PTfBnx51z+xNRn0u6uNL8EPeW4nhco6h0lI6jPrgjjmpP8AiJh+
Cn/RM/2k/wDw3U//AMcoA/RSkdwgyf5V+dn/ABEw/BT/AKJn+0n/AOG6n/8AjleRftV/8FPPGP8AwU+8NXXgv4M3
Xjr9nb4Q6DYPrXxd+KfjLR5dCutE0lXC/Y9NyWJuZl805XB+QAEZOQD6a/aU/wCCzOjeGvjXqnwe+BHgTxF+0V8Z
tLQ/b9I8OSRwaP4dfkL/AGlqUh8m3wwAKjcwOQdrcV+dn7Xn/BVjxX4e129sPjv+25ofw5vUmaOXwJ+zz4fGs3+n
li+6K51ef5UliXywdkn3t/yggGvzZ/bw/wCCukeueApPgZ+zNYX3wj/Z10kmH7PaOYNc8cSDAa/1a5U75Gkxnyc7
QMBs4UL6n/wb/fsN/swf8FRtK8Z/Bf4laf4j8N/GgWs2r+F/E+mavIqXdsqhZIvsr5haSBisu3GZUL/MvlkkA77x
B/wV3/Z31W8Frb+Ov+CkvxGWwjeWbUNQ+KUdg4iUkl/Jt2KhAOSxwRnk12Xwt/4Ky/Cy91+2l+Gf7cn7YHwk1jYV
tbb4uWNv4y8Pq652RSmIyMkbEgFyjMACSelfS/8Awb4f8EK9E/Z4+If7VGgfGOz07xF4kspX+GvkqSILjQr20juZ
LuNSMgXccsSq3VDBKo5DV8nfsR/8EAfhH8NP2ePjf+0F+07r+pX3wi+Hupapp3heDw9qAgPiaKyu5Lb7as23LpcS
oIYFUrvLMxIG3IB+ifwU/wCCv/x7+BHgF/GfxN8O/D79pX4KxsTcfEn4G3f2250RfkOb/S5CHUKCxcpt2Bec1+if
7M37U3w//bE+E9h43+GnizR/GPhrUOEvNPm3iJwATFKn3opVyN0bgMM8iv4rfg/+294u/ZJ/aUuPiJ8DdS1v4WSL
dGSysLXVHvUjtw2VtblpAFu4+BuEqbW/ujiv1b/YP/bnHx01nU/jv+y3oVh8Pv2jfCdsdT+KXwYs5WTw58XdKQlr
jUNNhX/V30ZZn8tQWUv8u/nzQD+jiivzl07/AIOavgxdafBLN8Lf2lLaaSNXeE/D2ZjExGSuRJg4PGRxxU3/ABEx
fBT/AKJn+0n/AOG6n/8AjlAH6KUV+df/ABExfBQf80z/AGk//DdT/wDxyvo3/gnt/wAFM/h5/wAFLfC/ivVPh/Y+
L9OTwXqi6Pqtt4i0r+zrqC4MYk2+XvY8A85wQe1AH0PRRRQB/NPfftR/ET9kb/ggZ8TvFXwy8Y674I8QyftS6pYP
f6VceTM9u9kWeIn+6WRCR6oK2v8Agl78Kf8AgoN/wVd/Zyn+I/gH9s+fSbSw1afRb/TdY1e/S7sriNY5MN5du6FW
jljcEN/FjqDXlH7TP/Kub8S/+zsNR/8ASGSv0F/4Mof+UcPxK/7KRcf+mzT6APnr4G/s3/tx/tU/8JHa/Cn/AIKO
eAPiJr/hmCSW70bSfF199tRlLKqsj2ylAzrs3sAoJ5NeLfs7/sIftyf8FtvCvxR+G3jv4/XVhL8JvEdvp3iTwj46
vb3MV3tkMMoEULo65SXHzfwhhkFTXlH/AAQB+D/xG+JX/BdfQdW8BQaoum+E/Ed9qXii/tpDHbW2l+ZKksczdCJS
yosZyWYjA+UkfvT/AME3viz4W+Mn/BVj9uPUvCUtvdWem3/hDRL+7t3DR3WoWmn3cNwQR/cZRCfeE0AfgV4T/wCD
fPT/ABz8X4PAOlftgfsr33jO61BtJh0ePXb37VLdqxQwBPswzJuBXb1J4HNaH7G//BJDxP8As/JH+0R4l/aP8Ffs
9eDPBfji68PeF/G8lpc39z4hvLSaWF7jT7QKpmgYxTriTG5Uk3Jt3Vuft6/8EFPiR8EdZ+Nnx8v/AIufBiPw14R1
u/8AEMyaF4imvdZtJZb4/ZbfyVhUR3DTSxJ8zjax6nFfb/8AwTht/gf/AMHAv/BGbw1+zP4s8TP4T+LHwtlNxA/m
xyagLlXmKapBEzD7TBIlw6TJkEMW5X925AOS/bt/aL/aF+Onxi/ZqHgH47/A7VrD456jP4J0/wCIvgODV9EufEot
3iIsNbginkaCPzLsAJCfMiaaRg0Qcinf8FwPgt8Yvib4P8DfBf41/tE/shfALwfo9nFqWkeDND/tq0tr2OMvBDM+
+2kJSPbIqICqqQTgkAj5i8C/8EsvjD/wSU/4LKfsu/D/AMbaj/b/AIB1n4m6Xq/h3WdOSUaXqE/2mGKU7XH7q6Ea
xCSPJ+Xy8Mwwa+9/+Dlf/gjz44/4KVftkfDXVPB/xA+DvhiWz8LDRo9M8V+JG07Ub+Y3lxIGt4VhkMiYkAyOdwIx
xQB+aXhj/g248S+PP2g/h78OfDvx9+BviPU/ih4XvfFnh6/0q8vruwvbS1lSNx5qW5+Zi0pXjH+jTAkMuD6J8Ef+
Deb4m/Ab9uKDwX4G/a5+BPhP4++FSLi20uz1jUrTWbYyW4l/dA2oL7reTcQu75GORjNXv+DeD4Xj4Bf8HDWl/DmH
xpaePLTwDpOu6PHqliZRp7Srau1wlssnIiW4aYZwN5DPgbq9S/bl/ZB+Kn7Sf/B2t9u8AaB4g+x+HPEfhTWdR1+G
0mWx0q1trDT5ZpJJwu0ZVHQDPzMdvrQBu/sxH9sfVIPit8I/iV8UvjuP2m01+w0DwFpsHi9rPSoreWKeW6124ZUY
T6dbxojmRc7maOIDe+K7HTv2af2h/E3jh/hpof8AwVd8P6n8cYbhrWXwiNQdYxOg+eBZhK0jSL3UQ7uCCowcfdnw
o/4KAfBT43/8F3Nf8BeGdT0fVvG/hv4Yvo02r2829Jp49S+0XOlxt913hTZK2wnB8wHmNgPwD/4Lhf8ABG34vf8A
BND9o/XviE8V1rfw48T+IbjVdE8W6UjqunyTTtNFBdYGba4Un5Tna+3KNkFVAPWPij/wUI+Pf7Ongfxx4V8c/tUf
FNPjV4C8TaloOpWNp4qQWYFv8kbxRsgebdIBt25yHk37Ngr9hv8AghnI037TP7cruSzv8ZZmYk9SbWMk1/KN8Zfj
N4h/aG+M2t+N/Fl2NQ8TeJ743+p3QQIbq4fG+QgcBmOWOOMk1/Vx/wAEMP8Ak5X9uP8A7LJN/wCkkdAH6K0UUUAf
y1ftM/8AKub8S/8As7DUf/SGSv0F/wCDKP8A5Rv/ABL/AOyj3H/ps0+vz6/aZ/5VzfiX/wBnYaj/AOkMlfoL/wAG
UJ/41w/Evn/mpFx/6bNPoA9c/wCCkOh6n+0B/wAEVNS+IX7F+sSfDCSJZPFNxa+DrePSbjWrWATx6hZP9nQMLhG8
xjtIYvblcndXyv8A8GP1zJefDr9o6aaR5ZpdW0N3kdizOxivySSeSSe9fHf/AATb/wCC2PxH/wCCHP7V3xD+GXxM
8Ja7qfw41PxRc3Oq+HLoG21Pw9cPMQ13aK/ytvj2lomISTahDqclvoif/gpd8Mf+COeg/Ev4mfsbt4G+L/hP45an
b69qun6tr/2S7+HUwRzHZvpgC3MkTSXMuJA4CbfKP3A7gH5yeLP2H/i9+3f/AMFQfjf4L+EfhfWfEF1f/EDW4b+S
ANFp1pD/AGpK2+7mOI0jVgrfMc5UbQWxXb/8Fpf+CZviD/gi9+0R8KIvCdxr1hE3haxvofGVjcyw/bteikkN7JBI
pBgZGMexAQRH5Z5JJr6/+IP/AAU6/wCCgXwx0C6vdA8Q/ATwRp2qaPovjG50vwzoFrGLGHXLu2t7SRxNCxaWSW6j
aQ75MYkJbKkV5X8DfiZ+3p8IfD/xB+EfinXPBGueHvh5balr+seH/irpkGvWthHYXFvFI8Anglco738DRbD5ZR9y
7VBNAH3x+zH+2Lr37dv/AAS9/Yt8S/FWwvdV+J0nx70DT9Hv2hC3WtDT7yRptTVByyCxWdZnAxuR2OAa+df+DxD9
nb4j/Fv9u74Tar4H8DeN/Esdp4KWA3mhaLdXqwzrqF06pvhRtrgEMBnIyD3r4s8JeP8A9sb/AIK3ajffFHQfiK9j
qHwxNr4V0LTdDvn8PNaSX8Unl2OmW9mixxmdbQqzM0attRWfG0DU8L/Er9uXT/At1rPhv9qbxRq3gTRG1ZdX8R2P
xGvZtL0V9OEPnrPI/wC8Jf7RbCHy1cTm4jEZbJwAfTn/AAbrf8E4vib+xZ/wVA+FXiv4waRf+D/EXj6w8RR6L4e1
Af8AE0ktoLENcX91HktDGZJoo0D4Z3MhwAg3fqX+1T+0Iv7e0n7X/wCx9ouvS+F/i54W0K3uvDL2121pNqlrc6da
3cTxujBzsumMM2ONk0echiK/GT4S+Nf29vBPxu0H4uW3xr8Gav4j1yx0rwb4T8ReIry21CLxXbagl1e2ljYy3Fs2
POmtLmJi/lMZ4hG7ZArxj/gor+0b+07+zh+3h4C+O/j2TQPA/wAdNQ0ex1iLVPDdtJB9tg+zRrGbpGBtpXMLeRKs
RaM+W8bjKnIBy3wg/wCCevxU/ZT/AGJp/wBsTVPEmrfCo+GfEcOk+A4reArq+saus7xyPhmUQW8Xk3AZmDmQwyJs
2kk/uL/wR+/4OEPhl/wVy8LxfA/4zaBpOj/EzXdMNjdadfxx3GheNgIz5wgVhhJGVS5t3HrsZsYHn3xB+LFl/wAH
N/8AwQk13QvBP9g6X8dPBd1b65qfhOBvJVNTt2lJ8pWORFeRSTGNySokkKM2VZq/JP8A4Jr/APBH39qHxp+354Bs
bT4YfEXwHJ4U8VWN5qniHVNGudPtdBS3uFkebznUKzARttVCd5wBwc0AR/8ABfz/AIJfWX/BLb9vufw34aFy/wAP
vGNsviHww0xLGzgeV0ksi5JLmCRSAx5KNETkkmv39/4IYf8AJyv7cf8A2WSb/wBJI6/HD/g7P/bo8Oftb/8ABRLR
fC3g/U7DWtB+EelnRZ76zlWWKbUpZjJdorrwRHthjPJw8clfsf8A8EMP+Tlf24/+yyTf+kkdAH6K0UUUAfzL+NPg
N42/aK/4N+vidoPgHwl4i8Z62n7VGp3bWGi2El7cLCtkytIUjBbaC6gnGAWHrXlPwk8fft7fst/sRW/wM+DvwS+M
Xw50XULq41TxHrmn+F70a3rd5OUVilwIgbaJIooo1WLDnaxL/PtH7XeCv+DdPwr8KLrXf+EH/aS/ay+H9h4i1e51
y803wx45h02xa6nbdJIIktcZICrkknCLknFbv/Di67/6PN/bm/8ADnj/AORqAP5f9V/4JbftS67qU95e/An4zXd3
dOZZp5/DF9JJM55LMxQkknqTVf8A4dQ/tOH/AJoD8X//AAlL3/43X9RH/Di67/6PN/bm/wDDnj/5Go/4cXXf/R5v
7c3/AIc8f/I1AH878nwj/bwnvdYuJPhF8XJZNe07QtJvd/geciS20VrZtOQDyfl8prSAkrgvsO7dubO4NG/4KBX2
lPp+sfCX4q+JtNl0W58OPaa14Clvon02e/i1B7U74CfLF1DG6DPyBdq4T5a/oE/4cXXf/R5v7c3/AIc8f/I1H/Di
67/6PN/bm/8ADnj/AORqAP5vvhj+yn+238F/DTaV4T+EPxo8P2x8S6f4uRrLwldRTQ6nYpcJazpIIty+WLmbCg7T
uGQcDHdxeE/27LGIWNj8BvG+l+GJY76O98M2Pw1aDRNV+2+R9ra5tVg2yvKbW2+ZuU8iPYU2iv6Bv+HF13/0eb+3
N/4c8f8AyNR/w4uu/wDo839ub/w54/8AkagD8A7K2/4KCaXDFFa/B/4kWsGn3lhqGiwxfDjEXhqewinis5NPT7Pt
tTCl1cbDGBhpWfl/mrzX4v8A7Hn7Yvxs8F+G/Dmr/AX4oQaD4UkubjT7Cy8E3VvDHc3IiF1ckCPJlm8iEuc4JQHA
JOf6Rv8Ahxdd/wDR5v7c3/hzx/8AI1H/AA4uu/8Ao839ub/w54/+RqAP5nfhJ+wP+2T8A/HNp4n8EfCf4/eEfEVh
n7Pqej6HqNldwg9QskahgD3GcHvXu/xO+IP/AAVH+MngqXw74jj/AGqL7RbmIwXFsmlX9sLqMjaUlaKNWkUjqGJB
5zX71/8ADi67/wCjzf25v/Dnj/5Go/4cXXf/AEeb+3N/4c8f/I1AH8vcP/BKX9ppLhG/4UB8XwAwOB4Uvf8A43X9
Mf8AwQzQx/tM/tyqwIZfjNOrA9QRaxgiuh/4cXXf/R5v7c3/AIc8f/I1ezf8E8v+Canhj/gnJpfjmLw94w+IPje+
+ImtDX9a1XxjqUWoahdXfliMuZUijLbgMksCSe9AH0bRRRQB/9lQSwMEFAAGAAgAAAAhADkxtZHbAAAA0AEAACMA
AAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc6yRzWrDMAyA74O+g9G9dtLDGKNOL2PQ69o9gGcr
iVkiG0tb17efdygspbDLbvpBnz6h7e5rntQnFo6JLLS6AYXkU4g0WHg9Pq8fQLE4Cm5KhBbOyLDrVnfbF5yc1CEe
Y2ZVKcQWRpH8aAz7EWfHOmWk2ulTmZ3UtAwmO//uBjSbprk35TcDugVT7YOFsg8bUMdzrpv/Zqe+jx6fkv+YkeTG
ChOKO9XLKtKVAcWC1pcaX4JWV2Uwt23a/7TJJZJgOaBIleKF1VXPXOWtfov0I2kWf+i+AQAA//8DAFBLAwQUAAYA
CAAAACEAspc8br4AAAAkAQAAIwAAAHhsL2RyYXdpbmdzL19yZWxzL2RyYXdpbmcxLnhtbC5yZWxzhI/LigIxEEX3
gv8Qam+q24UM0mk3MuB20A8okup0tPMgyQz69wbcjDAwy7qXew41HO5+ET+ci4tBQS87EBx0NC5YBZfz5+YDRKkU
DC0xsIIHFziM69XwxQvVNiqzS0U0SigK5lrTHrHomT0VGROH1kwxe6rtzBYT6RtZxm3X7TD/ZsD4xhQnoyCfTA/i
/EjN/D87TpPTfIz623OofyjQ+eZuQMqWqwIp0bNx9Mp7eU0WcBzw7bfxCQAA//8DAFBLAwQUAAYACAAAACEAXXiZ
6uECAACQLQAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxLmJpbuxazW7TQBD+1klFWyQQ
EhI/pyAuCCogEEAco4BEaUgtatTcKle2wBKJUaJQioTgEXgMxJkLJx6CKw9Ae+bALcysvWki22lQ2uLQmciOdzzr
nflm9mdW62AFj/EQDm7iFoBlNLCKOpWrKGETAd7iBXr0VEIN93AHZeKVYFOpTvKqCPUDX5ZOf4OloLC7GM579H8C
TcuCRfcClUA1yrhBLbyh/zKVD5IUN5BBn5aANbveNCK/vt//+pF4zH9mP9hYXW8Matq1epNtGkdsYzi/u3gN0U9x
gSpULgG/rwK+ZkRcw1PwCI9xWo60qIgIWIXlttNxt7PRisRGhNmF6UTeoW+OCN/OFC4khSuZwsWkcL3mZIj/LCTU
eOK2e+7LNPkiMQc67+OY6LUx8B8Lm+Y5WvJBedEjJ3BkqiE45d1DedJPouUgvcHziESfICAIHG8ELBoGeCEXrZ5o
SDg1vHZOGSJSltbjBhJ+VwAlLlqIF5k88MQ1NNPDO6BIj3xpVfrvgR1eFuubjFSzFqEKj9BPVdp4nv2/d+noi90c
P1Pmkj/i8ByKX61xFMkDdiQybB7XGTZQh/9cbJzpOYPeN7XRVXRoz8BFao430df76a6bUjN2qKUzbdqx4CSdTOeW
9rI2MwZN2FByHFK8A9K07MJ53inRFAdVlNtGrov8MWtdaub0jYNecBYEBAFBwCCgR+0NWnRWO0H6PmQsWQt7ncDv
lBr+Vip6TtDyu/y29DRsue2EzNp2azMcMwmuB+3nHl1d8Y0g8B8iMHfy3NmL6XZN2PUyQZGuJz1GEBiDAGd3O5R6
XKa8I6DCa0p8Ph9qRhvnjworCia3PO4uMrm1hKogcDQI5Cri9FmJD1Mb3sUr2kvaQkjXXVRwHT6dK/EH3+XzGEKC
wFEhYOKNptRDJ4UFff7InC3i/cr9d0dl3pHeMDsIeKRqvFk+Y25bkDATBAQBQeCvEbjgnrnClf4AAAD//wMAUEsD
BBQABgAIAAAAIQBpqmRMbgEAAKECAAARAAgBZG9jUHJvcHMvY29yZS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB8kl1LwzAYhe8F/0PJfZcmdXOGrkM3
duVgYEXxLiTvtmKTliTazV9v2q51fiD0ppzzPpxzSDI/qCJ4B2PzUs8QGUUoAC1KmevdDD1mq3CKAuu4lrwoNczQ
ESyap5cXiaiYKA1sTFmBcTnYwJO0ZaKaob1zFcPYij0obkfeob24LY3izv+aHa64eOU7wDSKJliB45I7jhtgWA1E
dEJKMSCrN1O0ACkwFKBAO4vJiOAvrwOj7J8HrXLmVLk7Vr7TKe45W4pOHNwHmw/Guq5HddzG8PkJfl7fP7RVw1w3
WwlAaSIFEwa4K0264MYHDZbwsOcfYBJ8pjU7Fty6tZ98m4O8O6a3Who/54JrbhP8W+9PNibXDmRKIzoOCQ1pnBHK
6A2L6ctw15t8nrZ+Fwpk4Auxrn6vPMWLZbZCHS+ahjTKaMSuxv7zvB/3TcEOqE7J/ydOwmgSkklGrhmN2ZicEXtA
2ob+/qjSTwAAAP//AwBQSwMEFAAGAAgAAAAhAGFJCRCJAQAAEQMAABAACAFkb2NQcm9wcy9hcHAueG1sIKIEASig
AAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnJJBb9sw
DIXvA/ofDN0bOd1QDIGsYkhX9LBhAZK2Z02mY6GyJIiskezXj7bR1Nl66o3ke3j6REndHDpf9JDRxVCJ5aIUBQQb
axf2lXjY3V1+FQWSCbXxMUAljoDiRl98UpscE2RygAVHBKxES5RWUqJtoTO4YDmw0sTcGeI272VsGmfhNtqXDgLJ
q7K8lnAgCDXUl+kUKKbEVU8fDa2jHfjwcXdMDKzVt5S8s4b4lvqnszlibKj4frDglZyLium2YF+yo6MulZy3amuN
hzUH68Z4BCXfBuoezLC0jXEZtepp1YOlmAt0f3htV6L4bRAGnEr0JjsTiLEG29SMtU9IWT/F/IwtAKGSbJiGYzn3
zmv3RS9HAxfnxiFgAmHhHHHnyAP+ajYm0zvEyznxyDDxTjjbgW86c843XplP+id7HbtkwpGFU/XDhWd8SLt4awhe
13k+VNvWZKj5BU7rPg3UPW8y+yFk3Zqwh/rV878wPP7j9MP18npRfi75XWczJd/+sv4LAAD//wMAUEsBAi0AFAAG
AAgAAAAhANNciaWKAQAArwUAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYA
CAAAACEAtVUwI/QAAABMAgAACwAAAAAAAAAAAAAAAADDAwAAX3JlbHMvLnJlbHNQSwECLQAUAAYACAAAACEAq29r
IHwDAADGCAAADwAAAAAAAAAAAAAAAADoBgAAeGwvd29ya2Jvb2sueG1sUEsBAi0AFAAGAAgAAAAhAIE+lJfzAAAA
ugIAABoAAAAAAAAAAAAAAAAAkQoAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhACZv
wqtwCQAAxDMAABgAAAAAAAAAAAAAAAAAxAwAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQItABQABgAIAAAA
IQD2YLRBuAcAABEiAAATAAAAAAAAAAAAAAAAAGoWAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAi0AFAAGAAgAAAAh
ALk2x3+YBAAA5SEAAA0AAAAAAAAAAAAAAAAAUx4AAHhsL3N0eWxlcy54bWxQSwECLQAUAAYACAAAACEAukyiH1MD
AACQBwAAFAAAAAAAAAAAAAAAAAAWIwAAeGwvc2hhcmVkU3RyaW5ncy54bWxQSwECLQAUAAYACAAAACEAd0ShdLIC
AAB7BQAAGAAAAAAAAAAAAAAAAACbJgAAeGwvZHJhd2luZ3MvZHJhd2luZzEueG1sUEsBAi0ACgAAAAAAAAAhAO8F
ju3WFAAA1hQAABMAAAAAAAAAAAAAAAAAgykAAHhsL21lZGlhL2ltYWdlMS5qcGdQSwECLQAUAAYACAAAACEAOTG1
kdsAAADQAQAAIwAAAAAAAAAAAAAAAACKPgAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwEC
LQAUAAYACAAAACEAspc8br4AAAAkAQAAIwAAAAAAAAAAAAAAAACmPwAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2lu
ZzEueG1sLnJlbHNQSwECLQAUAAYACAAAACEAXXiZ6uECAACQLQAAJwAAAAAAAAAAAAAAAAClQAAAeGwvcHJpbnRl
clNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEuYmluUEsBAi0AFAAGAAgAAAAhAGmqZExuAQAAoQIAABEAAAAAAAAA
AAAAAAAAy0MAAGRvY1Byb3BzL2NvcmUueG1sUEsBAi0AFAAGAAgAAAAhAGFJCRCJAQAAEQMAABAAAAAAAAAAAAAA
AAAAcEYAAGRvY1Byb3BzL2FwcC54bWxQSwUGAAAAAA8ADwD+AwAAL0kAAAAA
"""

# Column positions from the inbound report, 0-indexed:
# C,D,E,F,G = 2,3,4,5,6 ; LPN # = 7, which is Excel column H
TRAILER_COLS = [2, 3, 4, 5, 6]
LPN_COL = 7

HEADER_ROWS = 3       # First 3 rows are headers in the inbound report
THRESHOLD = 9         # Loads with 9 or fewer pallets get flagged for research

# Transfer log template positions
LOG_SHEET_NAME = "Sheet1"
LOG_START_ROW = 8
LOG_MAX_COL = 14
LOG_TRAILER_COL = 3   # Excel column C = Trailer Num.
LOG_PALLETS_COL = 12  # Excel column L = Total Pallets in this Trailer


def clean_part(value):
    """Turn one trailer-number piece into clean text."""
    if pd.isna(value):
        return ""

    # Avoid trailer pieces like 123.0 when Excel stores a number as float.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()

    # Extra guard for values that came through as text ending in .0.
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    return text


def build_trailer(row):
    """Concatenate inbound report columns C:G into one trailer number."""
    return "".join(clean_part(row[c]) for c in TRAILER_COLS).strip()


def last_3_numbers(trailer):
    """Return the last 3 numeric digits from the trailer number."""
    digits = re.findall(r"\d", str(trailer))
    if digits:
        return "".join(digits[-3:])

    # Fallback only if there are no numeric digits at all.
    text = str(trailer).strip()
    return text[-3:] if len(text) > 3 else text


def flag_red(row):
    return ["background-color: #ffb3b3; color: #800000; font-weight: bold"] * len(row)


def load_embedded_transfer_log():
    """Load the embedded Excel transfer log template into memory."""
    template_bytes = base64.b64decode(TRANSFER_LOG_TEMPLATE_BASE64)
    return load_workbook(BytesIO(template_bytes))


def copy_row_format(ws, source_row, target_row, max_col=LOG_MAX_COL):
    """Copy row height and cell formatting when the log needs extra rows."""
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(1, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)

        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def clear_log_body(ws):
    """Clear the data-entry area so only trailer number and pallets get filled."""
    for row in range(LOG_START_ROW, ws.max_row + 1):
        for col in range(1, LOG_MAX_COL + 1):
            ws.cell(row=row, column=col).value = None


def build_transfer_log_excel(loads_for_log):
    """
    Create a filled transfer log.

    Only non-research loads are written:
    - Trailer Num. column C gets the last 3 trailer digits.
    - Total Pallets in this Trailer column L gets the pallet count.
    - Everything else in the body is left blank.
    """
    wb = load_embedded_transfer_log()
    ws = wb[LOG_SHEET_NAME] if LOG_SHEET_NAME in wb.sheetnames else wb.active

    clear_log_body(ws)

    needed_rows = len(loads_for_log)
    last_needed_row = LOG_START_ROW + needed_rows - 1

    # Add more rows if the inbound report has more trailers than the template has lines.
    while ws.max_row < last_needed_row:
        insert_at = ws.max_row + 1
        ws.insert_rows(insert_at)
        copy_row_format(ws, insert_at - 1, insert_at)

    for idx, row in loads_for_log.reset_index(drop=True).iterrows():
        excel_row = LOG_START_ROW + idx
        ws.cell(row=excel_row, column=LOG_TRAILER_COL).value = row["Trailer_Last_3"]
        ws.cell(row=excel_row, column=LOG_PALLETS_COL).value = int(row["Pallets"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


if uploaded is not None:
    # Headers span the first 3 rows, so read with no header and skip them.
    df = pd.read_excel(uploaded, header=None, skiprows=HEADER_ROWS)

    # Keep only lines with an LPN. Each LPN counts as one pallet.
    df = df[df[LPN_COL].notna()].copy()

    # Build the trailer number from inbound report columns C:G.
    df["Trailer"] = df.apply(build_trailer, axis=1)
    df = df[df["Trailer"].astype(str).str.strip().ne("")]

    # Each unique LPN = one pallet.
    result = (
        df.groupby("Trailer")[LPN_COL]
        .nunique()
        .reset_index()
        .rename(columns={LPN_COL: "Pallets"})
        .sort_values("Pallets", ascending=False)
        .reset_index(drop=True)
    )

    result["Trailer_Last_3"] = result["Trailer"].apply(last_3_numbers)

    high = result[result["Pallets"] > THRESHOLD].reset_index(drop=True)
    low = result[result["Pallets"] <= THRESHOLD].reset_index(drop=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total trailers", len(result))
    c2.metric(f"Loads over {THRESHOLD}", len(high))
    c3.metric(f"Loads {THRESHOLD} or less", len(low))
    c4.metric("Transfer log rows", len(high))

    # ---- List 1: more than 9 pallets ----
    st.subheader(f"✅ Loads with more than {THRESHOLD} pallets")

    high_display = high[["Trailer", "Trailer_Last_3", "Pallets"]].copy()
    st.dataframe(high_display, use_container_width=True, hide_index=True)

    st.download_button(
        "⬇️ Download full loads (CSV)",
        data=high_display.to_csv(index=False).encode("utf-8"),
        file_name="loads_over_9.csv",
        mime="text/csv",
    )

    # ---- Transfer Log Excel output ----
    st.subheader("🧾 Transfer Trailer Log")

    if high.empty:
        st.info("No non-research loads to place into the transfer log.")
    else:
        st.write(
            "The Transfer Trailer Log below fills only **Trailer Num.** with the "
            "last 3 trailer digits and **Total Pallets in this Trailer** with the "
            "computed pallet count. All other body cells are left blank."
        )

        transfer_log_bytes = build_transfer_log_excel(high)

        st.download_button(
            "⬇️ Download filled Transfer Trailer Log (Excel)",
            data=transfer_log_bytes,
            file_name="filled_transfer_trailer_log.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ---- List 2: 9 or fewer pallets, flagged for research ----
    st.subheader(f"🚨 Loads with {THRESHOLD} or fewer pallets")

    if low.empty:
        st.success("No short loads — nothing to research.")
    else:
        low_display = low[["Trailer", "Trailer_Last_3", "Pallets"]].copy()
        low_display["Status"] = "research"

        styled = low_display.style.apply(flag_red, axis=1).hide(axis="index")
        st.dataframe(styled, use_container_width=True)

        st.download_button(
            "⬇️ Download research list (CSV)",
            data=low_display.to_csv(index=False).encode("utf-8"),
            file_name="loads_to_research.csv",
            mime="text/csv",
        )

else:
    st.info("Waiting for a file...")
