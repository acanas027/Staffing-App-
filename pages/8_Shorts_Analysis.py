import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import Alignment, PatternFill
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepInFrame 
from reportlab.lib.styles import getSampleStyleSheet

# =====================================================================
# PAGE CONFIG
# =====================================================================
st.set_page_config(
    page_title="Shorts Analysis Tool",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =====================================================================
# DESIGN TOKENS + GLOBAL STYLE
# =====================================================================
NAVY = "#12233F"
STEEL = "#3E5C76"
AMBER = "#E8871E"
SUCCESS = "#2E9E5B"
WARNING = "#E8871E"
DANGER = "#D64545"
LIGHT_PURPLE = "#E6D9F2"
BG = "#F5F6F8"
CARD = "#FFFFFF"
TEXT_MUTED = "#6B7280"

STATUS_COLORS = {
    "Full": SUCCESS,
    "Partial": WARNING,
    "Short": DANGER,
}

# Fixed color per wave number so a given wave (e.g. Wave 3 = green) looks the
# same in every chart. Keyed by the wave's integer and cycled for higher waves,
# so the color is stable no matter which waves appear in a particular chart.
WAVE_PALETTE = [STEEL, AMBER, SUCCESS, "#7A5C99", "#2A9D8F", DANGER]


def wave_color_map(wave_values):
    result = {}
    for w in wave_values:
        try:
            n = int(float(w))
        except (TypeError, ValueError):
            continue
        result[str(w)] = WAVE_PALETTE[(n - 1) % len(WAVE_PALETTE)]
    return result

# Distinct pastel palette used to tell different repeated (multi-trailer) items
# apart from each other. Cycled if there are more repeated items than colors.
ITEM_COLOR_PALETTE = [
    "#E6D9F2",  # light purple
    "#D9F2E6",  # light green
    "#F2E6D9",  # light peach
    "#D9E6F2",  # light blue
    "#F2D9E6",  # light pink
    "#F2F2D9",  # light yellow
    "#D9F2F2",  # light cyan
    "#F2D9D9",  # light salmon
]


def inject_dashboard_style():
    """Professional styling — only applied to the results dashboard,
    never to the upload screen, so that one stays plain/default."""
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Sora:wght@500;600;700&family=Inter:wght@400;500;600&display=swap');

    div[data-testid="stAppViewContainer"] div[data-testid="block-container"] h1,
    div[data-testid="stAppViewContainer"] div[data-testid="block-container"] h2,
    div[data-testid="stAppViewContainer"] div[data-testid="block-container"] h3 {{
        font-family: 'Sora', sans-serif;
        color: {NAVY};
    }}

    .dock-header {{
        background: linear-gradient(90deg, {NAVY} 0%, {STEEL} 100%);
        padding: 28px 32px;
        border-radius: 14px;
        margin-bottom: 24px;
        font-family: 'Inter', sans-serif;
    }}
    .dock-header h1 {{
        color: #FFFFFF !important;
        margin: 0;
        font-size: 1.7rem;
        font-family: 'Sora', sans-serif;
    }}
    .dock-header p {{
        color: #D7DEE8;
        margin: 6px 0 0 0;
        font-size: 0.95rem;
    }}

    .kpi-card {{
        background: {CARD};
        border-radius: 10px;
        padding: 10px 14px;
        box-shadow: 0 1px 3px rgba(18,35,63,0.08);
        border-left: 4px solid {AMBER};
        height: 100%;
        font-family: 'Inter', sans-serif;
    }}
    .kpi-label {{
        font-size: 0.72rem;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        color: {TEXT_MUTED};
        font-weight: 600;
        margin-bottom: 2px;
    }}
    .kpi-value {{
        font-family: 'Sora', sans-serif;
        font-size: 1.25rem;
        font-weight: 700;
        color: {NAVY};
    }}
    .kpi-sub {{
        font-size: 0.7rem;
        color: {TEXT_MUTED};
        margin-top: 1px;
    }}

    .section-card {{
        background: {CARD};
        border-radius: 10px;
        padding: 8px 12px;
        box-shadow: 0 1px 3px rgba(18,35,63,0.06);
        margin-bottom: 8px;
    }}
    </style>
    """, unsafe_allow_html=True)


PLOTLY_TEMPLATE = "plotly_white"
FONT = dict(family="Inter, sans-serif", color=NAVY)


def style_fig(fig, height=260):
    fig.update_layout(
        template=PLOTLY_TEMPLATE,
        font=FONT,
        title_font=dict(family="Sora, sans-serif", size=14, color=NAVY),
        margin=dict(l=10, r=10, t=40, b=10),
        height=height,
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    return fig


def kpi_card(label, value, sub=""):
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub">{sub}</div>
    </div>
    """, unsafe_allow_html=True)


STATUS_PDF_COLORS = {
    "Full": colors.HexColor("#C7F0D8"),
    "Partial": colors.HexColor("#FFE8A3"),
    "Short": colors.HexColor("#F5C2C7"),
}


def parse_dispatch_value(v):
    """
    Parse a dispatch time into a comparable whole-number HHMM value.
    The short sheet mixes two formats in the same column:
      - text clock times like "01:00", "02:00", "17:00"  (hours 1-23)
      - plain integers like 800, 1400, 2200
    Both mean the same thing (an HHMM time), so normalize both to an integer:
      "01:00" -> 100, "17:00" -> 1700, 800 -> 800, 2200 -> 2200.
    Anything unparseable returns NaN.
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return np.nan

    import datetime
    # Real time/datetime objects from Excel
    if isinstance(v, datetime.time):
        return float(v.hour * 100 + v.minute)
    if isinstance(v, datetime.datetime):
        return float(v.hour * 100 + v.minute)

    s = str(v).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return np.nan

    # "HH:MM" (or "H:MM") text time
    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 and parts[1] != "" else 0
            return float(h * 100 + m)
        except Exception:
            return np.nan

    # Plain number already in HHMM form (e.g. 800, 1400, 2200)
    try:
        return float(s)
    except Exception:
        return np.nan


def get_multi_trailer_items(df):
    """Return Item values that appear in Load Coverage from more than one real trailer."""
    if not {"Item", "Trailer"}.issubset(df.columns):
        return set()

    tmp = df[["Item", "Trailer"]].copy()
    tmp["Trailer"] = tmp["Trailer"].astype(str).str.strip()
    tmp = tmp[tmp["Trailer"].ne("No trailer found")]

    multi = tmp.groupby("Item")["Trailer"].nunique()
    return set(multi[multi > 1].index.astype(str))


def build_item_color_map(repeated_items):
    """Assign each distinct repeated item its OWN color, so two different
    repeated-item pairs never look like the same group. Deterministic order
    (sorted) so the same item always gets the same color across reruns."""
    items_sorted = sorted(map(str, repeated_items))
    return {
        item: ITEM_COLOR_PALETTE[i % len(ITEM_COLOR_PALETTE)]
        for i, item in enumerate(items_sorted)
    }


def whole_numbers(df, case_cols=(), percent_cols=()):
    """
    Return a copy of df with case_cols rounded to whole numbers (no decimals)
    and percent_cols converted from a 0-1 fraction into a whole-number percent,
    with the column renamed to '<col> %' so the unit is obvious.
    """
    df = df.copy()
    for col in case_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).round(0).astype(int)
    for col in percent_cols:
        if col in df.columns:
            df[col] = (pd.to_numeric(df[col], errors="coerce").fillna(0) * 100).round(0).astype(int)
            df = df.rename(columns={col: f"{col} %"})
    return df


def prettify_headers(df, rename_map=None):
    """
    Display-only header cleanup: apply any specific renames first (e.g. giving
    a column a more descriptive label), then turn every remaining underscore
    into a space so every header reads as plain words. Returns a NEW frame —
    the original (underscore) column names are what the charts/KPI code still
    uses, so this is only ever applied to a separate copy meant for on-screen
    tables, the Excel export, and the PDF.
    """
    df = df.copy()
    if rename_map:
        df = df.rename(columns=rename_map)
    df.columns = [str(c).replace("_", " ") for c in df.columns]
    return df


def build_status_table(df, status_col="Status", repeated_items=None, item_color_map=None):
    """Reportlab Table with Status cells colored and repeated multi-trailer Item
    cells colored per-item (a different color for each distinct repeated item)."""
    repeated_items = set() if repeated_items is None else set(map(str, repeated_items))
    item_color_map = item_color_map or build_item_color_map(repeated_items)

    data = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(data, repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]

    # Per-item highlight on the Item cells where that item is supplied by more
    # than one trailer in the Load Coverage page — each distinct repeated item
    # gets its own color so different repeated pairs are never confused.
    if repeated_items and "Item" in df.columns:
        item_col_idx = list(df.columns).index("Item")
        for i, item_value in enumerate(df["Item"].astype(str).tolist()):
            item_color = item_color_map.get(item_value)
            if item_color:
                style_commands.append((
                    "BACKGROUND",
                    (item_col_idx, i + 1),
                    (item_col_idx, i + 1),
                    colors.HexColor(item_color),
                ))

    if status_col in df.columns:
        col_idx = list(df.columns).index(status_col)
        for i, val in enumerate(df[status_col].astype(str).tolist()):
            color = STATUS_PDF_COLORS.get(val)
            if color:
                style_commands.append(("BACKGROUND", (col_idx, i + 1), (col_idx, i + 1), color))
    table.setStyle(TableStyle(style_commands))
    return table

def build_full_pdf(
    title, onlot_kpis, otr_kpis, figs, onlot_wave_df, otr_wave_df,
    coverage_summary_df, load_df, repeated_items=None, item_color_map=None
):
    """One PDF with separate source KPIs/waves and combined Load Coverage."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=portrait(letter),
        topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30
    )
    styles = getSampleStyleSheet()
    story = []

    # ---- Page 1: Overview ----
    overview_page = []

    overview_page.append(Paragraph(f"{title} — Overview", styles["Title"]))
    overview_page.append(Spacer(1, 14))

    def pdf_kpi_table(kpis):
        header_row = [k[0] for k in kpis]
        value_row = [k[1] for k in kpis]
        sub_row = [k[2] for k in kpis]
        kpi_col_width = min(105, doc.width / max(len(kpis), 1))
        table = Table(
            [header_row, value_row, sub_row],
            colWidths=[kpi_col_width] * len(kpis)
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, 1), 14),
            ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor(NAVY)),
            ("FONTSIZE", (0, 2), (-1, 2), 7),
            ("TEXTCOLOR", (0, 2), (-1, 2), colors.HexColor(TEXT_MUTED)),
            ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(NAVY)),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        return table

    overview_page.append(Paragraph("Topeka Transfers", styles["Heading2"]))
    overview_page.append(pdf_kpi_table(onlot_kpis))
    overview_page.append(Spacer(1, 8))
    overview_page.append(Paragraph("Over-the-Road Transfers", styles["Heading2"]))
    overview_page.append(pdf_kpi_table(otr_kpis))
    overview_page.append(Spacer(1, 12))

    images = []
    for chart_title, fig in figs:
        png_bytes = fig.to_image(format="png", scale=2, width=520, height=316)
        images.append(Image(BytesIO(png_bytes), width=255, height=155))

    rows = []
    for i in range(0, len(images), 2):
        pair = images[i:i + 2]
        if len(pair) == 1:
            pair.append("")
        rows.append(pair)

    if rows:
        chart_table = Table(rows, colWidths=[265, 265])
        chart_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        overview_page.append(chart_table)

    story.append(KeepInFrame(doc.width, doc.height, overview_page, mode="shrink"))

    # ---- Page 2: Separate Wave Plans ----
    story.append(PageBreak())
    story.append(Paragraph(f"{title} — Wave Plan", styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Topeka Transfer Plan — Trailers in Priority Order", styles["Heading2"]))
    story.append(Spacer(1, 6))
    story.append(build_status_table(onlot_wave_df))
    story.append(Spacer(1, 14))
    story.append(Paragraph("OTR Transfer Plan", styles["Heading2"]))
    story.append(Spacer(1, 6))
    story.append(build_status_table(otr_wave_df))

    # ---- Page 3: Load Coverage ----
    story.append(PageBreak())
    story.append(Paragraph(f"{title} — Load Coverage", styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Combined Coverage Summary", styles["Heading2"]))
    story.append(Spacer(1, 6))
    story.append(build_status_table(coverage_summary_df))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Load Coverage — Demand vs. Available by Trip", styles["Heading2"]))
    story.append(Spacer(1, 6))
    story.append(build_status_table(load_df, repeated_items=repeated_items, item_color_map=item_color_map))

    doc.build(story)
    buffer.seek(0)
    return buffer


# =====================================================================
# HEADER
# =====================================================================
st.title("Shorts analysis Tool")
st.write("Turn Topeka and over-the-road transfer inventory plus outbound order shorts into prioritized unloading plans.")

# =====================================================================
# INPUTS
# =====================================================================
st.markdown("### Data Inputs")
col1, col2, col3 = st.columns(3)
with col1:
    book_file = st.file_uploader("Topeka transfer inventory", type=["xlsx"])
with col2:
    otr_file = st.file_uploader("Over-the-road transfers", type=["xlsx"])
with col3:
    short_file = st.file_uploader("Order shorts (short sheet.xlsx)", type=["xlsx"])

run = st.button("Run Analysis")
st.caption(
    "Wave size: 4 trailers per wave. "
    "Priority: earliest dispatch on a short load first; ties broken by cases solved."
)
st.divider()

if not book_file or not otr_file or not short_file:
    st.info("Upload all three files above, then click **Run Analysis**.")
    st.stop()

if not run:
    st.info("Files loaded. Click **Run Analysis** to generate the dashboard.")
    st.stop()

# =====================================================================
# PROCESSING
# =====================================================================
try:
    # ---- LOAD TRAILER INVENTORY (BOOK) ----
    df = pd.read_excel(book_file)
    df = df.iloc[2:].reset_index(drop=True)
    df = df.iloc[:, 4:]

    df.columns = [
        "ColE", "ColF", "ColG", "ColH", "ColI",
        "ColJ", "ColK", "ColL",
        "Date1", "Time1", "Date2", "Time2", "User", "ExtraDate"
    ]

    df["ColE"] = df["ColE"].astype(str).str.replace("L", "", regex=False)
    df["ColE"] = pd.to_numeric(df["ColE"], errors="coerce").fillna(0).astype(int)
    df["ColF"] = pd.to_numeric(df["ColF"], errors="coerce").fillna(0).astype(int)
    df["ColG"] = pd.to_numeric(df["ColG"], errors="coerce").fillna(0).astype(int)
    df["Trailer"] = df["ColE"].astype(str) + df["ColF"].astype(str) + df["ColG"].astype(str)

    def _build_sku(j, k):
        """Build the padded SKU string. The Sku Number column comes in as object
        dtype because a few rows carry alphanumeric SKUs (sample items like
        'S12156'), and the suffix column can be a string of spaces — so coerce
        both to numbers first and skip anything that isn't a real numeric SKU."""
        j = pd.to_numeric(j, errors="coerce")
        if pd.isna(j):
            return None
        k = pd.to_numeric(k, errors="coerce")
        return f"{j:.2f}" + (f"{int(k):03d}" if pd.notna(k) else "000")

    df["SKU"] = [_build_sku(j, k) for j, k in zip(df["ColJ"], df["ColK"])]

    clean_df = df[["Trailer", "SKU", "ColH", "ColL"]].copy()
    clean_df.columns = ["Trailer", "SKU", "LPN", "Quantity"]
    clean_df["Quantity"] = pd.to_numeric(clean_df["Quantity"], errors="coerce")
    clean_df = clean_df.dropna(subset=["SKU", "Quantity"])

    # A "real" transfer trailer carries at least MIN_LPNS_PER_TRAILER pallets
    # (distinct LPNs). Apply this filter to the inventory ITSELF so it flows
    # through everywhere — allocation, dock plan, load coverage, and both KPIs.
    # This keeps the invariant: Trailers Involved can never exceed Transfer
    # Trailers, because the involved trailers are always a subset of this pool.
    MIN_LPNS_PER_TRAILER = 10
    lpns_per_trailer = clean_df.groupby("Trailer")["LPN"].nunique()
    valid_trailers = lpns_per_trailer[lpns_per_trailer >= MIN_LPNS_PER_TRAILER].index
    clean_df = clean_df[clean_df["Trailer"].isin(valid_trailers)].copy()

    # Total transfer trailers on the lot today = trailers left after the filter.
    total_transfer_trailers = int(clean_df["Trailer"].nunique())

    # ---- LOAD OVER-THE-ROAD TRANSFERS ----
    # Only status 66 and 99 are eligible for review. The usable supply column
    # depends on the status:
    #   status 66 (still in transit)      -> In Transit Quantity
    #   status 99 (received / reported)   -> Reported Quantity
    # (Status-99 rows carry 0 in In Transit Quantity, so using In Transit for
    # everything would drop all received transfers.) Whichever column applies,
    # a zero value means that row has no usable supply and must not contribute
    # to coverage, KPIs, priorities, or waves.
    otr_raw = pd.read_excel(otr_file)
    otr_raw.columns = [str(c).strip() for c in otr_raw.columns]

    required_otr_columns = {
        "Item", "Order Status", "In Transit Quantity", "Reported Quantity"
    }
    missing_otr_columns = sorted(required_otr_columns.difference(otr_raw.columns))
    if missing_otr_columns:
        raise ValueError(
            "OTR transfer report is missing required column(s): "
            + ", ".join(missing_otr_columns)
        )

    otr_raw["Order Status"] = pd.to_numeric(otr_raw["Order Status"], errors="coerce")
    otr_raw = otr_raw[otr_raw["Order Status"].isin([66, 99])].copy()

    otr_raw["Item"] = pd.to_numeric(otr_raw["Item"], errors="coerce").map(
        lambda v: f"{v:.5f}" if pd.notna(v) else None
    )
    otr_raw["In Transit Quantity"] = pd.to_numeric(
        otr_raw["In Transit Quantity"], errors="coerce"
    ).fillna(0)
    otr_raw["Reported Quantity"] = pd.to_numeric(
        otr_raw["Reported Quantity"], errors="coerce"
    ).fillna(0)
    # Pick the right supply column per row based on Order Status.
    otr_raw["Quantity"] = np.where(
        otr_raw["Order Status"] == 99,
        otr_raw["Reported Quantity"],
        otr_raw["In Transit Quantity"],
    )

    # The supplied export merges the last two headers into
    # "Delivery External Tracking Number". Some versions split them into
    # "Delivery" and "External Tracking Number", so accept either layout.
    trailer_id_candidates = [
        "Delivery",
        "Delivery External Tracking Number",
        "External Tracking Number",
    ]
    trailer_id_column = next(
        (c for c in trailer_id_candidates if c in otr_raw.columns), None
    )
    if trailer_id_column is None and "Order" not in otr_raw.columns:
        raise ValueError(
            "OTR transfer report needs Delivery, Delivery External Tracking Number, "
            "External Tracking Number, or Order to identify each trailer."
        )

    if trailer_id_column is not None:
        otr_raw["Trailer"] = otr_raw[trailer_id_column].astype(str).str.strip()
    else:
        otr_raw["Trailer"] = ""

    if (
        "External Tracking Number" in otr_raw.columns
        and trailer_id_column != "External Tracking Number"
    ):
        external = otr_raw["External Tracking Number"].astype(str).str.strip()
        blank_delivery = otr_raw["Trailer"].isin(["", "nan", "None"])
        otr_raw.loc[blank_delivery, "Trailer"] = external[blank_delivery]
    if "Order" in otr_raw.columns:
        order_fallback = "Order " + otr_raw["Order"].astype(str).str.strip()
        blank_delivery = otr_raw["Trailer"].isin(["", "nan", "None"])
        otr_raw.loc[blank_delivery, "Trailer"] = order_fallback[blank_delivery]

    otr_clean = otr_raw[["Trailer", "Item", "Quantity", "Order Status"]].copy()
    otr_clean = otr_clean.dropna(subset=["Item", "Quantity"])
    otr_clean = otr_clean[
        (~otr_clean["Trailer"].isin(["", "nan", "None"]))
        & (otr_clean["Quantity"] > 0)
    ].copy()

    # Preserve each eligible trailer's status for the separate OTR KPI and wave plan.
    otr_status = (
        otr_clean.groupby("Trailer")["Order Status"]
        .apply(lambda s: "/".join(str(int(v)) for v in sorted(set(s.dropna()))))
        .rename("Transfer_Status")
    )
    total_otr_trailers = int(otr_clean["Trailer"].nunique())

    # ---- LOAD SHORT SHEET ----
    short_df = pd.read_excel(short_file, header=2)
    short_df = short_df.iloc[:, :11].reset_index(drop=True)

    short_df.columns = [
        "Trip", "Destination", "Dispatch", "Status", "Order",
        "Item", "Description", "Cases", "W", "ProdETA", "Comments"
    ]

    short_clean = short_df[["Trip", "Dispatch", "Item", "Cases"]].copy()

    short_clean["Item"] = pd.to_numeric(short_clean["Item"], errors="coerce").map(
        lambda v: f"{v:.5f}" if pd.notna(v) else None
    )
    short_clean["Cases"] = pd.to_numeric(short_clean["Cases"], errors="coerce")
    # Dispatch may be text "HH:MM" or a plain HHMM integer — normalize both.
    short_clean["Dispatch"] = short_clean["Dispatch"].map(parse_dispatch_value)
    short_clean = short_clean.dropna(subset=["Item", "Cases"])

    # Forward-fill Trip and Dispatch: the short sheet only writes the Trip number
    # and Dispatch time on the first item row of each trip — continuation rows have
    # blank spaces. Strip spaces, treat blank as NaN, then ffill so every item row
    # inherits its parent trip's values.
    short_clean["Trip"] = short_clean["Trip"].astype(str).str.strip().replace("", None)
    short_clean["Trip"] = short_clean["Trip"].ffill()
    short_clean["Dispatch"] = short_clean["Dispatch"].ffill()

    # =====================================================================
    # CORE MODEL (corrected)
    # Every line on the short sheet is a SHORTAGE — cases an outbound load needs
    # that are not on the pick line yet. The cases that fix them live in the
    # transfer trailers (the inventory file). So:
    #   DEMAND  = short-sheet cases
    #   SUPPLY  = trailer inventory
    # We allocate each item's trailer stock to its short-sheet lines in DISPATCH
    # order (earliest dispatch filled first). A trailer's Fix_Cases is the number
    # of cases it actually contributes toward covering shortages — capped at what
    # is demanded. Every trailer that carries a short item therefore fixes cases.
    # =====================================================================

    # Build two explicitly separated supply pools.  On-lot inventory is applied
    # first because it is physically actionable now; OTR inventory then covers
    # the remaining shortage.  Source labels remain attached to every case so
    # the two KPI/wave plans stay separate while Load Coverage is combined.
    onlot_stock = clean_df.groupby(["SKU", "Trailer"], as_index=False)["Quantity"].sum()
    onlot_stock = onlot_stock.rename(columns={"SKU": "Item"})
    onlot_stock["Supply_Source"] = "On-Lot"

    otr_stock = otr_clean.groupby(["Item", "Trailer"], as_index=False)["Quantity"].sum()
    otr_stock["Supply_Source"] = "OTR"

    trailer_stock = pd.concat([onlot_stock, otr_stock], ignore_index=True)
    trailer_stock["Source_Order"] = trailer_stock["Supply_Source"].map(
        {"On-Lot": 0, "OTR": 1}
    )

    # Total inventory available per item across both independent supply pools.
    item_totals = trailer_stock.groupby("Item", as_index=False)["Quantity"].sum()
    item_totals = item_totals.rename(columns={"Quantity": "Total_Item_Inventory"})

    fix_rows = []           # trailer-level: how many cases each trailer contributes to each short line
    load_rows = []          # short-line level: demand, allocated, short, status, primary source trailer
    load_detail_rows = []   # load coverage detail: one row per trailer contribution to a short line

    # Process each item independently. Within an item, fill the earliest-dispatch
    # short line first. Draw from on-lot trailers first, then OTR; within each
    # source use the largest available stock first.
    demand_sorted = short_clean.sort_values(["Item", "Dispatch", "Trip"])

    for item, demand_group in demand_sorted.groupby("Item"):
        stock = trailer_stock[trailer_stock["Item"] == item].sort_values(
            ["Source_Order", "Quantity"], ascending=[True, False]
        )
        stock_remaining = {
            (row["Supply_Source"], row["Trailer"]): float(row["Quantity"])
            for _, row in stock.iterrows()
        }
        total_item_inv = float(item_totals.loc[item_totals["Item"] == item, "Total_Item_Inventory"].sum())

        for _, line in demand_group.iterrows():
            need = float(line["Cases"])
            allocated_total = 0.0
            sources = []

            for source_key in list(stock_remaining.keys()):
                if need - allocated_total <= 0:
                    break
                supply_source, trailer = source_key
                avail = stock_remaining[source_key]
                if avail <= 0:
                    continue
                take = min(avail, need - allocated_total)
                if take <= 0:
                    continue
                stock_remaining[source_key] = avail - take
                allocated_total += take

                # Keep the exact trailer contribution so Load Coverage can show
                # one separate row per trailer used on the same load/item.
                sources.append({
                    "Supply_Source": supply_source,
                    "Trailer": trailer,
                    "Allocated_Cases": take,
                })

                fix_rows.append({
                    "Supply_Source": supply_source,
                    "Trailer": trailer,
                    "Item": item,
                    "Trip": line["Trip"],
                    "Dispatch": line["Dispatch"],
                    "Allocated_Cases": take,
                })

            short_qty = need - allocated_total
            fill_rate = (allocated_total / need) if need > 0 else 0.0
            status = "Full" if fill_rate >= 1 else ("Partial" if fill_rate > 0 else "Short")
            primary_trailer = sources[0]["Trailer"] if sources else np.nan
            primary_source = sources[0]["Supply_Source"] if sources else "None"
            onlot_allocated = sum(
                s["Allocated_Cases"] for s in sources if s["Supply_Source"] == "On-Lot"
            )
            otr_allocated = sum(
                s["Allocated_Cases"] for s in sources if s["Supply_Source"] == "OTR"
            )

            # Line-level summary, kept for KPIs, charts, and load status math.
            load_rows.append({
                "Item": item,
                "Trip": line["Trip"],
                "Dispatch": line["Dispatch"],
                "Demand_Cases": need,
                "Allocated_Cases": allocated_total,
                "On_Lot_Allocated_Cases": onlot_allocated,
                "OTR_Allocated_Cases": otr_allocated,
                "Total_Item_Inventory": total_item_inv,
                "Fill_Rate": fill_rate,
                "Status": status,
                "Actual_short_cases": short_qty,
                "Supply_Source": primary_source,
                "Trailer": primary_trailer,
            })

            # Detail rows for the Load Coverage page/export.
            # If two trailers cover the same short line, this creates two rows.
            # Example: 3 cases from 198 and 9 cases from 605 = two rows.
            if sources:
                for source in sources:
                    load_detail_rows.append({
                        "Supply_Source": source["Supply_Source"],
                        "Item": item,
                        "Trip": line["Trip"],
                        "Dispatch": line["Dispatch"],
                        "Demand_Cases": need,
                        "Allocated_Cases": source["Allocated_Cases"],
                        "Combined_Allocated_Cases": allocated_total,
                        "Total_Item_Inventory": total_item_inv,
                        "Fill_Rate": fill_rate,
                        "Status": status,
                        "Actual_short_cases": short_qty,
                        "Trailer": source["Trailer"],
                    })
            else:
                load_detail_rows.append({
                    "Supply_Source": "None",
                    "Item": item,
                    "Trip": line["Trip"],
                    "Dispatch": line["Dispatch"],
                    "Demand_Cases": need,
                    "Allocated_Cases": 0.0,
                    "Combined_Allocated_Cases": 0.0,
                    "Total_Item_Inventory": total_item_inv,
                    "Fill_Rate": fill_rate,
                    "Status": status,
                    "Actual_short_cases": short_qty,
                    "Trailer": np.nan,
                })

    fix_df = pd.DataFrame(fix_rows, columns=[
        "Supply_Source", "Trailer", "Item", "Trip", "Dispatch", "Allocated_Cases"
    ])
    alloc = pd.DataFrame(load_rows, columns=[
        "Item", "Trip", "Dispatch", "Demand_Cases", "Allocated_Cases",
        "On_Lot_Allocated_Cases", "OTR_Allocated_Cases", "Total_Item_Inventory",
        "Fill_Rate", "Status", "Actual_short_cases", "Supply_Source", "Trailer"
    ])
    alloc_detail = pd.DataFrame(load_detail_rows, columns=[
        "Supply_Source", "Item", "Trip", "Dispatch", "Demand_Cases",
        "Allocated_Cases", "Combined_Allocated_Cases", "Total_Item_Inventory",
        "Fill_Rate", "Status", "Actual_short_cases", "Trailer"
    ])

    if alloc.empty:
        raise ValueError("No valid shortage lines were found in the short sheet.")

    _matched = int(alloc["Allocated_Cases"].gt(0).sum())
    st.success(f"{_matched} of {len(alloc)} short lines can be covered (fully or partially) from trailer inventory.")

    # =====================================================================
    # SOURCE-SPECIFIC TRAILER SUMMARIES AND WAVE PLANS
    # The same ranking logic is run independently for On-Lot and OTR so their
    # KPIs, priorities, and wave numbers never mix.
    # =====================================================================
    def build_source_priority(source_name, status_lookup=None):
        source_fix = fix_df[fix_df["Supply_Source"] == source_name].copy()
        summary_columns = ["Trailer", "Fix_Cases", "Loads_Impacted"]
        priority_columns = summary_columns + [
            "Earliest_Dispatch", "Demand_Served", "SKU_Count",
            "Trailer_Priority", "Wave"
        ]

        if source_fix.empty:
            return (
                pd.DataFrame(columns=summary_columns),
                pd.DataFrame(columns=priority_columns),
            )

        optimized = source_fix.groupby("Trailer").agg(
            Fix_Cases=("Allocated_Cases", "sum"),
            Loads_Impacted=("Trip", "nunique"),
        ).reset_index()

        earliest_dispatch = source_fix.groupby("Trailer")["Dispatch"].min().rename("Earliest_Dispatch")
        demand_served = source_fix.groupby("Trailer")["Allocated_Cases"].sum().rename("Demand_Served")
        sku_count = source_fix.groupby("Trailer")["Item"].nunique().rename("SKU_Count")

        priority = optimized.merge(
            earliest_dispatch, on="Trailer", how="left"
        ).merge(
            demand_served, on="Trailer", how="left"
        ).merge(
            sku_count, on="Trailer", how="left"
        )

        if status_lookup is not None:
            priority = priority.merge(status_lookup, on="Trailer", how="left")

        priority = priority.sort_values(
            by=["Earliest_Dispatch", "Fix_Cases", "Loads_Impacted"],
            ascending=[True, False, False]
        ).reset_index(drop=True)
        priority["Trailer_Priority"] = range(1, len(priority) + 1)
        priority["Wave"] = ((priority["Trailer_Priority"] - 1) // 4) + 1

        optimized = optimized.sort_values(
            by=["Fix_Cases", "Loads_Impacted"], ascending=[False, False]
        ).reset_index(drop=True)
        return optimized, priority

    onlot_optimized_trailers, onlot_trailer_priority = build_source_priority("On-Lot")
    otr_optimized_trailers, otr_trailer_priority = build_source_priority(
        "OTR", otr_status.reset_index()
    )

    onlot_top4_trailers = onlot_optimized_trailers.head(4).copy()
    otr_top4_trailers = otr_optimized_trailers.head(4).copy()
    for top_df in (onlot_top4_trailers, otr_top4_trailers):
        if not top_df.empty:
            top_df.insert(0, "Rank", range(1, len(top_df) + 1))

    # =====================================================================
    # LOAD COVERAGE TABLE
    # =====================================================================
    onlot_priority_lookup = onlot_trailer_priority[
        ["Trailer", "Wave", "Trailer_Priority"]
    ].copy()
    onlot_priority_lookup["Supply_Source"] = "On-Lot"
    otr_priority_lookup = otr_trailer_priority[
        ["Trailer", "Wave", "Trailer_Priority"]
    ].copy()
    otr_priority_lookup["Supply_Source"] = "OTR"
    priority_lookup = pd.concat(
        [onlot_priority_lookup, otr_priority_lookup], ignore_index=True
    )

    load_trailer = alloc_detail.merge(
        priority_lookup,
        on=["Supply_Source", "Trailer"],
        how="left"
    )

    load_trailer = load_trailer[
        ["Supply_Source", "Wave", "Trailer_Priority", "Trailer", "Trip", "Item",
         "Dispatch", "Demand_Cases", "Allocated_Cases", "Combined_Allocated_Cases",
         "Total_Item_Inventory", "Fill_Rate", "Status", "Actual_short_cases"]
    ]

    exceptions = load_trailer[load_trailer["Status"] != "Full"].copy()
    exceptions = exceptions.sort_values(by=["Dispatch", "Status"])

    # ---- FORMAT EXPORTS ----
    # Two independent Wave Plans. Every listed trailer fixes shortage cases.
    def format_dock_plan(priority_df):
        result = priority_df.drop(columns=["Trailer_Priority"]).copy()
        cols = ["Wave"] + [c for c in result.columns if c != "Wave"]
        return result[cols].reset_index(drop=True)

    onlot_dock_plan_export = format_dock_plan(onlot_trailer_priority)
    otr_dock_plan_export = format_dock_plan(otr_trailer_priority)

    load_export = load_trailer.copy()
    load_export["Fill_Rate"] = load_export["Fill_Rate"].round(2)

    # Items that repeat within the same wave should sit next to each other where
    # possible, WITHOUT disturbing rows that have no trailer/wave assigned (those
    # keep their original Dispatch/Trip/Status ordering, exactly as before).
    # Build the cluster key directly instead of assigning an integer array into
    # a pre-created float64 column. Newer pandas versions reject that implicit
    # dtype change with: Invalid value '[...]' for dtype 'float64'.
    load_export["_Item_Cluster_Sort"] = (
        load_export
        .groupby(["Supply_Source", "Wave", "Item"], dropna=False)["Trailer_Priority"]
        .transform("min")
    )
    load_export["_Item_Cluster_Sort"] = pd.to_numeric(
        load_export["_Item_Cluster_Sort"], errors="coerce"
    )

    load_export["_Source_Sort"] = load_export["Supply_Source"].map(
        {"On-Lot": 0, "OTR": 1, "None": 2}
    ).fillna(3)
    load_export["_Wave_Sort"] = load_export["Wave"].fillna(9999)
    load_export["_Trailer_Priority_Sort"] = load_export["Trailer_Priority"].fillna(9999)
    load_export["_No_Value_Sort"] = load_export["Wave"].isna().astype(int)
    load_export = load_export.sort_values(
        by=["_No_Value_Sort", "_Source_Sort", "_Wave_Sort", "_Item_Cluster_Sort", "Item", "_Trailer_Priority_Sort",
            "Dispatch", "Trip", "Status", "Actual_short_cases"],
        ascending=[True, True, True, True, True, True, True, True, True, False]
    ).drop(
        columns=["_Source_Sort", "_Wave_Sort", "_Item_Cluster_Sort", "_Trailer_Priority_Sort", "_No_Value_Sort", "Trailer_Priority"]
    ).reset_index(drop=True)

    load_export["Wave"] = load_export["Wave"].apply(lambda x: "" if pd.isna(x) else int(x))
    load_export["Trailer"] = load_export["Trailer"].fillna("No trailer found")

    # Items that repeat in Load Coverage because they are supplied by different trailers.
    repeated_multi_trailer_items = get_multi_trailer_items(load_export)
    item_color_map = build_item_color_map(repeated_multi_trailer_items)

    exception_export = exceptions.copy()
    exception_export["Fill_Rate"] = exception_export["Fill_Rate"].round(2)
    if "Trailer_Priority" in exception_export.columns:
        exception_export = exception_export.drop(columns=["Trailer_Priority"])
    exception_export["Wave"] = exception_export["Wave"].apply(lambda x: "" if pd.isna(x) else int(x))
    exception_export["Trailer"] = exception_export["Trailer"].fillna("No trailer found")

    # ---- WHOLE NUMBERS: every case/dispatch count as a whole number, and
    # Fill_Rate shown as a whole-number percent (e.g. 22%, not 0.22). ----
    CASE_COLS = [
        "Trip", "Dispatch", "Demand_Cases", "Allocated_Cases",
        "Combined_Allocated_Cases", "Total_Item_Inventory", "Actual_short_cases"
    ]
    load_export = whole_numbers(load_export, case_cols=CASE_COLS, percent_cols=["Fill_Rate"])
    exception_export = whole_numbers(exception_export, case_cols=CASE_COLS, percent_cols=["Fill_Rate"])
    onlot_dock_plan_export = whole_numbers(
        onlot_dock_plan_export,
        case_cols=["Fix_Cases", "Earliest_Dispatch", "Demand_Served"]
    )
    otr_dock_plan_export = whole_numbers(
        otr_dock_plan_export,
        case_cols=["Fix_Cases", "Earliest_Dispatch", "Demand_Served"]
    )
    onlot_optimized_trailers = whole_numbers(onlot_optimized_trailers, case_cols=["Fix_Cases"])
    otr_optimized_trailers = whole_numbers(otr_optimized_trailers, case_cols=["Fix_Cases"])
    onlot_top4_trailers = whole_numbers(onlot_top4_trailers, case_cols=["Fix_Cases"])
    otr_top4_trailers = whole_numbers(otr_top4_trailers, case_cols=["Fix_Cases"])

    # ---- DISPLAY-ONLY HEADER CLEANUP ----
    # Separate copies for tables/exports only — the charts and KPI code above
    # still use the original underscore column names (Fix_Cases, Loads_Impacted,
    # etc.), so renaming happens here, on copies, after everything that does math.
    SHORT_CASES_RENAME = {
        "Actual_short_cases": "Still Short After Both Sources",
        "Allocated_Cases": "This Trailer Solves",
        "Combined_Allocated_Cases": "Combined Cases Solved",
    }
    onlot_dock_plan_display = prettify_headers(onlot_dock_plan_export)
    otr_dock_plan_display = prettify_headers(otr_dock_plan_export)
    load_display = prettify_headers(load_export, rename_map=SHORT_CASES_RENAME)
    exception_display = prettify_headers(exception_export, rename_map=SHORT_CASES_RENAME)
    onlot_optimized_trailers_display = prettify_headers(onlot_optimized_trailers)
    otr_optimized_trailers_display = prettify_headers(otr_optimized_trailers)
    onlot_top4_trailers_display = prettify_headers(onlot_top4_trailers)
    otr_top4_trailers_display = prettify_headers(otr_top4_trailers)
    for display_df in (load_display, exception_display):
        if "Supply Source" in display_df.columns:
            display_df["Supply Source"] = display_df["Supply Source"].replace(
                {"On-Lot": "Topeka Transfers"}
            )

except Exception as e:
    st.error(f"Something went wrong while processing the files: {e}")
    st.stop()

# =====================================================================
# RESULTS DASHBOARD
# =====================================================================
inject_dashboard_style()

st.markdown("""
<div class="dock-header">
    <h1>Shorts Analysis Result</h1>
    <p>Prioritized unloading plan generated from your uploaded files.</p>
</div>
""", unsafe_allow_html=True)

# =====================================================================
# TWO SEPARATE KPI TABLES
# =====================================================================
total_cases_short = int(alloc["Demand_Cases"].sum())
onlot_cases_solved = int(alloc["On_Lot_Allocated_Cases"].sum())
otr_cases_solved = int(alloc["OTR_Allocated_Cases"].sum())
combined_cases_solved = int(alloc["Allocated_Cases"].sum())
total_shortage = int(alloc["Actual_short_cases"].clip(lower=0).sum())

# Short-flow numbers for the KPIs.
# Topeka runs first, so its "Still Short" is total demand minus what Topeka
# solved — and that same remaining number is what OTR starts with. OTR's
# "Still Short" is the final shortage after both sources (total_shortage).
topeka_still_short = max(total_cases_short - onlot_cases_solved, 0)
otr_incoming_short = topeka_still_short

onlot_trailers_involved = int(onlot_dock_plan_export["Trailer"].nunique())
otr_trailers_involved = int(otr_dock_plan_export["Trailer"].nunique())
onlot_waves = int(onlot_dock_plan_export["Wave"].nunique()) if not onlot_dock_plan_export.empty else 0
otr_waves = int(otr_dock_plan_export["Wave"].nunique()) if not otr_dock_plan_export.empty else 0

# ---- LOADS SOLVED (out of total short loads) ----
# Every trip on the short sheet is a "short load", so the denominator is the
# total number of distinct trips. A trip is "solved" when every one of its
# short lines is fully covered:
#   - Topeka solved: every line fully filled from ON-LOT cases alone.
#   - Combined solved (final): every line fully filled after both sources.
trip_status = alloc.groupby("Trip")["Status"].apply(lambda s: (s == "Full").all())
loads_met_count = int(trip_status.sum())
total_loads = int(trip_status.shape[0])

topeka_full_line = alloc["On_Lot_Allocated_Cases"] >= alloc["Demand_Cases"]
topeka_loads_met = int(topeka_full_line.groupby(alloc["Trip"]).all().sum())


def move_next_values(dock_plan, source_label):
    if dock_plan.empty:
        return "—", f"no {source_label} trailer fixes a remaining shortage"
    next_trailer = dock_plan.iloc[0]
    return (
        f"Trailer {next_trailer['Trailer']}",
        f"fixes {int(next_trailer['Fix_Cases']):,} cases "
        f"across {int(next_trailer['Loads_Impacted'])} load(s)",
    )


onlot_move_next_value, onlot_move_next_sub = move_next_values(
    onlot_dock_plan_export, "Topeka"
)
otr_move_next_value, otr_move_next_sub = move_next_values(
    otr_dock_plan_export, "OTR"
)

st.markdown("### Topeka Transfers")
k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
with k1:
    kpi_card("Transfer Trailers", f"{total_transfer_trailers}", "from Topeka")
with k2:
    kpi_card("Trailers Involved", f"{onlot_trailers_involved}", "fix shortages")
with k3:
    kpi_card("Cases Short", f"{total_cases_short:,}", "total demand")
with k4:
    kpi_card("Cases Fixed", f"{onlot_cases_solved:,}", "from Topeka inventory")
with k5:
    kpi_card("Still Short", f"{topeka_still_short:,}", "before OTR")
with k6:
    kpi_card("Loads Solved", f"{topeka_loads_met} of {total_loads}", "fully covered by Topeka")
with k7:
    kpi_card("Move Next", onlot_move_next_value, onlot_move_next_sub)

st.markdown("### Over-the-Road Transfers")
k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
with k1:
    kpi_card("Eligible OTR Trailers", f"{total_otr_trailers}")
with k2:
    kpi_card("OTR Trailers Involved", f"{otr_trailers_involved}", "fix remaining shortages")
with k3:
    kpi_card("Cases Short", f"{otr_incoming_short:,}", "remaining after Topeka")
with k4:
    kpi_card("Cases Fixed", f"{otr_cases_solved:,}", "from OTR inventory")
with k5:
    kpi_card("Still Short", f"{total_shortage:,}", "after both sources")
with k6:
    kpi_card("Loads Solved", f"{loads_met_count} of {total_loads}", "fully covered")
with k7:
    kpi_card("Move Next", otr_move_next_value, otr_move_next_sub)

coverage_summary_display = pd.DataFrame({
    "Metric": [
        "Total Cases Short", "Solved by Topeka", "Solved by OTR",
        "Combined Cases Solved", "Still Short", "Loads Fully Met"
    ],
    "Result": [
        total_cases_short, onlot_cases_solved, otr_cases_solved,
        combined_cases_solved, total_shortage, f"{loads_met_count} of {total_loads}"
    ],
})

# =====================================================================
# TABS
# =====================================================================
tab_overview, tab_wave, tab_load = st.tabs(
    ["Overview", "Wave Plan", "Load Coverage"]
)

# ---------------- OVERVIEW ----------------
with tab_overview:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        topeka_cases_chart_df = onlot_dock_plan_export.sort_values(
            "Fix_Cases", ascending=False
        ).head(15).copy()
        topeka_cases_chart_df["Wave"] = topeka_cases_chart_df["Wave"].astype(str)
        fig1 = px.bar(
            topeka_cases_chart_df,
            x="Trailer", y="Fix_Cases", color="Wave",
            title="Cases Fixed by Trailer — Topeka Transfers", text="Fix_Cases",
            color_discrete_sequence=[STEEL, AMBER, SUCCESS]
        )
        st.plotly_chart(style_fig(fig1), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        otr_cases_chart_df = otr_dock_plan_export.sort_values(
            "Fix_Cases", ascending=False
        ).head(15).copy()
        otr_cases_chart_df["Wave"] = otr_cases_chart_df["Wave"].astype(str)
        fig5 = px.bar(
            otr_cases_chart_df,
            x="Trailer", y="Fix_Cases", color="Wave",
            title="Cases Fixed by Trailer — Over-the-Road", text="Fix_Cases",
            color_discrete_sequence=[STEEL, AMBER]
        )
        st.plotly_chart(style_fig(fig5), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    combined_priority_chart_df = pd.concat([
        onlot_trailer_priority.assign(Supply_Source="Topeka Transfers", Source_Order=0),
        otr_trailer_priority.assign(Supply_Source="OTR", Source_Order=1),
    ], ignore_index=True)
    combined_priority_chart_df = combined_priority_chart_df.sort_values(
        ["Wave", "Source_Order", "Trailer_Priority"]
    )
    combined_priority_chart_df["Display_Trailer"] = (
        combined_priority_chart_df["Trailer"].astype(str)
    )
    duplicate_trailer_labels = combined_priority_chart_df["Display_Trailer"].duplicated(
        keep=False
    )
    combined_priority_chart_df.loc[duplicate_trailer_labels, "Display_Trailer"] = (
        combined_priority_chart_df.loc[duplicate_trailer_labels, "Supply_Source"]
        + " | "
        + combined_priority_chart_df.loc[duplicate_trailer_labels, "Trailer"].astype(str)
    )
    combined_priority_chart_df["Wave"] = (
        combined_priority_chart_df["Wave"].astype(int).astype(str)
    )
    # Plotly Express already handles the visual direction for horizontal bars.
    # Pass priority order directly so #1 renders at the top; manually reversing
    # this list causes the last priority to appear first.
    priority_category_order = combined_priority_chart_df["Display_Trailer"].tolist()
    fig2 = px.bar(
        combined_priority_chart_df,
        x="Fix_Cases", y="Display_Trailer", orientation="h",
        color="Wave",
        title="Trailer Priority Order — By Wave",
        category_orders={"Display_Trailer": priority_category_order},
        color_discrete_sequence=[STEEL, AMBER, SUCCESS, "#7A5C99", "#2A9D8F", DANGER],
        hover_data={
            "Supply_Source": True, "Trailer_Priority": True, "Trailer": True,
            "Display_Trailer": False, "Source_Order": False,
        }
    )
    fig2 = style_fig(fig2, height=320)
    fig2.update_layout(
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.22,
            xanchor="center",
            x=0.5,
            title_text="Wave",
        ),
        margin=dict(l=10, r=10, t=50, b=75),
    )
    st.plotly_chart(fig2, use_container_width=True)
    st.caption(
        "Trailers are ordered and color-coded by wave. Source and exact trailer "
        "priority remain available in the hover details."
    )
    st.markdown('</div>', unsafe_allow_html=True)

    c3, c4 = st.columns(2)
    with c3:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        status_summary = alloc.groupby("Status").size().reset_index(name="Count")
        fig3 = px.pie(
            status_summary, names="Status", values="Count",
            title="Combined Load Status Breakdown", color="Status",
            color_discrete_map=STATUS_COLORS, hole=0.55
        )
        st.plotly_chart(style_fig(fig3), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with c4:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        short_items = alloc[alloc["Actual_short_cases"] > 0]
        top_skus = short_items.groupby("Item").agg(
            Actual_short_cases=("Actual_short_cases", "sum")
        ).reset_index()
        top_skus = top_skus.sort_values(by="Actual_short_cases", ascending=False).head(10)
        fig4 = px.bar(
            top_skus, x="Item", y="Actual_short_cases", title="Top Still-Short Items After Both Sources",
            text="Actual_short_cases", color_discrete_sequence=[DANGER]
        )
        st.plotly_chart(style_fig(fig4), use_container_width=True)
        st.caption("Items still short after Topeka and OTR inventory are allocated.")
        st.markdown('</div>', unsafe_allow_html=True)

    # The two cases-fixed graphs and the single priority graph above intentionally
    # replace the previous combined cases-fixed view.

# ---------------- WAVE PLAN ----------------
with tab_wave:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Topeka Transfer Plan — Trailers in Priority Order")
    st.caption(
        "Every Topeka transfer trailer here carries cases that fix a shortage. "
        "Priority: earliest dispatch on a short load first; ties broken by most cases fixed. "
        "Waves are groups of 4."
    )
    st.dataframe(onlot_dock_plan_display, use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("OTR Transfer Plan — Trailers in Priority Order")
    st.caption(
        "Only OTR transfers on status 66/99 with usable supply greater than zero are included "
        "(In Transit Quantity for status 66, Reported Quantity for status 99). "
        "The OTR waves are ranked independently and contain 4 trailers each."
    )
    st.dataframe(otr_dock_plan_display, use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ---------------- LOAD COVERAGE ----------------
with tab_load:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Load Coverage — Demand vs. Available by Trip")
    st.caption(
        "This is the only combined analysis: Topeka inventory is allocated first, "
        "then eligible OTR inventory covers the remaining shortage. "
        "Each row is one source/trailer contribution to one load/item. "
        "Status: green = Full, yellow = Partial, red = Short. "
        "Each repeated (multi-trailer) item gets its own highlight color."
    )
    st.dataframe(coverage_summary_display, use_container_width=True, hide_index=True)
    st.caption(
        "Combined Cases Solved equals Topeka Cases Solved plus OTR Cases Solved; "
        "Still Short is calculated once per short line, never once per trailer row."
    )

    def highlight_load_coverage(row):
        styles = [""] * len(row)
        columns = list(row.index)

        if "Status" in columns:
            status_idx = columns.index("Status")
            color_map = {"Full": "#C7F0D8", "Partial": "#FFE8A3", "Short": "#F5C2C7"}
            status_color = color_map.get(row["Status"], "")
            if status_color:
                styles[status_idx] = f"background-color: {status_color}"

        if "Item" in columns:
            item_value = str(row["Item"])
            item_color = item_color_map.get(item_value)
            if item_color:
                item_idx = columns.index("Item")
                styles[item_idx] = f"background-color: {item_color}; font-weight: 700"

        return styles

    styled_load = load_display.style.apply(highlight_load_coverage, axis=1)
    st.dataframe(styled_load, use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)


# =====================================================================
# EXPORT
# =====================================================================
def build_excel():
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        coverage_summary_display.to_excel(writer, sheet_name="Coverage Summary", index=False)
        onlot_dock_plan_display.to_excel(writer, sheet_name="Topeka Wave Plan", index=False)
        otr_dock_plan_display.to_excel(writer, sheet_name="OTR Wave Plan", index=False)
        load_display.to_excel(writer, sheet_name="Load Coverage", index=False)
        exception_display.to_excel(writer, sheet_name="Exception Report", index=False)
        onlot_optimized_trailers_display.to_excel(writer, sheet_name="Topeka Optimized", index=False)
        otr_optimized_trailers_display.to_excel(writer, sheet_name="OTR Optimized", index=False)
        onlot_top4_trailers_display.to_excel(writer, sheet_name="Topeka Top 4", index=False)
        otr_top4_trailers_display.to_excel(writer, sheet_name="OTR Top 4", index=False)

        for sheet in writer.sheets:
            ws = writer.sheets[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Per-item highlight on repeated Item cells in the Load Coverage sheet —
        # each distinct repeated item gets its own fill color.
        ws = writer.sheets.get("Load Coverage")
        if ws is not None:
            headers = [cell.value for cell in ws[1]]
            if "Item" in headers:
                item_col = headers.index("Item") + 1
                fills_by_item = {
                    item: PatternFill(start_color=color.lstrip("#"), end_color=color.lstrip("#"), fill_type="solid")
                    for item, color in item_color_map.items()
                }
                for row_num in range(2, ws.max_row + 1):
                    item_value = str(ws.cell(row=row_num, column=item_col).value)
                    fill = fills_by_item.get(item_value)
                    if fill:
                        ws.cell(row=row_num, column=item_col).fill = fill
    buffer.seek(0)
    return buffer

st.markdown("### Export")
c1, c2 = st.columns(2)
with c1:
    st.download_button(
        label="Download Full Report (.xlsx)",
        data=build_excel(),
        file_name="Short_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
with c2:
    st.download_button(
        label="Download Full Report (.pdf)",
        data=build_full_pdf(
            "Shorts Analysis Result",
            onlot_kpis=[
                ("Transfer Trailers", str(total_transfer_trailers), "on lot"),
                ("Involved", str(onlot_trailers_involved), "fix shortages"),
                ("Cases Short", f"{total_cases_short:,}", "total demand"),
                ("Cases Fixed", f"{onlot_cases_solved:,}", "Topeka supply"),
                ("Still Short", f"{topeka_still_short:,}", "before OTR"),
                ("Loads Solved", f"{topeka_loads_met} of {total_loads}", "by Topeka"),
                ("Move Next", onlot_move_next_value, onlot_move_next_sub),
            ],
            otr_kpis=[
                ("Eligible OTR", str(total_otr_trailers), ""),
                ("Involved", str(otr_trailers_involved), "fix shortages"),
                ("Cases Short", f"{otr_incoming_short:,}", "after Topeka"),
                ("Cases Fixed", f"{otr_cases_solved:,}", "OTR supply"),
                ("Still Short", f"{total_shortage:,}", "after both"),
                ("Loads Solved", f"{loads_met_count} of {total_loads}", "Fully Covered"),
                ("Move Next", otr_move_next_value, otr_move_next_sub),
            ],
            figs=[
                ("Topeka Cases Fixed", fig1),
                ("OTR Cases Fixed", fig5),
                ("Combined Priority Order", fig2),
                ("Combined Load Status", fig3),
                ("Top Still-Short Items", fig4),
            ],
            onlot_wave_df=onlot_dock_plan_display,
            otr_wave_df=otr_dock_plan_display,
            coverage_summary_df=coverage_summary_display,
            load_df=load_display,
            repeated_items=repeated_multi_trailer_items,
            item_color_map=item_color_map
        ),
        file_name="Shorts_Analysis_Report.pdf",
        mime="application/pdf"
    )
